from flask import Flask, render_template, request, jsonify, Response, session, redirect, g
from flask_cors import CORS
from database import get_db, init_db
from datetime import datetime, date
import json, os, shutil, sqlite3, re, csv, io, math, requests as http_req

try:
    import yfinance as yf
    HAS_YFINANCE = True
except ImportError:
    HAS_YFINANCE = False

try:
    from apscheduler.schedulers.background import BackgroundScheduler
    from apscheduler.triggers.cron import CronTrigger
    from apscheduler.triggers.interval import IntervalTrigger
    HAS_SCHEDULER = True
except ImportError:
    HAS_SCHEDULER = False

try:
    from pykrx import stock as krx_stock
    HAS_PYKRX = True
except ImportError:
    HAS_PYKRX = False

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import xlrd
    HAS_XLRD = True
except ImportError:
    HAS_XLRD = False

app = Flask(__name__)
CORS(app)
app.secret_key = os.environ.get('FLASK_SECRET_KEY', 'richdadtechtree-money-secret-key-1029384756!')
GOOGLE_CLIENT_ID = os.environ.get('GOOGLE_CLIENT_ID', '881691238914-testlocalgoogleclientid.apps.googleusercontent.com')

from flask_caching import Cache
cache = Cache(app, config={
    'CACHE_TYPE': 'SimpleCache',
    'CACHE_DEFAULT_TIMEOUT': 120  # 기본 2분
})

# gunicorn 등 production 서버에서도 DB 초기화/마이그레이션 실행
with app.app_context():
    try:
        init_db()
    except Exception as _init_err:
        print(f"[init_db] {_init_err}")

def _fix_recurring_day_of_month():
    """
    recurring_budget.day_of_month=1 인 템플릿의 day_of_month 를 실제 날짜로 수정.

    전략 1: recurring_id 로 직접 연결된 과거 항목 중 1일이 아닌 날짜가 있으면 사용
    전략 2: 이름(trim·소문자) 이 같은 '고정지출' 항목 중 가장 최근 비-1일 날짜를 사용
    이후: auto-generated 항목 날짜도 템플릿 day_of_month 에 맞게 수정
    """
    try:
        db = get_db()
        cur = db.cursor()

        # ── 전략 1: recurring_id 직접 연결 ──
        cur.execute("""
            UPDATE recurring_budget rb
            SET day_of_month = EXTRACT(DAY FROM b.date::date)::INTEGER
            FROM (
                SELECT DISTINCT ON (recurring_id) recurring_id, date
                FROM budget
                WHERE recurring_id IS NOT NULL
                  AND EXTRACT(DAY FROM date::date) <> 1
                ORDER BY recurring_id, date DESC
            ) b
            WHERE rb.id = b.recurring_id
              AND rb.day_of_month = 1
        """)
        updated_rc1 = cur.rowcount

        # ── 전략 2: 이름 매칭 (고정지출 유형, 1일이 아닌 항목) ──
        cur.execute("""
            UPDATE recurring_budget rb
            SET day_of_month = EXTRACT(DAY FROM b.date::date)::INTEGER
            FROM (
                SELECT DISTINCT ON (TRIM(LOWER(name)))
                    TRIM(LOWER(name)) AS norm_name, date
                FROM budget
                WHERE type = '고정지출'
                  AND EXTRACT(DAY FROM date::date) <> 1
                ORDER BY TRIM(LOWER(name)), date DESC
            ) b
            WHERE TRIM(LOWER(rb.name)) = b.norm_name
              AND rb.day_of_month = 1
              AND rb.is_active = TRUE
        """)
        updated_rc2 = cur.rowcount

        # ── auto-generated 항목 날짜 수정 ──
        cur.execute("""
            UPDATE budget b
            SET date = (
                DATE_TRUNC('month', b.date::date) +
                MAKE_INTERVAL(days => LEAST(
                    rb.day_of_month,
                    EXTRACT(DAY FROM (DATE_TRUNC('month', b.date::date)
                                      + INTERVAL '1 month - 1 day'))::INTEGER
                ) - 1)
            )::date
            FROM recurring_budget rb
            WHERE b.recurring_id = rb.id
              AND b.is_auto_generated = TRUE
              AND EXTRACT(DAY FROM b.date::date) <> rb.day_of_month
        """)
        updated_b = cur.rowcount

        db.commit()
        cur.close()
        db.close()
        total_rc = updated_rc1 + updated_rc2
        if total_rc or updated_b:
            print(f"[fix_recurring] 템플릿 {total_rc}건 day_of_month 수정 "
                  f"(직접연결 {updated_rc1}, 이름매칭 {updated_rc2}), "
                  f"항목 {updated_b}건 날짜 수정")
    except Exception as e:
        import traceback
        print(f"[fix_recurring] {e}\n{traceback.format_exc()}")

with app.app_context():
    try:
        _fix_recurring_day_of_month()
    except Exception as _fix_err:
        print(f"[fix_recurring] {_fix_err}")

def _scheduled_snapshot():
    """매일 23:59:30 — 오늘 순자산을 daily_snapshots에 저장"""
    try:
        with app.app_context():
            db = get_db()
            _save_daily_snapshot(db)
            db.commit()
            db.close()
            print("[scheduler] 일일 스냅샷 저장 완료")
    except Exception as e:
        print(f"[scheduler] 스냅샷 저장 실패: {e}")

def _keep_alive():
    """10분마다 자기 자신에 HTTP 요청 — Render 슬립 방지"""
    try:
        base = os.environ.get('RENDER_EXTERNAL_URL', '')
        if base:
            http_req.get(f"{base}/api/ping", timeout=10)
            print("[scheduler] keep-alive ping 완료")
    except Exception as e:
        print(f"[scheduler] keep-alive 실패: {e}")

def _scheduled_price_update():
    """매일 09:00, 16:00 KST — 주식/ETF/코인 현재가 자동 업데이트"""
    try:
        with app.app_context():
            _run_price_update_logic()
            print("[scheduler] 가격 업데이트 완료")
    except Exception as e:
        print(f"[scheduler] 가격 업데이트 실패: {e}")

if HAS_SCHEDULER:
    _scheduler = BackgroundScheduler(timezone='Asia/Seoul')
    _scheduler.add_job(_scheduled_snapshot, CronTrigger(hour=23, minute=59, second=30))
    _scheduler.add_job(_scheduled_price_update, CronTrigger(hour=9, minute=0, second=0))
    _scheduler.add_job(_scheduled_price_update, CronTrigger(hour=16, minute=0, second=0))
    _scheduler.add_job(_keep_alive, IntervalTrigger(minutes=10))
    _scheduler.start()
    print("[scheduler] 시작: 매일 09:00/16:00 가격 업데이트 + 23:59:30 스냅샷 + 10분 keep-alive")

def _clear_summary_cache():
    """수입/지출/자산 데이터 변경 시 모든 캐시 삭제"""
    try:
        cache.clear()
    except Exception as e:
        print("Error clearing cache:", e)

@app.after_request
def clear_cache_on_modification(response):
    if request.method in ('POST', 'PUT', 'DELETE'):
        if 200 <= response.status_code < 400:
            _clear_summary_cache()
    return response

@app.teardown_appcontext
def close_db_connection(exception=None):
    from flask import g
    db_conn = getattr(g, 'db_conn', None)
    if db_conn is not None:
        try:
            from database import _pool
            if _pool:
                _pool.putconn(db_conn)
            else:
                db_conn.close()
        except Exception as e:
            print("Error returning connection to pool on teardown:", e)

def _retry_on_db_error(fn):
    """DB 연결 끊김 시 연결 초기화 후 1회 재시도하는 래퍼"""
    import functools, psycopg2
    @functools.wraps(fn)
    def wrapper(*args, **kwargs):
        for attempt in range(2):
            try:
                if attempt > 0 and hasattr(g, 'db_conn'):
                    try:
                        from database import _pool
                        if _pool:
                            _pool.putconn(g.db_conn, close=True)
                    except Exception:
                        pass
                    delattr(g, 'db_conn')
                return fn(*args, **kwargs)
            except (psycopg2.DatabaseError, psycopg2.OperationalError):
                if attempt == 0:
                    continue
                raise
    return wrapper

from flask.json.provider import DefaultJSONProvider
import math as _math
class CustomJSONProvider(DefaultJSONProvider):
    def default(self, obj):
        import decimal
        if isinstance(obj, (date, datetime)):
            return obj.isoformat()
        if isinstance(obj, decimal.Decimal):
            return float(obj)
        if isinstance(obj, float) and (_math.isnan(obj) or _math.isinf(obj)):
            return None
        return super().default(obj)

    def dumps(self, obj, **kwargs):
        import decimal
        # float NaN/Infinity → null, Decimal → float 치환 후 직렬화
        def _sanitize(o):
            if isinstance(o, decimal.Decimal):
                return float(o)
            if isinstance(o, float) and (_math.isnan(o) or _math.isinf(o)):
                return None
            if isinstance(o, dict):
                return {k: _sanitize(v) for k, v in o.items()}
            if isinstance(o, list):
                return [_sanitize(v) for v in o]
            return o
        return super().dumps(_sanitize(obj), **kwargs)

app.json = CustomJSONProvider(app)

# ── 버전 정보 ────────────────────────────────────────────────
def _auto_increment_version():
    version_file = os.path.join(os.path.dirname(__file__), 'version.json')
    try:
        data = {"version": "1.00", "updated": ""}
        if os.path.exists(version_file):
            with open(version_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
        try:
            v_num = float(data.get("version", "1.00"))
            v_num = round(v_num + 0.01, 2)
            data["version"] = f"{v_num:.2f}"
        except ValueError:
            data["version"] = "1.00"
        from datetime import date
        data["updated"] = str(date.today())
        with open(version_file, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
    except Exception as e:
        print("Version increment failed:", e)

_auto_increment_version()

def _load_version():
    try:
        with open(os.path.join(os.path.dirname(__file__), 'version.json'), 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return {"version": "1.00", "updated": ""}

@app.context_processor
def inject_version():
    v = _load_version()
    return {"APP_VERSION": v.get("version", "1.00"), "APP_UPDATED": v.get("updated", "")}

# ── 인증 미들웨어 및 구글 로그인 라우터 ──────────────────────────────────
@app.before_request
def enforce_auth():
    # 1. 로그인 필요 없는 예외 경로 (로그인 페이지, 구글 로그인 API, 정적 에셋 등)
    if request.path in ['/login', '/api/auth/google'] or request.path.startswith('/static'):
        return None
    
    # 2. 세션에 로그인 사용자 정보가 없으면 로그인 페이지로 리디렉션
    if 'user' not in session:
        return redirect('/login')

@app.route('/login')
def login_page():
    if 'user' in session:
        return redirect('/')
    return render_template('login.html', google_client_id=GOOGLE_CLIENT_ID)

@app.route('/logout')
def logout():
    session.clear()
    return redirect('/login')

ALLOWED_EMAILS = [
    'bbonoyo@gmail.com',
    'mybpilatesmyb@gmail.com'
]

@app.route('/api/auth/google', methods=['POST'])
def api_auth_google():
    d = request.json or {}
    id_token = d.get('credential')
    if not id_token:
        return jsonify({'ok': False, 'error': '인증 토큰이 제공되지 않았습니다.'}), 400
    
    # 구글 서버를 통해 id_token 검증 (Zero dependency)
    try:
        res = http_req.get(f"https://oauth2.googleapis.com/tokeninfo?id_token={id_token}", timeout=5)
        if res.status_code != 200:
            return jsonify({'ok': False, 'error': '유효하지 않은 구글 인증 토큰입니다.'}), 401
        
        payload = res.json()
        email = payload.get('email', '').strip().lower()
        
        if email not in [addr.lower() for addr in ALLOWED_EMAILS]:
            return jsonify({'ok': False, 'error': '승인되지 않은 구글 계정입니다. 접근 권한이 없습니다.'}), 403
        
        session['user'] = {
            'email': payload.get('email'),
            'name': payload.get('name'),
            'picture': payload.get('picture')
        }
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': f'구글 서버 인증 통신 중 오류: {str(e)}'}), 500

# ── 페이지 라우터 ────────────────────────────────────────────
@app.route('/api/ping')
def api_ping():
    return jsonify({'ok': True})

@app.route('/')
def dashboard():
    return render_template('dashboard.html')

@app.route('/income')
def income():
    return render_template('income.html')

@app.route('/budget')
def budget():
    return render_template('budget.html')

@app.route('/cards')
def cards():
    return render_template('cards.html')

@app.route('/investments')
def investments():
    return render_template('investments.html')

@app.route('/realestate')
def realestate():
    return render_template('realestate.html')

@app.route('/loans')
def loans():
    return render_template('loans.html')

@app.route('/pension')
def pension():
    return render_template('pension.html')

@app.route('/cash')
def cash():
    return render_template('cash.html')

@app.route('/goals')
def goals():
    return render_template('goals.html')

@app.route('/monthly')
def monthly():
    return render_template('monthly.html')

@app.route('/tech-tree')
def tech_tree():
    return render_template('tech_tree.html')

@app.route('/analysis/calculator')
def analysis_calculator():
    return render_template('analysis_calculator.html')


# ── 공통 헬퍼 ────────────────────────────────────────────────
def _month_range(year_str, month_str):
    """'YYYY', 'MM' -> ('YYYY-MM-01', 'YYYY-MM-01' (next month))"""
    y, m = int(year_str), int(month_str)
    start = f"{y:04d}-{m:02d}-01"
    if m == 12:
        end = f"{y+1:04d}-01-01"
    else:
        end = f"{y:04d}-{m+1:02d}-01"
    return start, end

def rows_to_list(rows):
    import decimal
    res = []
    for r in rows:
        d = dict(r)
        for k, v in d.items():
            if isinstance(v, (date, datetime)):
                d[k] = v.isoformat()
            elif isinstance(v, decimal.Decimal):
                d[k] = float(v)
        res.append(d)
    return res


# ── API: 수입 ────────────────────────────────────────────────
@app.route('/api/income', methods=['GET', 'POST'])
def api_income():
    db = get_db()
    if request.method == 'GET':
        year  = request.args.get('year')
        month = request.args.get('month')
        query = "SELECT * FROM income"
        params = []
        if year and month:
            query += " WHERE date >= %s AND date < %s"
            start_date, end_date = _month_range(year, month)
            params = [start_date, end_date]
        query += " ORDER BY date DESC"
        cur = db.cursor()
        cur.execute(query, params)
        rows = cur.fetchall()
        cur.close()
        db.close()
        return jsonify(rows_to_list(rows))

    data = request.json
    base_date_str = data['date']          # 'YYYY-MM-DD'
    is_recurring  = data.get('is_recurring', False)
    repeat_months = int(data.get('repeat_months') or 1)

    currency = data.get('currency', 'KRW')
    exchange_rate = float(data.get('exchange_rate') or 1.0)
    original_amount = float(data.get('original_amount') or 0.0)

    if currency == 'USD':
        amount = int(original_amount * exchange_rate)
    else:
        amount = int(data.get('amount') or 0)
        original_amount = float(amount)
        exchange_rate = 1.0

    if is_recurring and repeat_months > 1:
        # base_date 에서 repeat_months 개월치를 순서대로 INSERT
        from datetime import date as _date
        import calendar as _cal

        base_date = _date.fromisoformat(base_date_str)
        for i in range(repeat_months):
            y = base_date.year + (base_date.month - 1 + i) // 12
            m = (base_date.month - 1 + i) % 12 + 1
            # 말일 초과 보정 (예: 1월31일 → 2월28일)
            d = min(base_date.day, _cal.monthrange(y, m)[1])
            tx_date = _date(y, m, d).isoformat()
            cur = db.cursor()
            cur.execute(
            "INSERT INTO income (date, category, name, memo, amount, currency, exchange_rate, original_amount) VALUES (%s,%s,%s,%s,%s,%s,%s,%s)",
            (tx_date, data.get('category'), data.get('name'), data.get('memo'), amount, currency, exchange_rate, original_amount)
            )
            cur.close()
    else:
        cur = db.cursor()
        cur.execute(
        "INSERT INTO income (date, category, name, memo, amount, currency, exchange_rate, original_amount) VALUES (%s,%s,%s,%s,%s,%s,%s,%s)",
        (base_date_str, data.get('category'), data.get('name'), data.get('memo'), amount, currency, exchange_rate, original_amount)
        )
        cur.close()

    db.commit()
    db.close()
    return jsonify({'ok': True}), 201


@app.route('/api/income/<int:rid>', methods=['PUT', 'DELETE'])
def api_income_detail(rid):
    db = get_db()
    if request.method == 'PUT':
        data = request.json
        currency = data.get('currency', 'KRW')
        exchange_rate = float(data.get('exchange_rate') or 1.0)
        original_amount = float(data.get('original_amount') or 0.0)

        if currency == 'USD':
            amount = int(original_amount * exchange_rate)
        else:
            amount = int(data.get('amount') or 0)
            original_amount = float(amount)
            exchange_rate = 1.0

        cur = db.cursor()
        cur.execute(
        "UPDATE income SET date=%s, category=%s, name=%s, memo=%s, amount=%s, currency=%s, exchange_rate=%s, original_amount=%s WHERE id=%s",
        (data.get('date'), data.get('category'), data.get('name'),
        data.get('memo'), amount, currency, exchange_rate, original_amount, rid)
        )
        cur.close()
        db.commit()
        db.close()
        return jsonify({'ok': True})
    cur = db.cursor()
    cur.execute("DELETE FROM income WHERE id = %s", (rid,))
    cur.close()
    db.commit()
    db.close()
    return jsonify({'ok': True})


# ── API: 가계부 ──────────────────────────────────────────────
import calendar as _cal

def _generate_recurring_budget(db, year: int, month: int):
    """
    지정한 연월에 대해 활성 템플릿을 조회하고,
    아직 생성되지 않은 고정지출 항목을 budget 테이블에 INSERT.

    규칙:
      1. 해당 월이 start_month~end_month 범위 내여야 함
      2. budget에 같은 (recurring_id, 해당월) 조합이 없어야 함 (중복 방지)
      3. day_of_month가 해당 월의 말일을 초과하면 말일로 보정
      4. 생성된 날짜가 오늘 이전인 경우에만 INSERT
         (해당 날짜가 오지 않은 건 생성하지 않음)
    """
    ym_str  = f"{year}-{month:02d}"
    today   = date.today()

    # 이미 생성된 항목 중 날짜가 오늘 이후인 것은 삭제 (잘못 선생성된 경우 정리)
    cur = db.cursor()
    cur.execute("""
        DELETE FROM budget
        WHERE is_auto_generated = TRUE
          AND to_char(date::date, 'YYYY-MM') = %s
          AND date::date > %s
    """, (ym_str, today))
    cur.close()

    cur = db.cursor()
    cur.execute("""
        SELECT * FROM recurring_budget
        WHERE is_active = TRUE
          AND start_month <= %s
          AND (end_month IS NULL OR end_month >= %s)
    """, (ym_str, ym_str))
    templates = cur.fetchall()
    cur.close()

    inserted = 0
    for t in templates:
        # 실제 날짜 계산 (말일 초과 보정)
        max_day    = _cal.monthrange(year, month)[1]
        actual_day = min(t['day_of_month'], max_day)
        tx_date    = date(year, month, actual_day)

        # ── 핵심 조건: 해당 날짜가 오늘 이후면 생성하지 않음 ──
        if tx_date > today:
            continue

        # 이미 이 템플릿으로 해당 월에 생성된 항목이 있는지 확인
        cur = db.cursor()
        cur.execute("""
            SELECT id, date::date as entry_date FROM budget
            WHERE recurring_id = %s
              AND to_char(date::date, 'YYYY-MM') = %s
        """, (t['id'], ym_str))
        existing = cur.fetchone()
        cur.close()

        if existing:
            # 날짜가 올바른 날짜와 다르면 수정 (예: 1일로 잘못 생성된 경우)
            if existing['entry_date'] != tx_date:
                cur = db.cursor()
                cur.execute("UPDATE budget SET date=%s WHERE id=%s",
                            (tx_date.isoformat(), existing['id']))
                cur.close()
                inserted += 1  # 수정도 카운트
            continue  # 이미 생성됨 → 신규 INSERT 스킵

        # 새 budget 행 INSERT
        cur = db.cursor()
        cur.execute("""
            INSERT INTO budget
                (date, category, name, type, payment_method,
                 amount, memo, card_id, recurring_id, is_auto_generated)
            VALUES (%s, %s, %s, '지출', %s, %s, %s, %s, %s, TRUE)
            RETURNING id
        """, (tx_date.isoformat(), t['category'], t['name'],
              t['payment_method'], t['amount'], t['memo'],
              t['card_id'], t['id']))
        budget_id = cur.fetchone()[0]
        cur.close()

        # 카드 연동 처리 (기존 _sync_card_tx 활용)
        if t['card_id']:
            _sync_card_tx(db, budget_id, {
                'card_id':  t['card_id'],
                'date':     tx_date.isoformat(),
                'name':     t['name'],
                'category': t['category'],
                'amount':   t['amount'],
                'memo':     t['memo'],
            })

        inserted += 1

    if inserted > 0:
        db.commit()
    return inserted


def _sync_card_tx(db, budget_id, data):
    """budget 저장 시 card_tx 자동 동기화"""
    card_id = data.get('card_id') or None
    if card_id:
        cur = db.cursor()
        cur.execute("SELECT id FROM card_tx WHERE budget_id = %s", (budget_id,))
        existing = cur.fetchone()
        cur.close()
        if existing:
            cur = db.cursor()
            cur.execute(
            "UPDATE card_tx SET card_id=%s, date=%s, name=%s, category=%s, amount=%s, memo=%s WHERE budget_id=%s",
            (card_id, data.get('date'), data.get('name'), data.get('category'),
            data.get('amount', 0), data.get('memo'), budget_id)
            )
            cur.close()
        else:
            cur = db.cursor()
            cur.execute(
            "INSERT INTO card_tx (card_id, date, name, category, amount, installment, memo, budget_id) VALUES (%s,%s,%s,%s,%s,1,%s,%s)",
            (card_id, data.get('date'), data.get('name'), data.get('category'),
            data.get('amount', 0), data.get('memo'), budget_id)
            )
            cur.close()
    else:
        cur = db.cursor()
        cur.execute("DELETE FROM card_tx WHERE budget_id = %s", (budget_id,))
        cur.close()


@app.route('/api/budget', methods=['GET', 'POST'])
def api_budget():
    db = get_db()
    if request.method == 'GET':
        year  = request.args.get('year')
        month = request.args.get('month')

        # ── [추가] 해당 연월의 고정지출 자동 생성 ──
        if year and month:
            try:
                _generate_recurring_budget(db, int(year), int(month))
            except Exception as e:
                print("Error in _generate_recurring_budget:", e)

        query = """SELECT b.*, c.card_name
                   FROM budget b
                   LEFT JOIN card_info c ON b.card_id = c.id"""
        params = []
        if year and month:
            query += " WHERE to_char(b.date::date, 'YYYY') = %s AND to_char(b.date::date, 'MM') = %s"
            params = [year, month.zfill(2)]
        query += " ORDER BY b.date DESC"
        cur = db.cursor()
        cur.execute(query, params)
        rows = cur.fetchall()
        cur.close()
        db.close()
        return jsonify(rows_to_list(rows))

    data = request.json
    type_ = data.get('type', '')

    # 카테고리 미설정 시 규칙 자동 적용
    if not data.get('category'):
        auto_cat = _apply_budget_category_rule(db, data.get('name', ''))
        if auto_cat:
            data['category'] = auto_cat

    # recurring_id 컬럼 존재 여부에 따라 INSERT 분기
    cur = db.cursor()
    cur.execute(
        "INSERT INTO budget (date,category,name,type,payment_method,amount,memo,card_id,is_auto_generated) "
        "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,FALSE) RETURNING id",
        (data['date'], data.get('category'), data.get('name'), type_,
         data.get('payment_method'), data['amount'], data.get('memo'),
         data.get('card_id') or None)
    )
    budget_id = cur.fetchone()[0]
    cur.close()
    _sync_card_tx(db, budget_id, data)
    db.commit()
    db.close()
    return jsonify({'ok': True, 'id': budget_id,
                    'type': type_, 'name': data.get('name'),
                    'date': data.get('date'),
                    'amount': data.get('amount'),
                    'category': data.get('category'),
                    'payment_method': data.get('payment_method'),
                    'card_id': data.get('card_id'),
                    'memo': data.get('memo'),
                    'recurring_id': None}), 201


@app.route('/api/budget/receipt', methods=['POST'])
def api_budget_receipt():
    if 'file' not in request.files:
        return jsonify({'error': '파일이 없습니다.'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': '선택된 파일이 없습니다.'}), 400

    filename = file.filename
    img_bytes = file.read()

    tx_date = date.today().isoformat()
    amount = 0
    name = ''
    memo = f'영수증 첨부: {filename}'
    payment_method = '신용카드'
    type_ = '변동지출'
    category = ''

    # 1. 파일명에서 파싱
    date_m = re.search(r'(20\d{2})[-_.]?(\d{2})[-_.]?(\d{2})', filename)
    if date_m:
        tx_date = f"{date_m.group(1)}-{date_m.group(2)}-{date_m.group(3)}"

    amount_m = re.findall(r'(\d{1,3}(?:,\d{3})+|\d{3,7})\s*(?:원|KRW)', filename)
    if amount_m:
        amount = int(amount_m[-1].replace(',', ''))
    else:
        nums = re.findall(r'\d+', filename)
        for num in nums:
            val = int(num)
            if val >= 1000 and val != int(tx_date.replace('-', '')):
                amount = val
                break

    common_stores = {
        '스타벅스': '식비', '이마트': '쇼핑', '다이소': '쇼핑', '쿠팡': '쇼핑', 
        '택시': '교통비', 'CU': '식비', 'GS25': '식비', '세븐일레븐': '식비',
        '맥도날드': '식비', '카페': '식비', '식당': '식비', '파리바게뜨': '식비',
        '올리브영': '쇼핑', '주유소': '교통비', '병원': '의료비', '약국': '의료비'
    }
    for store, cat in common_stores.items():
        if store in filename:
            name = store
            category = cat
            break

    # 2. OCR 시도 (설치되어 있는 경우)
    extracted_text = ""
    try:
        from PIL import Image
        import io
        img = Image.open(io.BytesIO(img_bytes))
        try:
            import pytesseract
            extracted_text = pytesseract.image_to_string(img, lang='kor+eng')
        except Exception:
            pass
    except Exception:
        pass

    if extracted_text:
        if amount == 0:
            amts = re.findall(r'(?:합계|금액|결제금액|총액|승인금액)\s*[:=]?\s*([\d,]+)\s*(?:원)?', extracted_text)
            if amts:
                amount = int(amts[0].replace(',', ''))
            else:
                nums = re.findall(r'\b\d{1,3}(?:,\d{3})+\b', extracted_text)
                if nums:
                    amount = int(nums[-1].replace(',', ''))
        
        if date_m is None:
            dt_m = re.search(r'(20\d{2})[-/.](\d{2})[-/.](\d{2})', extracted_text)
            if dt_m:
                tx_date = f"{dt_m.group(1)}-{dt_m.group(2)}-{dt_m.group(3)}"

        if not name:
            lines = [line.strip() for line in extracted_text.split('\n') if line.strip()]
            for line in lines:
                for store, cat in common_stores.items():
                    if store in line:
                        name = store
                        category = cat
                        break
                if name: break
            if not name and lines:
                cleaned = re.sub(r'[^\w\s]', '', lines[0]).strip()
                if cleaned and not cleaned.isdigit():
                    name = cleaned[:15]

    if not name:
        name = os.path.splitext(filename)[0][:15]
    if amount == 0:
        amount = 10000
    
    db = get_db()
    if not category:
        auto_cat = _apply_budget_category_rule(db, name)
        if auto_cat:
            category = auto_cat
        else:
            category = '식비'

    cur = db.cursor()
    try:
        cur.execute(
            "INSERT INTO budget (date, category, name, type, payment_method, amount, memo) "
            "VALUES (%s, %s, %s, %s, %s, %s, %s) RETURNING id",
            (tx_date, category, name, type_, payment_method, amount, memo)
        )
        budget_id = cur.fetchone()[0]
        db.commit()
    except Exception as e:
        db.rollback()
        cur.close()
        db.close()
        return jsonify({'error': f'DB 저장 실패: {str(e)}'}), 500
    cur.close()
    db.close()

    return jsonify({
        'ok': True,
        'budget_id': budget_id,
        'parsed': {
            'date': tx_date,
            'category': category,
            'name': name,
            'amount': amount,
            'memo': memo
        }
    }), 201


@app.route('/api/budget/<int:rid>', methods=['PUT', 'DELETE'])
def api_budget_detail(rid):
    db = get_db()
    if request.method == 'PUT':
        data = request.json
        type_ = data.get('type', '')

        # 현재 행의 recurring_id, date 조회 (컬럼 없으면 None)
        try:
            cur = db.cursor()
            cur.execute("SELECT recurring_id, date FROM budget WHERE id=%s", (rid,))
            row = cur.fetchone()
            recurring_id = row[0] if row else None
            row_date     = row[1] if row else None
            cur.close()
        except Exception:
            db.rollback()
            recurring_id = None
            row_date     = None

        cur = db.cursor()
        new_recurring_id = data.get('recurring_id') or recurring_id or None
        cur.execute(
            "UPDATE budget SET date=%s,category=%s,name=%s,type=%s,payment_method=%s,amount=%s,memo=%s,card_id=%s,recurring_id=%s WHERE id=%s",
            (data.get('date'), data.get('category'), data.get('name'), type_,
             data.get('payment_method'), data.get('amount', 0), data.get('memo'),
             data.get('card_id') or None, new_recurring_id, rid)
        )
        recurring_id = new_recurring_id
        cur.close()

        # 고정지출: 마스터 + 이후 모든 월 동기화
        if type_ == '고정지출' and recurring_id:
            try:
                cur = db.cursor()
                cur.execute(
                    "UPDATE recurring_budget SET name=%s,category=%s,payment_method=%s,card_id=%s,amount=%s,memo=%s WHERE id=%s",
                    (data.get('name'), data.get('category'), data.get('payment_method'),
                     data.get('card_id') or None, data.get('amount', 0), data.get('memo'), recurring_id)
                )
                cur.close()
                cur = db.cursor()
                cur.execute(
                    "UPDATE budget SET name=%s,category=%s,payment_method=%s,card_id=%s,amount=%s,memo=%s "
                    "WHERE recurring_id=%s AND date > %s",
                    (data.get('name'), data.get('category'), data.get('payment_method'),
                     data.get('card_id') or None, data.get('amount', 0), data.get('memo'),
                     recurring_id, row_date)
                )
                cur.close()
            except Exception:
                db.rollback()

        _sync_card_tx(db, rid, data)

        # 카테고리가 설정된 경우 항목명을 키워드로 학습
        new_category = data.get('category', '')
        new_name     = data.get('name', '')
        rule_id = None
        if new_category and new_name:
            rule_id = _learn_budget_category(db, new_name, new_category)

        db.commit()
        db.close()
        return jsonify({'ok': True, 'id': rid,
                        'type': type_, 'name': new_name,
                        'date': data.get('date'),
                        'amount': data.get('amount'),
                        'category': new_category,
                        'payment_method': data.get('payment_method'),
                        'card_id': data.get('card_id'),
                        'memo': data.get('memo'),
                        'recurring_id': recurring_id,
                        'learned':   rule_id is not None,
                        'rule_id':   rule_id,
                        'keyword':   new_name     if rule_id else None})

    # DELETE
    mode = request.args.get('mode', 'single')  # 'single' | 'forward'
    cur = db.cursor()
    cur.execute("SELECT recurring_id, date FROM budget WHERE id=%s", (rid,))
    row = cur.fetchone()
    recurring_id = row[0] if row else None
    row_date     = row[1] if row else None
    cur.close()

    if mode == 'forward' and recurring_id:
        # 이 날짜 이후 + 이 행 포함 모두 삭제, 반복 비활성화
        cur = db.cursor()
        cur.execute(
            "DELETE FROM card_tx WHERE budget_id IN "
            "(SELECT id FROM budget WHERE recurring_id=%s AND date >= %s)",
            (recurring_id, row_date)
        )
        cur.close()
        cur = db.cursor()
        cur.execute("DELETE FROM budget WHERE recurring_id=%s AND date >= %s", (recurring_id, row_date))
        cur.close()
        cur = db.cursor()
        cur.execute("UPDATE recurring_budget SET is_active=FALSE WHERE id=%s", (recurring_id,))
        cur.close()
    else:
        cur = db.cursor()
        cur.execute("DELETE FROM card_tx WHERE budget_id=%s", (rid,))
        cur.close()
        cur = db.cursor()
        cur.execute("DELETE FROM budget WHERE id=%s", (rid,))
        cur.close()

    db.commit()
    db.close()
    return jsonify({'ok': True})


# ── 고정지출 템플릿 관리 ──────────────────────────────────────
@app.route('/api/recurring-budget', methods=['GET', 'POST'])
def api_recurring_budget():
    db = get_db()
    if request.method == 'GET':
        cur = db.cursor()
        cur.execute("""
            SELECT r.*, c.card_name
            FROM recurring_budget r
            LEFT JOIN card_info c ON r.card_id = c.id
            ORDER BY r.day_of_month, r.name
        """)
        rows = cur.fetchall()
        cur.close(); db.close()
        return jsonify(rows_to_list(rows))

    d = request.json or {}
    cur = db.cursor()
    cur.execute("""
        INSERT INTO recurring_budget
            (name, category, payment_method, amount, card_id,
             day_of_month, start_month, end_month, memo)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
    """, (d.get('name'), d.get('category'), d.get('payment_method'),
          d.get('amount', 0), d.get('card_id') or None,
          d.get('day_of_month', 1), d.get('start_month'),
          d.get('end_month') or None, d.get('memo')))
    cur.close(); db.commit(); db.close()
    return jsonify({'ok': True}), 201


@app.route('/api/recurring-budget/<int:rid>', methods=['PUT', 'DELETE'])
def api_recurring_budget_detail(rid):
    db = get_db()
    if request.method == 'DELETE':
        cur = db.cursor()
        cur.execute("UPDATE recurring_budget SET is_active=FALSE WHERE id=%s", (rid,))
        cur.close(); db.commit(); db.close()
        return jsonify({'ok': True})
    d = request.json or {}
    cur = db.cursor()
    cur.execute("""
        UPDATE recurring_budget
        SET name=%s, category=%s, payment_method=%s, amount=%s, card_id=%s,
            day_of_month=%s, start_month=%s, end_month=%s, memo=%s, is_active=%s
        WHERE id=%s
    """, (d.get('name'), d.get('category'), d.get('payment_method'),
          d.get('amount', 0), d.get('card_id') or None,
          d.get('day_of_month', 1), d.get('start_month'),
          d.get('end_month') or None, d.get('memo'),
          d.get('is_active', True), rid))
    cur.close(); db.commit(); db.close()
    return jsonify({'ok': True})


# ── API: 가계부 카테고리 ─────────────────────────────────────
@app.route('/api/budget-categories', methods=['GET', 'POST'])
def api_budget_categories():
    db = get_db()
    cur = db.cursor()
    if request.method == 'GET':
        cur.execute("SELECT id, name FROM budget_categories ORDER BY sort_order, id")
        rows = cur.fetchall()
        db.close()
        return jsonify([{'id': r['id'], 'name': r['name']} for r in rows])
    data = request.json or {}
    name = (data.get('name') or '').strip()
    if not name:
        db.close()
        return jsonify({'error': 'name required'}), 400
    cur.execute("SELECT COALESCE(MAX(sort_order),0)+1 FROM budget_categories")
    next_order = cur.fetchone()[0]
    cur.execute(
        "INSERT INTO budget_categories (name, sort_order) VALUES (%s, %s) ON CONFLICT (name) DO NOTHING",
        (name, next_order)
    )
    db.commit()
    db.close()
    return jsonify({'ok': True})

@app.route('/api/budget-categories/<int:cid>', methods=['DELETE'])
def api_budget_category_delete(cid):
    db = get_db()
    cur = db.cursor()
    cur.execute("DELETE FROM budget_categories WHERE id=%s", (cid,))
    db.commit()
    db.close()
    return jsonify({'ok': True})


# ── API: 가계부 자동 분류 규칙 ──────────────────────────────────
@app.route('/api/budget-category-rules', methods=['GET', 'POST'])
def api_budget_category_rules():
    db = get_db()
    cur = db.cursor()
    if request.method == 'GET':
        cur.execute("SELECT id, keyword, category FROM budget_category_rules ORDER BY id DESC")
        rows = cur.fetchall()
        db.close()
        return jsonify(rows_to_list(rows))
    data = request.json or {}
    kw  = (data.get('keyword') or '').strip()
    cat = (data.get('category') or '').strip()
    if not kw or not cat:
        db.close()
        return jsonify({'error': 'keyword and category required'}), 400
    cur.execute(
        "INSERT INTO budget_category_rules (keyword, category) VALUES (%s, %s) "
        "ON CONFLICT (keyword) DO UPDATE SET category = EXCLUDED.category",
        (kw, cat)
    )
    db.commit()
    db.close()
    return jsonify({'ok': True})

@app.route('/api/budget-category-rules/<int:rid>', methods=['DELETE', 'PUT'])
def api_budget_category_rule_detail(rid):
    db = get_db()
    cur = db.cursor()
    if request.method == 'PUT':
        data = request.json or {}
        kw  = (data.get('keyword') or '').strip()
        cat = (data.get('category') or '').strip()
        if kw and cat:
            cur.execute("DELETE FROM budget_category_rules WHERE id=%s", (rid,))
            cur.execute(
                "INSERT INTO budget_category_rules (keyword, category) VALUES (%s, %s) "
                "ON CONFLICT (keyword) DO UPDATE SET category = EXCLUDED.category",
                (kw, cat)
            )
    else:
        cur.execute("DELETE FROM budget_category_rules WHERE id=%s", (rid,))
    db.commit()
    db.close()
    return jsonify({'ok': True})


def _apply_budget_category_rule(db, name):
    """항목명에 매칭되는 규칙이 있으면 카테고리를 반환."""
    if not name:
        return None
    try:
        cur = db.cursor()
        cur.execute(
            "SELECT keyword, category FROM budget_category_rules ORDER BY LENGTH(keyword) DESC"
        )
        rules = cur.fetchall()
        cur.close()
        name_lower = name.lower()
        for r in rules:
            if r['keyword'].lower() in name_lower:
                return r['category']
    except Exception:
        pass
    return None


def _learn_budget_category(db, name, category):
    """항목명을 키워드로 학습. 새로 등록된 경우 rule_id 반환, 아니면 None."""
    if not name or not category:
        return None
    try:
        cur = db.cursor()
        cur.execute("SELECT keyword FROM budget_category_rules")
        rules = cur.fetchall()
        cur.close()
        name_lower = name.lower()
        # 이미 커버되는 규칙이 있으면 등록하지 않음
        if any(r['keyword'].lower() in name_lower for r in rules):
            return None
        cur = db.cursor()
        cur.execute(
            "INSERT INTO budget_category_rules (keyword, category) VALUES (%s, %s) "
            "ON CONFLICT (keyword) DO UPDATE SET category = EXCLUDED.category RETURNING id",
            (name, category)
        )
        row = cur.fetchone()
        cur.close()
        return row[0] if row else None
    except Exception:
        return None


# ── API: 카드 정보 ───────────────────────────────────────────
@app.route('/api/cards', methods=['GET', 'POST'])
def api_cards():
    db = get_db()
    if request.method == 'GET':
        cur = db.cursor()
        cur.execute("SELECT * FROM card_info ORDER BY card_num")
        rows = cur.fetchall()
        cur.close()
        db.close()
        return jsonify(rows_to_list(rows))

    data = request.json
    cur = db.cursor()
    cur.execute(
    "INSERT INTO card_info (card_num, card_name, limit_amount, payment_day, billing_day, benefit) VALUES (%s,%s,%s,%s,%s,%s)",
    (data['card_num'], data.get('card_name'), data.get('limit_amount', 0),
    data.get('payment_day'), data.get('billing_day'), data.get('benefit'))
    )
    cur.close()
    db.commit()
    db.close()
    return jsonify({'ok': True}), 201


@app.route('/api/cards/<int:cid>', methods=['PUT', 'DELETE'])
def api_card_detail(cid):
    db = get_db()
    if request.method == 'PUT':
        data = request.json
        cur = db.cursor()
        cur.execute(
        "UPDATE card_info SET card_num=%s, card_name=%s, limit_amount=%s, payment_day=%s, billing_day=%s, benefit=%s WHERE id=%s",
        (data.get('card_num'), data.get('card_name'), data.get('limit_amount', 0),
        data.get('payment_day'), data.get('billing_day'), data.get('benefit'), cid)
        )
        cur.close()
        db.commit()
        db.close()
        return jsonify({'ok': True})
    cur = db.cursor()
    cur.execute("DELETE FROM card_info WHERE id = %s", (cid,))
    cur.close()
    db.commit()
    db.close()
    return jsonify({'ok': True})


# ── API: 카드 거래내역 ───────────────────────────────────────
@app.route('/api/card-tx', methods=['GET', 'POST'])
def api_card_tx():
    db = get_db()
    if request.method == 'GET':
        card_id = request.args.get('card_id')
        year    = request.args.get('year')
        month   = request.args.get('month')
        query   = "SELECT t.*, c.card_name FROM card_tx t LEFT JOIN card_info c ON t.card_id = c.id"
        params  = []
        conds   = []
        if card_id:
            conds.append("t.card_id = %s"); params.append(card_id)
        if year and month:
            conds.append("to_char(t.date::date, 'YYYY') = %s"); params.append(year)
            conds.append("to_char(t.date::date, 'MM') = %s"); params.append(month.zfill(2))
        if conds:
            query += " WHERE " + " AND ".join(conds)
        query += " ORDER BY t.date DESC"
        cur = db.cursor()
        cur.execute(query, params)
        rows = cur.fetchall()
        cur.close()
        total = sum(r['amount'] for r in rows)
        # 카테고리별 집계 (없는 건 → '미분류')
        cat_map = {}
        for r in rows:
            cat = r['category'] or '미분류'
            cat_map[cat] = cat_map.get(cat, 0) + r['amount']
        by_category = sorted(
            [{'category': c, 'total': t} for c, t in cat_map.items()],
            key=lambda x: x['total'], reverse=True
        )
        # 자금 그룹별 집계
        cur = db.cursor()
        cur.execute("SELECT id, name FROM fund_groups")
        fund_group_names = {r['id']: r['name'] for r in cur.fetchall()}
        cur.close()
        fund_map = {}
        for r in rows:
            gid = r['fund_group_id']
            name = fund_group_names.get(gid, '미지정') if gid else '미지정'
            fund_map[name] = fund_map.get(name, 0) + r['amount']
        by_fund_group = sorted(
            [{'name': n, 'total': t} for n, t in fund_map.items()],
            key=lambda x: x['total'], reverse=True
        )
        db.close()
        return jsonify({'rows': rows_to_list(rows), 'total': total,
                        'by_category': by_category, 'by_fund_group': by_fund_group})

    data = request.json
    cur = db.cursor()
    cur.execute(
    "INSERT INTO card_tx (card_id, date, name, category, amount, installment, memo) VALUES (%s,%s,%s,%s,%s,%s,%s)",
    (data.get('card_id'), data['date'], data.get('name'), data.get('category'),
    data['amount'], data.get('installment', 1), data.get('memo'))
    )
    cur.close()
    db.commit()
    db.close()
    return jsonify({'ok': True}), 201


@app.route('/api/card-tx/<int:rid>', methods=['PUT', 'DELETE'])
def api_card_tx_detail(rid):
    db = get_db()
    if request.method == 'PUT':
        data = request.json
        category = data.get('category') or ''
        locked = 1 if category else 0
        fund_group_id = data.get('fund_group_id')
        fund_group_locked = 1 if fund_group_id else 0
        cur = db.cursor()
        cur.execute(
        "UPDATE card_tx SET card_id=%s, date=%s, name=%s, category=%s, amount=%s, installment=%s, memo=%s,"
        " category_locked=%s, fund_group_id=%s, fund_group_locked=%s WHERE id=%s",
        (data.get('card_id'), data.get('date'), data.get('name'), category,
        data.get('amount', 0), data.get('installment', 1), data.get('memo'),
        locked, fund_group_id, fund_group_locked, rid)
        )
        cur.close()
        db.commit()
        db.close()
        return jsonify({'ok': True})
    cur = db.cursor()
    cur.execute("DELETE FROM card_tx WHERE id = %s", (rid,))
    cur.close()
    db.commit()
    db.close()
    return jsonify({'ok': True})


@app.route('/api/card-tx/bulk', methods=['DELETE'])
def api_card_tx_bulk_delete():
    data = request.json or {}
    ids  = data.get('ids', [])
    if not ids:
        return jsonify({'error': 'ids가 필요합니다'}), 400
    db = get_db()
    placeholders = ','.join('%s' * len(ids))
    cur = db.cursor()
    cur.execute(f"DELETE FROM card_tx WHERE id IN ({placeholders})", ids)
    deleted = cur.rowcount
    cur.close()
    db.commit(); db.close()
    return jsonify({'ok': True, 'deleted': deleted})


# ── API: 가계부 대조 ──────────────────────────────────────────
@app.route('/api/card-reconcile')
def api_card_reconcile():
    """카드 거래내역 ↔ 가계부 자동 매칭"""
    card_id = request.args.get('card_id', type=int)
    year    = request.args.get('year',    type=int)
    month   = request.args.get('month',   type=int)
    if not all([card_id, year, month]):
        return jsonify({'error': 'card_id, year, month 필요'}), 400

    from datetime import timedelta
    first_day   = date(year, month, 1)
    last_month  = month % 12 + 1
    last_year   = year + (1 if month == 12 else 0)
    last_day    = date(last_year, last_month, 1) - timedelta(days=1)
    range_start = (first_day - timedelta(days=3)).isoformat()
    range_end   = (last_day  + timedelta(days=3)).isoformat()
    ym = f"{year}-{month:02d}"

    db  = get_db()
    cur = db.cursor()

    cur.execute("""
        SELECT id, date, name, category, amount, installment, memo
        FROM card_tx
        WHERE card_id = %s AND to_char(date::date, 'YYYY-MM') = %s
        ORDER BY date, id
    """, (card_id, ym))
    card_txs = rows_to_list(cur.fetchall())

    cur.execute("""
        SELECT id, date, name, category, amount, payment_method, memo
        FROM budget
        WHERE date >= %s AND date <= %s AND amount > 0
        ORDER BY date, id
    """, (range_start, range_end))
    budget_entries = rows_to_list(cur.fetchall())
    cur.close(); db.close()

    used = set()
    result = []
    for tx in card_txs:
        tx_date   = date.fromisoformat(str(tx['date'])[:10])
        tx_amount = int(tx['amount'])
        best = None; best_days = 999
        for b in budget_entries:
            if b['id'] in used or int(b['amount']) != tx_amount:
                continue
            diff = abs((tx_date - date.fromisoformat(str(b['date'])[:10])).days)
            if diff <= 3 and diff < best_days:
                best_days = diff; best = b
        if best:
            used.add(best['id'])
            result.append({**tx, 'status': 'matched',
                           'budget_id': best['id'], 'budget_name': best['name'],
                           'budget_date': str(best['date'])[:10],
                           'budget_cat': best.get('category',''), 'date_diff': best_days})
        else:
            result.append({**tx, 'status': 'unmatched', 'budget_id': None})

    matched   = sum(1 for r in result if r['status'] == 'matched')
    unmatched = len(result) - matched
    return jsonify({'items': result, 'matched': matched,
                    'unmatched': unmatched, 'total': len(result)})


@app.route('/api/card-reconcile/add-budget', methods=['POST'])
def api_card_reconcile_add_budget():
    """미매칭 카드 거래를 가계부에 일괄 추가"""
    data        = request.json or {}
    card_tx_ids = data.get('card_tx_ids', [])
    card_id     = data.get('card_id')
    if not card_tx_ids:
        return jsonify({'ok': True, 'added': 0})

    db  = get_db()
    cur = db.cursor()
    placeholders = ','.join(['%s'] * len(card_tx_ids))
    cur.execute(f"SELECT id, date, name, category, amount, memo FROM card_tx WHERE id IN ({placeholders})",
                card_tx_ids)
    txs = rows_to_list(cur.fetchall())
    added = 0
    for tx in txs:
        cur.execute("""
            INSERT INTO budget (date, category, name, type, payment_method, amount, memo, card_id)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        """, (str(tx['date'])[:10], tx.get('category') or '', tx.get('name') or '',
              '변동', '카드', int(tx['amount']), tx.get('memo') or '', card_id))
        added += 1
    cur.close(); db.commit(); db.close()
    return jsonify({'ok': True, 'added': added})


@app.route('/api/card-tx/auto-categorize', methods=['POST'])
def api_card_tx_auto_categorize():
    data    = request.json or {}
    card_id = data.get('card_id')
    year    = data.get('year')
    month   = data.get('month')

    db = get_db()
    query  = "SELECT id, name FROM card_tx WHERE category_locked = 0"
    params = []
    if card_id:
        query += " AND card_id = %s"; params.append(card_id)
    if year and month:
        query += " AND to_char(date::date, 'YYYY') = %s AND to_char(date::date, 'MM') = %s"
        params += [year, month.zfill(2)]

    cur = db.cursor()
    cur.execute(query, params)
    rows = cur.fetchall()
    cur.close()
    updated = 0
    for row in rows:
        hint = _get_category_hint(db, row['name'])
        if hint:
            cur = db.cursor()
            cur.execute("UPDATE card_tx SET category=%s WHERE id=%s", (hint, row['id']))
            cur.close()
            updated += 1
    db.commit()
    db.close()
    return jsonify({'ok': True, 'updated': updated})


# ── 포지션 계산 헬퍼 ─────────────────────────────────────────
def calc_position(transactions):
    qty = 0.0
    avg_cost = 0.0
    realized = 0.0
    for tx in transactions:
        tq = float(tx['quantity'] or 0)
        tp = float(tx['price'] or 0)
        if tq <= 0:
            continue
        if tx['tx_type'] in ('buy', '매수'):
            new_qty = qty + tq
            avg_cost = (qty * avg_cost + tq * tp) / new_qty if new_qty > 0 else 0.0
            qty = new_qty
        else:  # sell
            realized += (tp - avg_cost) * tq
            qty = max(0.0, qty - tq)
            if qty == 0.0:
                avg_cost = 0.0
    return qty, avg_cost, realized


def recalc_realized_pnl(db, item_id, is_etf=False):
    """해당 종목의 모든 거래를 날짜순으로 다시 훑어 매도 건의 실현손익을 일괄 재계산/저장한다.
    과거 거래 하나를 수정/삭제해도 이후 매도 건들의 실현손익이 항상 정합성을 유지하도록 한다."""
    cur = db.cursor()
    table = 'etf_tx' if is_etf else 'stock_tx'
    col = 'etf_id' if is_etf else 'stock_id'
    cur.execute(f"SELECT id, tx_date, tx_type, price, quantity FROM {table} WHERE {col} = %s ORDER BY tx_date, id", (item_id,))
    txs = cur.fetchall()

    qty = 0.0
    avg_cost = 0.0
    for tx in txs:
        tq = float(tx['quantity'] or 0)
        tp = float(tx['price'] or 0)
        if tq <= 0:
            continue
        if tx['tx_type'] in ('buy', '매수'):
            new_qty = qty + tq
            avg_cost = (qty * avg_cost + tq * tp) / new_qty if new_qty > 0 else 0.0
            qty = new_qty
        elif tx['tx_type'] in ('sell', '매도'):
            realized = (tp - avg_cost) * tq
            cur.execute(f"UPDATE {table} SET realized_pnl = %s WHERE id = %s", (realized, tx['id']))
            qty = max(0.0, qty - tq)
            if qty == 0.0:
                avg_cost = 0.0
    cur.close()


def compute_realized_pnl_map(rows, key_col):
    """주어진 거래 목록을 종목별로 묶어 날짜순으로 재계산한 매도 건 실현손익 맵(tx id -> 값)을 반환한다.
    저장된 컬럼값에 의존하지 않고 항상 거래내역 자체에서 다시 계산하므로 쓰기 경합/누락에도 안전하다."""
    from collections import defaultdict
    groups = defaultdict(list)
    for r in rows:
        groups[r[key_col]].append(r)
    result = {}
    for grp in groups.values():
        grp_sorted = sorted(grp, key=lambda r: (r['tx_date'], r['id']))
        qty = 0.0
        avg_cost = 0.0
        for tx in grp_sorted:
            tq = float(tx['quantity'] or 0)
            tp = float(tx['price'] or 0)
            if tq <= 0:
                result[tx['id']] = 0.0
                continue
            if tx['tx_type'] in ('buy', '매수'):
                new_qty = qty + tq
                avg_cost = (qty * avg_cost + tq * tp) / new_qty if new_qty > 0 else 0.0
                qty = new_qty
                result[tx['id']] = 0.0
            elif tx['tx_type'] in ('sell', '매도'):
                result[tx['id']] = (tp - avg_cost) * tq
                qty = max(0.0, qty - tq)
                if qty == 0.0:
                    avg_cost = 0.0
            else:
                result[tx['id']] = 0.0
    return result


@app.route('/api/recalc-realized-pnl', methods=['POST'])
def api_recalc_realized_pnl():
    db = get_db()
    cur = db.cursor()
    cur.execute("SELECT id FROM stocks")
    stock_ids = [r['id'] for r in cur.fetchall()]
    cur.execute("SELECT id FROM etf")
    etf_ids = [r['id'] for r in cur.fetchall()]
    cur.close()
    for sid in stock_ids:
        recalc_realized_pnl(db, sid)
    for eid in etf_ids:
        recalc_realized_pnl(db, eid, is_etf=True)
    db.commit()
    db.close()
    return jsonify({'ok': True})


def _qty_before(db, item_id, tx_date, tx_type, table, col, exclude_id=None):
    """해당 거래 시점(같은 날짜는 id 기준) 이전까지의 보유 수량을 계산한다. 매도 검증용."""
    cur = db.cursor()
    query = f"SELECT tx_type, quantity FROM {table} WHERE {col} = %s AND tx_date <= %s"
    params = [item_id, tx_date]
    if exclude_id is not None:
        query += " AND id != %s"
        params.append(exclude_id)
    cur.execute(query, params)
    rows = cur.fetchall()
    cur.close()
    qty = 0.0
    for r in rows:
        tq = float(r['quantity'] or 0)
        if r['tx_type'] in ('buy', '매수'):
            qty += tq
        elif r['tx_type'] in ('sell', '매도'):
            qty -= tq
    return qty


# ── API: 주식 ────────────────────────────────────────────────
@app.route('/api/stocks', methods=['GET', 'POST'])
def api_stocks():
    db = get_db()
    if request.method == 'GET':
        cur = db.cursor()
        cur.execute("SELECT id, name, ticker, current_price, dividend, memo, category, COALESCE(ath,0) as ath, realized_pnl_override FROM stocks ORDER BY name")
        stocks = [dict(r) for r in cur.fetchall()]
        # SQL 기반 qty (대시보드·테크트리와 동일 기준, 음수 0 처리)
        cur.execute("""
            SELECT s.id,
                GREATEST(0, COALESCE(SUM(CASE WHEN t.tx_type IN ('buy','매수') THEN t.quantity ELSE 0 END), 0)
                          - COALESCE(SUM(CASE WHEN t.tx_type IN ('sell','매도') THEN t.quantity ELSE 0 END), 0)) AS qty
            FROM stocks s LEFT JOIN stock_tx t ON t.stock_id = s.id
            GROUP BY s.id
        """)
        sql_qty = {r['id']: float(r['qty'] or 0) for r in cur.fetchall()}
        # avg_price / realized_pnl 은 calc_position(FIFO) 사용
        cur.execute("SELECT stock_id, tx_type, price, quantity, COALESCE(fee,0) as fee, COALESCE(realized_pnl,0) as realized_pnl FROM stock_tx ORDER BY stock_id, tx_date, id")
        all_tx = cur.fetchall()
        cur.close()

        from collections import defaultdict
        tx_by_stock = defaultdict(list)
        for tx in all_tx:
            tx_by_stock[tx['stock_id']].append(tx)

        result = []
        try:
            for s in stocks:
                qty      = sql_qty.get(s['id'], 0.0)
                _, avg, realized = calc_position(tx_by_stock[s['id']])
                avg      = avg if (avg is not None and avg == avg) else 0.0  # NaN guard
                eval_amt = round(qty * _sf(s['current_price']))
                cost_amt = round(qty * avg)
                s['quantity']       = qty
                s['avg_price']      = avg if qty > 0 else None
                s['eval_amount']    = eval_amt
                s['unrealized_pnl'] = eval_amt - cost_amt
                s['return_rate']    = round((eval_amt - cost_amt) / cost_amt * 100, 2) if cost_amt else 0
                override = s.pop('realized_pnl_override', None)
                s['realized_pnl']       = round(override) if override is not None else round(realized)
                s['realized_pnl_fixed']  = override is not None
                result.append(s)
        except Exception as e:
            import traceback
            db.close()
            return jsonify({'error': str(e), 'trace': traceback.format_exc()}), 500
        db.close()
        return jsonify(result)

    data = request.json
    cur = db.cursor()
    cur.execute(
    "INSERT INTO stocks (name, ticker, current_price, dividend, memo, category) VALUES (%s,%s,%s,%s,%s,%s)",
    (data.get('name'), data.get('ticker'),
    data.get('current_price', 0), data.get('dividend', 0), data.get('memo'), data.get('category'))
    )
    cur.close()
    db.commit()
    db.close()
    return jsonify({'ok': True}), 201


@app.route('/api/stocks/<int:rid>', methods=['PUT', 'DELETE'])
def api_stocks_detail(rid):
    db = get_db()
    if request.method == 'PUT':
        data = request.json
        cur = db.cursor()
        cur.execute(
        "UPDATE stocks SET name=%s, ticker=%s, current_price=%s, dividend=%s, memo=%s, category=%s WHERE id=%s",
        (data.get('name'), data.get('ticker'),
        data.get('current_price', 0), data.get('dividend', 0), data.get('memo'), data.get('category'), rid)
        )
        cur.close()
        db.commit()
        db.close()
        return jsonify({'ok': True})
    cur = db.cursor()
    cur.execute("DELETE FROM stock_tx WHERE stock_id = %s", (rid,))
    cur.close()
    cur = db.cursor()
    cur.execute("DELETE FROM stocks WHERE id = %s", (rid,))
    cur.close()
    db.commit()
    db.close()
    return jsonify({'ok': True})


# ── API: 주식 구분 카테고리 ──────────────────────────────────
@app.route('/api/stock-categories', methods=['GET', 'POST'])
def api_stock_categories():
    db = get_db()
    cur = db.cursor()
    if request.method == 'GET':
        cur.execute("SELECT id, name FROM stock_categories ORDER BY sort_order, id")
        rows = rows_to_list(cur.fetchall())
        cur.close(); db.close()
        return jsonify(rows)
    data = request.json
    cur.execute("SELECT COALESCE(MAX(sort_order),0)+1 FROM stock_categories")
    order = cur.fetchone()[0]
    cur.execute(
        "INSERT INTO stock_categories (name, sort_order) VALUES (%s, %s) ON CONFLICT (name) DO NOTHING",
        (data.get('name'), order)
    )
    cur.close(); db.commit(); db.close()
    return jsonify({'ok': True}), 201

@app.route('/api/stock-categories/<int:cid>', methods=['DELETE'])
def api_stock_category_delete(cid):
    db = get_db()
    cur = db.cursor()
    cur.execute("DELETE FROM stock_categories WHERE id=%s", (cid,))
    cur.close(); db.commit(); db.close()
    return jsonify({'ok': True})


# ── API: 주식 거래내역 ────────────────────────────────────────
@app.route('/api/stock-tx', methods=['GET', 'POST'])
def api_stock_tx():
    db = get_db()
    if request.method == 'GET':
        stock_id = request.args.get('stock_id')
        query  = "SELECT t.*, s.name, s.ticker FROM stock_tx t LEFT JOIN stocks s ON t.stock_id = s.id"
        params = []
        if stock_id:
            query += " WHERE t.stock_id = %s"
            params.append(stock_id)
        query += " ORDER BY t.tx_date DESC, t.id DESC"
        cur = db.cursor()
        cur.execute(query, params)
        rows = [dict(r) for r in cur.fetchall()]
        cur.close()
        db.close()
        pnl_map = compute_realized_pnl_map(rows, 'stock_id')
        for r in rows:
            r['realized_pnl'] = pnl_map.get(r['id'], 0.0)
        return jsonify(rows_to_list(rows))

    data = request.json
    cur = db.cursor()
    cur.execute("SELECT name, ticker FROM stocks WHERE id=%s", (data.get('stock_id'),))
    s = cur.fetchone()
    ticker = s['ticker'] if s else ''
    ex = get_current_exchange_rate() if is_foreign_ticker(ticker) else 1.0

    if data.get('tx_type') in ('sell', '매도'):
        held = _qty_before(db, data.get('stock_id'), data.get('tx_date'), data.get('tx_type'), 'stock_tx', 'stock_id')
        if float(data.get('quantity', 0)) > held + 1e-9:
            cur.close()
            db.close()
            return jsonify({'error': f'보유 수량({held:g})보다 많은 수량을 매도할 수 없습니다.'}), 400

    cur.execute(
    "INSERT INTO stock_tx (stock_id, tx_date, tx_type, price, quantity, fee, memo, exchange_rate, realized_pnl) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id",
    (data.get('stock_id'), data.get('tx_date'), data.get('tx_type'),
    data.get('price', 0), data.get('quantity', 0), data.get('fee', 0), data.get('memo'), ex, 0.0)
    )
    new_id = cur.fetchone()[0]
    tx_type = data.get('tx_type')
    sname = f"{s['name']}({s['ticker']})" if s and s.get('ticker') else (s['name'] if s else '주식')
    if tx_type in ('buy', '매수'):
        amt = -round((float(data.get('price', 0)) * float(data.get('quantity', 0)) + float(data.get('fee', 0))) * ex)
        _upsert_cash_adj(cur, 'stock_tx', new_id, amt, f"{sname} 매수", data.get('tx_date'))
    elif tx_type in ('sell', '매도'):
        amt = round((float(data.get('price', 0)) * float(data.get('quantity', 0)) - float(data.get('fee', 0))) * ex)
        _upsert_cash_adj(cur, 'stock_tx', new_id, amt, f"{sname} 매도", data.get('tx_date'))
    recalc_realized_pnl(db, data.get('stock_id'))
    cur.close()
    db.commit()
    db.close()
    return jsonify({'ok': True}), 201


@app.route('/api/stock-tx/<int:rid>', methods=['PUT', 'DELETE'])
def api_stock_tx_detail(rid):
    db = get_db()
    if request.method == 'PUT':
        data = request.json
        cur = db.cursor()
        cur.execute("SELECT stock_id FROM stock_tx WHERE id=%s", (rid,))
        old_row = cur.fetchone()
        old_stock_id = old_row['stock_id'] if old_row else None
        cur.execute("SELECT name, ticker FROM stocks WHERE id=%s", (data.get('stock_id'),))
        s = cur.fetchone()
        ticker = s['ticker'] if s else ''
        ex = get_current_exchange_rate() if is_foreign_ticker(ticker) else 1.0

        if data.get('tx_type') in ('sell', '매도'):
            held = _qty_before(db, data.get('stock_id'), data.get('tx_date'), data.get('tx_type'), 'stock_tx', 'stock_id', exclude_id=rid)
            if float(data.get('quantity', 0)) > held + 1e-9:
                cur.close()
                db.close()
                return jsonify({'error': f'보유 수량({held:g})보다 많은 수량을 매도할 수 없습니다.'}), 400

        cur.execute(
        "UPDATE stock_tx SET stock_id=%s, tx_date=%s, tx_type=%s, price=%s, quantity=%s, fee=%s, memo=%s, exchange_rate=%s WHERE id=%s",
        (data.get('stock_id'), data.get('tx_date'), data.get('tx_type'),
        data.get('price', 0), data.get('quantity', 0), data.get('fee', 0), data.get('memo'), ex, rid)
        )
        tx_type = data.get('tx_type')
        sname = f"{s['name']}({s['ticker']})" if s and s.get('ticker') else (s['name'] if s else '주식')
        if tx_type in ('buy', '매수'):
            amt = -round((float(data.get('price', 0)) * float(data.get('quantity', 0)) + float(data.get('fee', 0))) * ex)
            _upsert_cash_adj(cur, 'stock_tx', rid, amt, f"{sname} 매수", data.get('tx_date'))
        elif tx_type in ('sell', '매도'):
            amt = round((float(data.get('price', 0)) * float(data.get('quantity', 0)) - float(data.get('fee', 0))) * ex)
            _upsert_cash_adj(cur, 'stock_tx', rid, amt, f"{sname} 매도", data.get('tx_date'))
        else:
            _remove_cash_adj(cur, 'stock_tx', rid)
        if old_stock_id is not None and old_stock_id != data.get('stock_id'):
            recalc_realized_pnl(db, old_stock_id)
        recalc_realized_pnl(db, data.get('stock_id'))
        cur.close()
        db.commit()
        db.close()
        return jsonify({'ok': True})
    cur = db.cursor()
    cur.execute("SELECT stock_id FROM stock_tx WHERE id=%s", (rid,))
    old_row = cur.fetchone()
    old_stock_id = old_row['stock_id'] if old_row else None
    _remove_cash_adj(cur, 'stock_tx', rid)
    cur.execute("DELETE FROM stock_tx WHERE id = %s", (rid,))
    if old_stock_id is not None:
        recalc_realized_pnl(db, old_stock_id)
    cur.close()
    db.commit()
    db.close()
    return jsonify({'ok': True})


# ── API: ETF ─────────────────────────────────────────────────
@app.route('/api/etf', methods=['GET', 'POST'])
def api_etf():
    db = get_db()
    if request.method == 'GET':
        cur = db.cursor()
        cur.execute("SELECT id, name, ticker, current_price, etf_type, category, memo, COALESCE(ath,0) as ath FROM etf ORDER BY name")
        etfs = [dict(r) for r in cur.fetchall()]
        # SQL 기반 qty (대시보드·테크트리와 동일 기준, 음수 0 처리)
        cur.execute("""
            SELECT e.id,
                GREATEST(0, COALESCE(SUM(CASE WHEN t.tx_type IN ('buy','매수') THEN t.quantity ELSE 0 END), 0)
                          - COALESCE(SUM(CASE WHEN t.tx_type IN ('sell','매도') THEN t.quantity ELSE 0 END), 0)) AS qty,
                COALESCE(SUM(CASE WHEN t.tx_type IN ('buy','매수') THEN t.quantity ELSE 0 END), 0) AS buy_qty,
                COALESCE(SUM(CASE WHEN t.tx_type IN ('sell','매도') THEN t.quantity ELSE 0 END), 0) AS sell_qty
            FROM etf e LEFT JOIN etf_tx t ON t.etf_id = e.id
            GROUP BY e.id
        """)
        sql_etf = {r['id']: {'qty': float(r['qty'] or 0), 'buy_qty': float(r['buy_qty'] or 0), 'sell_qty': float(r['sell_qty'] or 0)}
                   for r in cur.fetchall()}
        cur.execute("SELECT etf_id, tx_type, price, quantity, COALESCE(fee,0) as fee, COALESCE(realized_pnl,0) as realized_pnl FROM etf_tx ORDER BY etf_id, tx_date, id")
        all_tx = cur.fetchall()
        cur.close()

        from collections import defaultdict
        tx_by_etf = defaultdict(list)
        for tx in all_tx:
            tx_by_etf[tx['etf_id']].append(tx)

        result = []
        try:
            for e in etfs:
                info     = sql_etf.get(e['id'], {'qty': 0.0, 'buy_qty': 0.0, 'sell_qty': 0.0})
                qty      = info['qty']
                _, avg, realized = calc_position(tx_by_etf[e['id']])
                avg      = avg if (avg is not None and avg == avg) else 0.0  # NaN guard
                eval_amt = round(qty * _sf(e['current_price']))
                cost_amt = round(qty * avg)
                e['quantity']       = qty
                e['avg_price']      = avg if qty > 0 else None
                e['eval_amount']    = eval_amt
                e['unrealized_pnl'] = eval_amt - cost_amt
                e['return_rate']    = round((eval_amt - cost_amt) / cost_amt * 100, 2) if cost_amt else 0
                e['realized_pnl']   = round(realized)
                e['buy_qty']        = info['buy_qty']
                e['sell_qty']       = info['sell_qty']
                result.append(e)
        except Exception as ex:
            import traceback
            db.close()
            return jsonify({'error': str(ex), 'trace': traceback.format_exc()}), 500
        db.close()
        return jsonify(result)

    data = request.json
    cur = db.cursor()
    cur.execute(
    "INSERT INTO etf (name, ticker, current_price, etf_type, category, memo) VALUES (%s,%s,%s,%s,%s,%s)",
    (data.get('name'), data.get('ticker'),
    data.get('current_price', 0), data.get('etf_type'), data.get('category'), data.get('memo'))
    )
    cur.close()
    db.commit()
    db.close()
    return jsonify({'ok': True}), 201


@app.route('/api/etf/<int:rid>', methods=['PUT', 'DELETE'])
def api_etf_detail(rid):
    db = get_db()
    if request.method == 'PUT':
        data = request.json
        cur = db.cursor()
        cur.execute(
        "UPDATE etf SET name=%s, ticker=%s, current_price=%s, etf_type=%s, category=%s, memo=%s WHERE id=%s",
        (data.get('name'), data.get('ticker'),
        data.get('current_price', 0), data.get('etf_type'), data.get('category'), data.get('memo'), rid)
        )
        cur.close()
        db.commit()
        db.close()
        return jsonify({'ok': True})
    cur = db.cursor()
    cur.execute("DELETE FROM etf_tx WHERE etf_id = %s", (rid,))
    cur.execute("DELETE FROM etf WHERE id = %s", (rid,))
    cur.close()
    db.commit()
    db.close()
    return jsonify({'ok': True})


# ── API: ETF 거래내역 ──────────────────────────────────────────
@app.route('/api/etf-tx', methods=['GET', 'POST'])
def api_etf_tx():
    db = get_db()
    if request.method == 'GET':
        etf_id = request.args.get('etf_id')
        query  = "SELECT t.*, e.name, e.ticker FROM etf_tx t LEFT JOIN etf e ON t.etf_id = e.id"
        params = []
        if etf_id:
            query += " WHERE t.etf_id = %s"
            params.append(etf_id)
        query += " ORDER BY t.tx_date DESC, t.id DESC"
        cur = db.cursor()
        cur.execute(query, params)
        rows = [dict(r) for r in cur.fetchall()]
        cur.close(); db.close()
        pnl_map = compute_realized_pnl_map(rows, 'etf_id')
        for r in rows:
            r['realized_pnl'] = pnl_map.get(r['id'], 0.0)
        return jsonify(rows_to_list(rows))

    data = request.json
    cur = db.cursor()
    cur.execute("SELECT name, ticker FROM etf WHERE id=%s", (data.get('etf_id'),))
    e = cur.fetchone()
    ticker = e['ticker'] if e else ''
    ex = get_current_exchange_rate() if is_foreign_ticker(ticker) else 1.0

    if data.get('tx_type') in ('sell', '매도'):
        held = _qty_before(db, data.get('etf_id'), data.get('tx_date'), data.get('tx_type'), 'etf_tx', 'etf_id')
        if float(data.get('quantity', 0)) > held + 1e-9:
            cur.close()
            db.close()
            return jsonify({'error': f'보유 수량({held:g})보다 많은 수량을 매도할 수 없습니다.'}), 400

    cur.execute(
    "INSERT INTO etf_tx (etf_id, tx_date, tx_type, price, quantity, fee, memo, exchange_rate, realized_pnl) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id",
    (data.get('etf_id'), data.get('tx_date'), data.get('tx_type'),
    data.get('price', 0), data.get('quantity', 0), data.get('fee', 0), data.get('memo'), ex, 0.0)
    )
    new_id = cur.fetchone()[0]
    tx_type = data.get('tx_type')
    ename = f"{e['name']}({e['ticker']})" if e and e.get('ticker') else (e['name'] if e else 'ETF')
    if tx_type in ('buy', '매수'):
        amt = -round((float(data.get('price', 0)) * float(data.get('quantity', 0)) + float(data.get('fee', 0))) * ex)
        _upsert_cash_adj(cur, 'etf_tx', new_id, amt, f"{ename} ETF 매수", data.get('tx_date'))
    elif tx_type in ('sell', '매도'):
        amt = round((float(data.get('price', 0)) * float(data.get('quantity', 0)) - float(data.get('fee', 0))) * ex)
        _upsert_cash_adj(cur, 'etf_tx', new_id, amt, f"{ename} ETF 매도", data.get('tx_date'))
    recalc_realized_pnl(db, data.get('etf_id'), is_etf=True)
    cur.close()
    db.commit(); db.close()
    return jsonify({'ok': True}), 201


@app.route('/api/etf-tx/<int:rid>', methods=['PUT', 'DELETE'])
def api_etf_tx_detail(rid):
    db = get_db()
    if request.method == 'DELETE':
        cur = db.cursor()
        cur.execute("SELECT etf_id FROM etf_tx WHERE id=%s", (rid,))
        old_row = cur.fetchone()
        old_etf_id = old_row['etf_id'] if old_row else None
        _remove_cash_adj(cur, 'etf_tx', rid)
        cur.execute("DELETE FROM etf_tx WHERE id = %s", (rid,))
        if old_etf_id is not None:
            recalc_realized_pnl(db, old_etf_id, is_etf=True)
        cur.close()
        db.commit(); db.close()
        return jsonify({'ok': True})
    data = request.json
    cur = db.cursor()
    cur.execute("SELECT etf_id FROM etf_tx WHERE id=%s", (rid,))
    old_row = cur.fetchone()
    old_etf_id = old_row['etf_id'] if old_row else None
    cur.execute("SELECT name, ticker FROM etf WHERE id=%s", (data.get('etf_id'),))
    e = cur.fetchone()
    ticker = e['ticker'] if e else ''
    ex = get_current_exchange_rate() if is_foreign_ticker(ticker) else 1.0

    if data.get('tx_type') in ('sell', '매도'):
        held = _qty_before(db, data.get('etf_id'), data.get('tx_date'), data.get('tx_type'), 'etf_tx', 'etf_id', exclude_id=rid)
        if float(data.get('quantity', 0)) > held + 1e-9:
            cur.close()
            db.close()
            return jsonify({'error': f'보유 수량({held:g})보다 많은 수량을 매도할 수 없습니다.'}), 400

    cur.execute(
    "UPDATE etf_tx SET etf_id=%s, tx_date=%s, tx_type=%s, price=%s, quantity=%s, fee=%s, memo=%s, exchange_rate=%s WHERE id=%s",
    (data.get('etf_id'), data.get('tx_date'), data.get('tx_type'),
    data.get('price', 0), data.get('quantity', 0), data.get('fee', 0), data.get('memo'), ex, rid)
    )
    tx_type = data.get('tx_type')
    ename = f"{e['name']}({e['ticker']})" if e and e.get('ticker') else (e['name'] if e else 'ETF')
    if tx_type in ('buy', '매수'):
        amt = -round((float(data.get('price', 0)) * float(data.get('quantity', 0)) + float(data.get('fee', 0))) * ex)
        _upsert_cash_adj(cur, 'etf_tx', rid, amt, f"{ename} ETF 매수", data.get('tx_date'))
    elif tx_type in ('sell', '매도'):
        amt = round((float(data.get('price', 0)) * float(data.get('quantity', 0)) - float(data.get('fee', 0))) * ex)
        _upsert_cash_adj(cur, 'etf_tx', rid, amt, f"{ename} ETF 매도", data.get('tx_date'))
    else:
        _remove_cash_adj(cur, 'etf_tx', rid)
    if old_etf_id is not None and old_etf_id != data.get('etf_id'):
        recalc_realized_pnl(db, old_etf_id, is_etf=True)
    recalc_realized_pnl(db, data.get('etf_id'), is_etf=True)
    cur.close()
    db.commit(); db.close()
    return jsonify({'ok': True})


# ── API: 코인 ────────────────────────────────────────────────
@app.route('/api/crypto', methods=['GET', 'POST'])
def api_crypto():
    db = get_db()
    if request.method == 'GET':
        cur = db.cursor()
        cur.execute("SELECT * FROM crypto ORDER BY name")
        rows = cur.fetchall()
        cur.close()
        db.close()
        return jsonify(rows_to_list(rows))

    data = request.json
    cur = db.cursor()
    cur.execute(
    "INSERT INTO crypto (name, symbol, exchange, buy_date, buy_price, quantity, current_price, memo) VALUES (%s,%s,%s,%s,%s,%s,%s,%s)",
    (data.get('name'), data.get('symbol'), data.get('exchange'), data.get('buy_date'),
    data.get('buy_price', 0), data.get('quantity', 0),
    data.get('current_price', 0), data.get('memo'))
    )
    cur.close()
    db.commit()
    db.close()
    return jsonify({'ok': True}), 201


@app.route('/api/crypto/<int:rid>', methods=['PUT', 'DELETE'])
def api_crypto_detail(rid):
    db = get_db()
    if request.method == 'PUT':
        data = request.json
        cur = db.cursor()
        cur.execute(
        "UPDATE crypto SET name=%s, symbol=%s, exchange=%s, buy_date=%s, buy_price=%s, quantity=%s, current_price=%s, memo=%s WHERE id=%s",
        (data.get('name'), data.get('symbol'), data.get('exchange'), data.get('buy_date'),
        data.get('buy_price', 0), data.get('quantity', 0),
        data.get('current_price', 0), data.get('memo'), rid)
        )
        cur.close()
        db.commit()
        db.close()
        return jsonify({'ok': True})
    cur = db.cursor()
    cur.execute("DELETE FROM crypto WHERE id = %s", (rid,))
    cur.close()
    db.commit()
    db.close()
    return jsonify({'ok': True})


# ── API: 코인 매도기록 ───────────────────────────────────────
def _ensure_crypto_sell_table(db):
    cur = db.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS crypto_sell (
            id SERIAL PRIMARY KEY, sell_date DATE NOT NULL,
            name TEXT NOT NULL, pnl INTEGER DEFAULT 0,
            memo TEXT, created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    db.commit()
    cur.close()

@app.route('/api/crypto-sell', methods=['GET', 'POST'])
def api_crypto_sell():
    db = get_db()
    _ensure_crypto_sell_table(db)
    if request.method == 'GET':
        cur = db.cursor()
        cur.execute("SELECT * FROM crypto_sell ORDER BY sell_date DESC, id DESC")
        rows = rows_to_list(cur.fetchall())
        cur.close()
        db.close()
        return jsonify(rows)
    data = request.json
    cur = db.cursor()
    cur.execute(
        "INSERT INTO crypto_sell (sell_date, name, pnl, memo) VALUES (%s, %s, %s, %s)",
        (data.get('sell_date'), data.get('name'), data.get('pnl', 0), data.get('memo'))
    )
    cur.close()
    db.commit()
    db.close()
    return jsonify({'ok': True}), 201

@app.route('/api/crypto-sell/<int:rid>', methods=['PUT', 'DELETE'])
def api_crypto_sell_detail(rid):
    db = get_db()
    _ensure_crypto_sell_table(db)
    if request.method == 'PUT':
        data = request.json
        cur = db.cursor()
        cur.execute(
            "UPDATE crypto_sell SET sell_date=%s, name=%s, pnl=%s, memo=%s WHERE id=%s",
            (data.get('sell_date'), data.get('name'), data.get('pnl', 0), data.get('memo'), rid)
        )
        cur.close()
        db.commit()
        db.close()
        return jsonify({'ok': True})
    cur = db.cursor()
    cur.execute("DELETE FROM crypto_sell WHERE id = %s", (rid,))
    cur.close()
    db.commit()
    db.close()
    return jsonify({'ok': True})


# ── API: 공모주 ──────────────────────────────────────────────
@cache.cached(timeout=180)
def _get_ipo_cached():
    db = get_db()
    cur = db.cursor()
    cur.execute("SELECT * FROM ipo ORDER BY listing_date DESC, id DESC")
    rows = cur.fetchall()
    cur.close()
    db.close()
    return rows_to_list(rows)

@app.route('/api/ipo', methods=['GET', 'POST'])
def api_ipo():
    if request.method == 'GET':
        return jsonify(_get_ipo_cached())

    db = get_db()
    data = request.json
    cur = db.cursor()
    cur.execute(
        "INSERT INTO ipo (name, listing_date, ipo_price, quantity, realized_pnl, fee, memo, lockup_ratio, floating_ratio) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)",
        (data.get('name'), data.get('listing_date'), data.get('ipo_price', 0),
         data.get('quantity', 0), data.get('realized_pnl', 0), data.get('fee', 0), data.get('memo'),
         data.get('lockup_ratio', 0), data.get('floating_ratio', 0))
    )
    cur.close()
    db.commit()
    db.close()
    return jsonify({'ok': True}), 201


@app.route('/api/ipo/<int:rid>', methods=['PUT', 'DELETE'])
def api_ipo_detail(rid):
    db = get_db()
    if request.method == 'PUT':
        data = request.json
        cur = db.cursor()
        cur.execute(
            "UPDATE ipo SET name=%s, listing_date=%s, ipo_price=%s, quantity=%s, realized_pnl=%s, fee=%s, memo=%s, lockup_ratio=%s, floating_ratio=%s WHERE id=%s",
            (data.get('name'), data.get('listing_date'), data.get('ipo_price', 0),
             data.get('quantity', 0), data.get('realized_pnl', 0), data.get('fee', 0), data.get('memo'),
             data.get('lockup_ratio', 0), data.get('floating_ratio', 0), rid)
        )
        cur.close()
        db.commit()
        db.close()
        return jsonify({'ok': True})
    cur = db.cursor()
    cur.execute("DELETE FROM ipo WHERE id = %s", (rid,))
    cur.close()
    db.commit()
    db.close()
    return jsonify({'ok': True})


# ── API: 종목별 투자계획 ──────────────────────────────────────
@cache.cached(timeout=180)
def _get_split_buy_plans_cached():
    db = get_db()
    cur = db.cursor()
    cur.execute("SELECT * FROM split_buy_plans ORDER BY id ASC")
    rows = cur.fetchall()
    cur.close()
    db.close()
    return rows_to_list(rows)

def _calc_invest_plan_steps(target_price, upper_pct, lower_pct,
                             split_count, total_budget, strategy, usd_krw=None):
    """
    상단 매수가 기준으로 분할매수 차수를 계산.

    예: target=60,000 / upper=0% / lower=20% / 5차
      1차: 60,000원 (0%)
      2차: 57,000원 (-5%)
      3차: 54,000원 (-10%)
      4차: 51,000원 (-15%)
      5차: 48,000원 (-20%)

    전략별 배분:
      equal:           [20%, 20%, 20%, 20%, 20%]
      inverse_pyramid: 하락할수록 비중↑ (역피라미딩)
      pyramid:         하락할수록 비중↓ (정피라미딩 / 추세추종)
    """
    total_range = upper_pct + lower_pct   # 전체 범위 %
    step_pct    = total_range / (split_count - 1) if split_count > 1 else 0

    # 각 차수의 기준가 대비 하락률
    pct_points = [upper_pct - step_pct * i for i in range(split_count)]

    # 배분 가중치 결정
    if strategy == 'equal':
        weights = [1.0 / split_count] * split_count

    elif strategy == 'inverse_pyramid':
        # 1.0, 1.5, 2.25, 3.375, 5.0625 ... (1.5배씩 증가)
        raw = [1.0 * (1.5 ** i) for i in range(split_count)]
        total_w = sum(raw)
        weights = [w / total_w for w in raw]

    elif strategy == 'pyramid':
        raw = [1.0 * (1.5 ** i) for i in range(split_count - 1, -1, -1)]
        total_w = sum(raw)
        weights = [w / total_w for w in raw]

    else:
        weights = [1.0 / split_count] * split_count

    is_foreign = usd_krw and usd_krw > 1
    steps = []
    cumulative = 0
    for i, (pct, weight) in enumerate(zip(pct_points, weights)):
        raw_trigger   = target_price * (1 + pct / 100)
        trigger_price = round(raw_trigger, 2) if is_foreign else round(raw_trigger)
        amount        = round(total_budget * weight)
        if is_foreign:
            shares = round(amount / (trigger_price * usd_krw), 4) if trigger_price > 0 else 0
        else:
            shares = round(amount / trigger_price, 4) if trigger_price > 0 else 0
        cumulative   += amount

        steps.append({
            'step_no':       i + 1,
            'pct_from_target': round(pct, 1),
            'trigger_price': trigger_price,
            'weight_pct':    round(weight * 100, 1),
            'amount':        amount,
            'shares':        shares,
            'cumulative':    cumulative,
            'label':         (f"+{pct:.1f}%" if pct > 0
                              else f"{pct:.1f}%" if pct < 0
                              else "상단 매수가"),
        })

    return steps


@app.route('/api/rebalance', methods=['GET'])
def api_rebalance_get():
    """올웨더 자산 배분 현황 조회"""
    db = get_db()
    cur = db.cursor()

    # 저장된 배분 설정 조회
    cur.execute("SELECT * FROM rebalance_assignments")
    assignments = {(r['source_type'], r['source_id']): dict(r) for r in cur.fetchall()}

    # 올웨더 ETF 조회 (현재 평가액 포함)
    cur.execute("""
        SELECT e.id, e.name, e.ticker, e.current_price, e.category,
               GREATEST(0,
                 COALESCE(SUM(CASE WHEN t.tx_type IN ('buy','매수') THEN t.quantity ELSE 0 END),0)
               - COALESCE(SUM(CASE WHEN t.tx_type IN ('sell','매도') THEN t.quantity ELSE 0 END),0)
               ) AS qty
        FROM etf e
        LEFT JOIN etf_tx t ON t.etf_id = e.id
        WHERE LOWER(e.category) LIKE '%올웨더%'
        GROUP BY e.id
    """)
    etfs = rows_to_list(cur.fetchall())

    # 올웨더 주식 조회
    cur.execute("""
        SELECT s.id, s.name, s.ticker, s.current_price, s.category,
               GREATEST(0,
                 COALESCE(SUM(CASE WHEN t.tx_type IN ('buy','매수') THEN t.quantity ELSE 0 END),0)
               - COALESCE(SUM(CASE WHEN t.tx_type IN ('sell','매도') THEN t.quantity ELSE 0 END),0)
               ) AS qty
        FROM stocks s
        LEFT JOIN stock_tx t ON t.stock_id = s.id
        WHERE LOWER(s.category) LIKE '%올웨더%'
        GROUP BY s.id
    """)
    stocks = rows_to_list(cur.fetchall())
    cur.close()
    db.close()

    # 현재가 × 수량 = 평가액 계산
    result_items = []
    for item in etfs:
        key = ('etf', item['id'])
        asgn = assignments.get(key, {})
        eval_amt = round(float(item['qty'] or 0) * float(item['current_price'] or 0))
        result_items.append({
            'source_type': 'etf', 'source_id': item['id'],
            'name': item['name'], 'ticker': item['ticker'],
            'eval_amount': eval_amt,
            'asset_class': asgn.get('asset_class', ''),
        })
    for item in stocks:
        key = ('stock', item['id'])
        asgn = assignments.get(key, {})
        eval_amt = round(float(item['qty'] or 0) * float(item['current_price'] or 0))
        result_items.append({
            'source_type': 'stock', 'source_id': item['id'],
            'name': item['name'], 'ticker': item['ticker'],
            'eval_amount': eval_amt,
            'asset_class': asgn.get('asset_class', ''),
        })

    # 현금 항목
    cash_row = assignments.get(('cash', 0), {})
    cash_amount = int(cash_row.get('cash_amount', 0) or 0)

    return jsonify({'items': result_items, 'cash': cash_amount})


@app.route('/api/rebalance', methods=['POST'])
def api_rebalance_save():
    """올웨더 자산 배분 설정 저장"""
    db = get_db()
    d = request.json or {}
    assignments = d.get('assignments', [])  # [{source_type, source_id, asset_class}]
    cash = int(d.get('cash', 0) or 0)
    cur = db.cursor()
    for a in assignments:
        cur.execute("""
            INSERT INTO rebalance_assignments (source_type, source_id, asset_class)
            VALUES (%s, %s, %s)
            ON CONFLICT (source_type, source_id)
            DO UPDATE SET asset_class = EXCLUDED.asset_class
        """, (a['source_type'], a['source_id'], a['asset_class']))
    # 현금 저장
    cur.execute("""
        INSERT INTO rebalance_assignments (source_type, source_id, asset_class, cash_amount)
        VALUES ('cash', 0, 'cash', %s)
        ON CONFLICT (source_type, source_id)
        DO UPDATE SET asset_class='cash', cash_amount=EXCLUDED.cash_amount
    """, (cash,))
    cur.close()
    db.commit(); db.close()
    return jsonify({'ok': True})


@app.route('/api/invest-plans', methods=['GET', 'POST'])
def api_invest_plans():
    db = get_db()
    if request.method == 'GET':
        stock_id = request.args.get('stock_id')
        etf_id   = request.args.get('etf_id')
        cur = db.cursor()
        where = ""
        params = []
        if stock_id:
            where = "WHERE p.stock_id = %s"
            params = [stock_id]
        elif etf_id:
            where = "WHERE p.etf_id = %s"
            params = [etf_id]
        cur.execute(f"""
            SELECT p.*,
                   s.name AS stock_name, s.ticker AS stock_ticker,
                   e.name AS etf_name,   e.ticker AS etf_ticker,
                   s.current_price AS stock_current_price,
                   e.current_price AS etf_current_price
            FROM invest_plans p
            LEFT JOIN stocks s ON p.stock_id = s.id
            LEFT JOIN etf    e ON p.etf_id   = e.id
            {where}
            ORDER BY p.created_at DESC
        """, params)
        rows = cur.fetchall(); cur.close()

        result = []
        for row in rows:
            r = dict(row)
            cur = db.cursor()
            cur.execute("""
                SELECT * FROM invest_plan_steps
                WHERE plan_id = %s ORDER BY step_no
            """, (row['id'],))
            r['steps'] = rows_to_list(cur.fetchall())
            cur.close()

            # 체결된 차수 수 / 총 차수
            r['executed_count']  = sum(1 for s in r['steps'] if s['is_executed'])
            r['total_steps']     = len(r['steps'])
            r['executed_amount'] = sum(s['executed_amount'] or 0
                                       for s in r['steps'] if s['is_executed'])
            # 현재가
            r['current_price'] = _sf(row['stock_current_price'] or row['etf_current_price'])
            result.append(r)

        db.close()
        return jsonify(result)

    # ── POST: 계획 생성 + steps 자동 계산 ──
    d = request.json or {}
    target_price = float(d.get('target_price', 0))
    upper_pct    = float(d.get('upper_pct',  0))
    lower_pct    = float(d.get('lower_pct',  20))
    split_count  = int(d.get('split_count',  5))
    total_budget = int(d.get('total_budget', 0))
    strategy     = d.get('strategy', 'inverse_pyramid')
    usd_krw      = float(d['usd_krw']) if d.get('usd_krw') else None
    _preview_only = d.get('_preview_only', False)

    if not target_price or not total_budget:
        db.close()
        return jsonify({'error': '상단 매수가와 총 예산은 필수입니다.'}), 400

    steps_data = _calc_invest_plan_steps(
        target_price, upper_pct, lower_pct,
        split_count, total_budget, strategy, usd_krw
    )

    if _preview_only:
        db.close()
        return jsonify({'ok': True, 'steps': steps_data}), 200

    cur = db.cursor()
    cur.execute("""
        INSERT INTO invest_plans
            (stock_id, etf_id, plan_name, target_price, upper_pct, lower_pct,
             split_count, total_budget, strategy, memo)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id
    """, (d.get('stock_id') or None, d.get('etf_id') or None,
          d.get('plan_name', ''), target_price, upper_pct, lower_pct,
          split_count, total_budget, strategy, d.get('memo', '')))
    plan_id = cur.fetchone()[0]
    cur.close()

    for step in steps_data:
        cur = db.cursor()
        cur.execute("""
            INSERT INTO invest_plan_steps
                (plan_id, step_no, trigger_price, target_amount, target_shares, weight_pct)
            VALUES (%s,%s,%s,%s,%s,%s)
        """, (plan_id, step['step_no'], step['trigger_price'],
              step['amount'], step['shares'], step['weight_pct']))
        cur.close()

    db.commit(); db.close()
    return jsonify({'ok': True, 'plan_id': plan_id, 'steps': steps_data}), 201


@app.route('/api/invest-plans/<int:plan_id>', methods=['PUT', 'DELETE'])
def api_invest_plan_detail(plan_id):
    db = get_db()
    if request.method == 'DELETE':
        cur = db.cursor()
        cur.execute("DELETE FROM invest_plans WHERE id=%s", (plan_id,))
        cur.close(); db.commit(); db.close()
        return jsonify({'ok': True})

    # PUT: 계획 수정 (steps 재계산)
    d = request.json or {}
    target_price = float(d.get('target_price', 0))
    upper_pct    = float(d.get('upper_pct',  0))
    lower_pct    = float(d.get('lower_pct',  20))
    split_count  = int(d.get('split_count',  5))
    total_budget = int(d.get('total_budget', 0))
    strategy     = d.get('strategy', 'inverse_pyramid')
    usd_krw      = float(d['usd_krw']) if d.get('usd_krw') else None

    if not target_price or not total_budget:
        db.close()
        return jsonify({'error': '상단 매수가와 총 예산은 필수입니다.'}), 400

    steps_data = _calc_invest_plan_steps(
        target_price, upper_pct, lower_pct, split_count, total_budget, strategy, usd_krw
    )

    cur = db.cursor()
    cur.execute("""
        UPDATE invest_plans
        SET stock_id=%s, etf_id=%s, plan_name=%s, target_price=%s,
            upper_pct=%s, lower_pct=%s, split_count=%s,
            total_budget=%s, strategy=%s, memo=%s, updated_at=CURRENT_TIMESTAMP
        WHERE id=%s
    """, (d.get('stock_id') or None, d.get('etf_id') or None,
          d.get('plan_name', ''), target_price, upper_pct, lower_pct,
          split_count, total_budget, strategy, d.get('memo', ''), plan_id))
    cur.close()

    # 기존 steps 삭제 후 재생성
    cur = db.cursor()
    cur.execute("DELETE FROM invest_plan_steps WHERE plan_id=%s", (plan_id,))
    cur.close()

    for step in steps_data:
        cur = db.cursor()
        cur.execute("""
            INSERT INTO invest_plan_steps
                (plan_id, step_no, trigger_price, target_amount, target_shares, weight_pct)
            VALUES (%s,%s,%s,%s,%s,%s)
        """, (plan_id, step['step_no'], step['trigger_price'],
              step['amount'], step['shares'], step['weight_pct']))
        cur.close()

    db.commit(); db.close()
    return jsonify({'ok': True, 'steps': steps_data})


@app.route('/api/invest-plans/<int:plan_id>/steps', methods=['GET'])
def api_invest_plan_steps(plan_id):
    db = get_db()
    cur = db.cursor()
    cur.execute("SELECT * FROM invest_plan_steps WHERE plan_id=%s ORDER BY step_no", (plan_id,))
    rows = rows_to_list(cur.fetchall())
    cur.close(); db.close()
    return jsonify(rows)


@app.route('/api/invest-plan-steps/<int:step_id>/execute', methods=['POST'])
def api_invest_plan_step_execute(step_id):
    """특정 차수 매수 체결 기록"""
    try:
        d = request.json or {}
        db = get_db()
        cur = db.cursor()
        cur.execute("""
            UPDATE invest_plan_steps
            SET is_executed=TRUE,
                executed_at=%s, executed_price=%s,
                executed_shares=%s, executed_amount=%s
            WHERE id=%s
        """, (d.get('executed_at', date.today().isoformat()),
              d.get('executed_price'), d.get('executed_shares'),
              d.get('executed_amount'), step_id))
        cur.close(); db.commit(); db.close()
        return jsonify({'ok': True})
    except Exception as e:
        import traceback
        print(f"[execute step] {e}\n{traceback.format_exc()}")
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/invest-plan-steps/<int:step_id>/execute', methods=['DELETE'])
def api_invest_plan_step_unexecute(step_id):
    """체결 기록 취소"""
    try:
        db = get_db()
        cur = db.cursor()
        cur.execute("""
            UPDATE invest_plan_steps
            SET is_executed=FALSE, executed_at=NULL,
                executed_price=NULL, executed_shares=NULL, executed_amount=NULL
            WHERE id=%s
        """, (step_id,))
        cur.close(); db.commit(); db.close()
        return jsonify({'ok': True})
    except Exception as e:
        import traceback
        print(f"[unexecute step] {e}\n{traceback.format_exc()}")
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/split-buy-plans', methods=['GET', 'POST'])
def api_split_buy_plans():
    if request.method == 'GET':
        return jsonify(_get_split_buy_plans_cached())

    db = get_db()

    # POST
    data = request.json
    name = data.get('name')
    ticker = data.get('ticker')
    total_budget = int(data.get('total_budget', 0))
    ath = float(data.get('ath', 0))
    current_price = data.get('current_price')
    if current_price is not None and current_price != '':
        current_price = float(current_price)
    else:
        current_price = None
    drop_from  = float(data.get('drop_from', 30))
    drop_to    = float(data.get('drop_to', 70))
    step_count = int(data.get('step_count', 5))

    if current_price and current_price > 0 and ath <= 0:
        ath = current_price
    if not name or ath <= 0:
        db.close()
        return jsonify({'error': '종목명과 현재가(또는 ATH) 중 하나는 필수입니다.'}), 400

    cur = db.cursor()
    cur.execute(
        "INSERT INTO split_buy_plans (name, ticker, total_budget, ath, current_price, drop_from, drop_to, step_count) VALUES (%s, %s, %s, %s, %s, %s, %s, %s) RETURNING id",
        (name, ticker, total_budget, ath, current_price, drop_from, drop_to, step_count)
    )
    plan_id = cur.fetchone()['id']

    # steps가 요청에 있으면 저장, 없으면 범위 기반 균등 분배
    steps = data.get('steps')
    if steps:
        for s in steps:
            cur.execute(
                "INSERT INTO split_buy_plan_steps (plan_id, step_number, drawdown_pct, ratio) VALUES (%s, %s, %s, %s)",
                (plan_id, int(s.get('step_number')), float(s.get('drawdown_pct')), float(s.get('ratio')))
            )
    else:
        n = max(1, step_count)
        ratio = round(100.0 / n, 6)
        for i in range(n):
            if n == 1:
                drop = drop_from
            else:
                drop = drop_from + (drop_to - drop_from) * i / (n - 1)
            cur.execute(
                "INSERT INTO split_buy_plan_steps (plan_id, step_number, drawdown_pct, ratio) VALUES (%s, %s, %s, %s)",
                (plan_id, i + 1, round(drop, 2), ratio)
            )

    db.commit()
    cur.close()
    db.close()
    return jsonify({'ok': True, 'id': plan_id}), 201


@app.route('/api/split-buy-plans/<int:rid>', methods=['PUT', 'DELETE'])
def api_split_buy_plan_detail(rid):
    db = get_db()
    if request.method == 'PUT':
        data = request.json
        name = data.get('name')
        ticker = data.get('ticker')
        total_budget = int(data.get('total_budget', 0))
        ath = float(data.get('ath', 0))
        current_price = data.get('current_price')
        if current_price is not None and current_price != '':
            current_price = float(current_price)
        else:
            current_price = None

        drop_from  = float(data.get('drop_from', 30))
        drop_to    = float(data.get('drop_to', 70))
        step_count = int(data.get('step_count', 5))

        if not name or ath <= 0:
            db.close()
            return jsonify({'error': '종목명과 최고가(ATH)는 필수 입력 사항입니다.'}), 400

        cur = db.cursor()
        cur.execute(
            "UPDATE split_buy_plans SET name=%s, ticker=%s, total_budget=%s, ath=%s, current_price=%s, drop_from=%s, drop_to=%s, step_count=%s WHERE id=%s",
            (name, ticker, total_budget, ath, current_price, drop_from, drop_to, step_count, rid)
        )
        db.commit()
        cur.close()
        db.close()
        return jsonify({'ok': True})

    # DELETE
    cur = db.cursor()
    cur.execute("DELETE FROM split_buy_plans WHERE id = %s", (rid,))
    db.commit()
    cur.close()
    db.close()
    return jsonify({'ok': True})


@cache.memoize(timeout=180)
def _get_split_buy_plan_steps_cached(pid):
    db = get_db()
    cur = db.cursor()
    cur.execute("SELECT * FROM split_buy_plan_steps WHERE plan_id = %s ORDER BY step_number ASC", (pid,))
    rows = cur.fetchall()
    cur.close()
    db.close()
    return rows_to_list(rows)

@app.route('/api/split-buy-plans/<int:pid>/steps', methods=['GET', 'POST'])
def api_split_buy_plan_steps(pid):
    if request.method == 'GET':
        return jsonify(_get_split_buy_plan_steps_cached(pid))

    db = get_db()

    # POST
    steps = request.json
    if not isinstance(steps, list):
        db.close()
        return jsonify({'error': '데이터가 리스트 형식이 아닙니다.'}), 400

    cur = db.cursor()
    # 기존 steps 삭제 후 재생성
    cur.execute("DELETE FROM split_buy_plan_steps WHERE plan_id = %s", (pid,))
    for s in steps:
        cur.execute(
            "INSERT INTO split_buy_plan_steps (plan_id, step_number, drawdown_pct, ratio) VALUES (%s, %s, %s, %s)",
            (pid, int(s.get('step_number')), float(s.get('drawdown_pct')), float(s.get('ratio')))
        )
    db.commit()
    cur.close()
    db.close()
    return jsonify({'ok': True})


@app.route('/api/split-buy-plans/<int:pid>/tx', methods=['GET', 'POST'])
def api_split_buy_plan_tx(pid):
    db = get_db()
    if request.method == 'GET':
        cur = db.cursor()
        cur.execute("SELECT * FROM split_buy_transactions WHERE plan_id = %s ORDER BY tx_date ASC, id ASC", (pid,))
        rows = cur.fetchall()
        cur.close()
        db.close()
        return jsonify(rows_to_list(rows))

    # POST
    data = request.json
    tx_type = data.get('tx_type')
    step_number = data.get('step_number')
    if step_number is not None and step_number != '':
        step_number = int(step_number)
    else:
        step_number = None
    price = float(data.get('price', 0))
    quantity = float(data.get('quantity', 0))
    tx_date = data.get('tx_date')
    memo = data.get('memo')

    if tx_type not in ['buy', 'sell'] or price <= 0 or quantity <= 0 or not tx_date:
        db.close()
        return jsonify({'error': '올바르지 않은 거래 데이터입니다.'}), 400

    cur = db.cursor()
    cur.execute(
        "INSERT INTO split_buy_transactions (plan_id, tx_type, step_number, price, quantity, tx_date, memo) VALUES (%s, %s, %s, %s, %s, %s, %s) RETURNING id",
        (pid, tx_type, step_number, price, quantity, tx_date, memo)
    )
    tx_id = cur.fetchone()['id']
    db.commit()
    cur.close()
    db.close()
    return jsonify({'ok': True, 'id': tx_id}), 201


@app.route('/api/split-buy-plans/<int:pid>/tx/<int:tx_id>', methods=['DELETE'])
def api_split_buy_plan_tx_delete(pid, tx_id):
    db = get_db()
    cur = db.cursor()
    cur.execute("DELETE FROM split_buy_transactions WHERE plan_id = %s AND id = %s", (pid, tx_id))
    db.commit()
    cur.close()
    db.close()
    return jsonify({'ok': True})


@app.route('/api/split-buy-plans/<int:pid>/refresh', methods=['POST'])
def api_split_buy_plan_refresh(pid):
    """현재가 및 역대 최고가(ATH)를 자동으로 조회하여 업데이트"""
    db = get_db()
    cur = db.cursor()
    cur.execute("SELECT ticker, name FROM split_buy_plans WHERE id = %s", (pid,))
    plan = cur.fetchone()
    if not plan:
        cur.close(); db.close()
        return jsonify({'error': '계획을 찾을 수 없습니다.'}), 404

    ticker = (plan['ticker'] or '').strip()
    if not ticker:
        cur.close(); db.close()
        return jsonify({'error': '티커가 없어 가격을 조회할 수 없습니다.'}), 400

    # ── 현재가 조회 (기존 price-update와 동일한 소스) ──
    current_price = None
    try:
        current_price = _fetch_stock_price(ticker)
    except Exception as e:
        print(f'[refresh] price error {ticker}: {e}', file=sys.stderr)

    # ── 역사상 최고가(ATH) 조회 — Yahoo Finance 전체 이력 월봉 최대 고가 ──
    ath = _fetch_all_time_high(ticker)

    # ── DB 업데이트 ──
    if current_price or ath:
        fields, vals = [], []
        if current_price:
            fields.append("current_price=%s"); vals.append(current_price)
        if ath:
            fields.append("ath=%s"); vals.append(ath)
        vals.append(pid)
        cur.execute(f"UPDATE split_buy_plans SET {', '.join(fields)} WHERE id=%s", vals)
        db.commit()

    cur.close(); db.close()
    return jsonify({
        'ok': True,
        'current_price': current_price,
        'ath': ath,
        'price_ok': current_price is not None,
        'ath_ok': ath is not None,
    })


@app.route('/api/split-buy-plans/<int:pid>/linked-tx', methods=['GET', 'POST'])
def api_split_buy_plan_linked_tx(pid):
    db = get_db()
    cur = db.cursor()
    cur.execute("SELECT ticker FROM split_buy_plans WHERE id = %s", (pid,))
    plan = cur.fetchone()
    if not plan:
        cur.close(); db.close()
        return jsonify({'error': 'Plan not found'}), 404
    ticker = (plan['ticker'] or '').strip()

    if request.method == 'GET':
        if not ticker:
            cur.close(); db.close()
            return jsonify([])
        cur.execute("""
            SELECT t.id, t.tx_date::text AS tx_date, t.tx_type,
                   CAST(t.price AS float) AS price,
                   CAST(t.quantity AS float) AS quantity,
                   CAST(COALESCE(t.fee, 0) AS float) AS fee,
                   t.memo, 'stock' AS source, s.id AS source_id
            FROM stock_tx t
            JOIN stocks s ON t.stock_id = s.id
            WHERE UPPER(s.ticker) = UPPER(%s)
        """, (ticker,))
        stock_txs = rows_to_list(cur.fetchall())
        cur.execute("""
            SELECT t.id, t.tx_date::text AS tx_date, t.tx_type,
                   CAST(t.price AS float) AS price,
                   CAST(t.quantity AS float) AS quantity,
                   CAST(COALESCE(t.fee, 0) AS float) AS fee,
                   t.memo, 'etf' AS source, e.id AS source_id
            FROM etf_tx t
            JOIN etf e ON t.etf_id = e.id
            WHERE UPPER(e.ticker) = UPPER(%s)
        """, (ticker,))
        etf_txs = rows_to_list(cur.fetchall())
        cur.close(); db.close()
        all_txs = sorted(stock_txs + etf_txs, key=lambda x: (x.get('tx_date') or '', x.get('id', 0)))
        return jsonify(all_txs)

    # POST — add transaction directly to stock_tx or etf_tx
    data = request.json
    tx_type  = data.get('tx_type')
    price    = float(data.get('price', 0))
    quantity = float(data.get('quantity', 0))
    tx_date  = data.get('tx_date')
    memo     = data.get('memo')
    fee      = float(data.get('fee', 0))

    if tx_type not in ['buy', 'sell'] or price <= 0 or quantity <= 0 or not tx_date:
        cur.close(); db.close()
        return jsonify({'error': '올바르지 않은 거래 데이터입니다.'}), 400
    if not ticker:
        cur.close(); db.close()
        return jsonify({'error': '이 계획에 티커가 설정되지 않았습니다.'}), 400

    # stocks 테이블 먼저 확인
    cur.execute("SELECT id, name, ticker FROM stocks WHERE UPPER(ticker) = UPPER(%s) LIMIT 1", (ticker,))
    stock = cur.fetchone()
    if stock:
        ex = get_current_exchange_rate() if is_foreign_ticker(stock['ticker']) else 1.0
        cur.execute(
            "INSERT INTO stock_tx (stock_id, tx_date, tx_type, price, quantity, fee, memo, exchange_rate) VALUES (%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id",
            (stock['id'], tx_date, tx_type, price, quantity, fee, memo, ex)
        )
        new_id = cur.fetchone()['id']
        if tx_type == 'buy':
            sname = f"{stock['name']}({stock['ticker']})" if stock.get('ticker') else stock['name']
            amt = -round((price * quantity + fee) * ex)
            _upsert_cash_adj(cur, 'stock_tx', new_id, amt, f"{sname} 매수", tx_date)
        db.commit(); cur.close(); db.close()
        return jsonify({'ok': True, 'id': new_id, 'source': 'stock'}), 201

    # ETF 테이블 확인
    cur.execute("SELECT id, name, ticker FROM etf WHERE UPPER(ticker) = UPPER(%s) LIMIT 1", (ticker,))
    etf = cur.fetchone()
    if etf:
        ex = get_current_exchange_rate() if is_foreign_ticker(etf['ticker']) else 1.0
        cur.execute(
            "INSERT INTO etf_tx (etf_id, tx_date, tx_type, price, quantity, fee, memo, exchange_rate) VALUES (%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id",
            (etf['id'], tx_date, tx_type, price, quantity, fee, memo, ex)
        )
        new_id = cur.fetchone()['id']
        if tx_type == 'buy':
            ename = f"{etf['name']}({etf['ticker']})" if etf.get('ticker') else etf['name']
            amt = -round((price * quantity + fee) * ex)
            _upsert_cash_adj(cur, 'etf_tx', new_id, amt, f"{ename} ETF 매수", tx_date)
        db.commit(); cur.close(); db.close()
        return jsonify({'ok': True, 'id': new_id, 'source': 'etf'}), 201

    cur.close(); db.close()
    return jsonify({'error': f'티커 {ticker}에 해당하는 종목이 없습니다. 주식 또는 ETF 탭에서 먼저 종목을 추가해주세요.'}), 400


@app.route('/api/split-buy-plans/<int:pid>/linked-tx/<string:source>/<int:tx_id>', methods=['DELETE'])
def api_split_buy_plan_linked_tx_delete(pid, source, tx_id):
    db = get_db()
    cur = db.cursor()
    if source == 'stock':
        _remove_cash_adj(cur, 'stock_tx', tx_id)
        cur.execute("DELETE FROM stock_tx WHERE id = %s", (tx_id,))
    elif source == 'etf':
        _remove_cash_adj(cur, 'etf_tx', tx_id)
        cur.execute("DELETE FROM etf_tx WHERE id = %s", (tx_id,))
    else:
        cur.close(); db.close()
        return jsonify({'error': 'Invalid source'}), 400
    db.commit(); cur.close(); db.close()
    return jsonify({'ok': True})



# ── API: 현재가 업데이트 ─────────────────────────────────────
import time

def _is_krx_ticker(ticker: str) -> bool:
    """6자리 숫자면 국내 KRX 종목으로 판단"""
    return bool(re.match(r'^\d{6}$', ticker))


def _fetch_price_and_52w_high(ticker: str) -> tuple:
    """Yahoo Finance API 한 번 호출로 현재가 + 52주 최고가 동시 조회.
    국내 종목은 .KS → .KQ 순으로 시도. 반환: (current_price, high_52w) — 실패 시 None"""
    ticker = (ticker or '').strip()
    if not ticker:
        return None, None
    if not _is_krx_ticker(ticker):
        ticker = ticker.upper()
    syms = ([ticker + s for s in ['.KS', '.KQ']] if _is_krx_ticker(ticker) else [ticker])
    for sym in syms:
        try:
            res = http_req.get(
                f'https://query2.finance.yahoo.com/v8/finance/chart/{sym}',
                params={'interval': '1d', 'range': '5d'},
                headers={'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36', 'Accept': 'application/json'},
                timeout=8
            )
            if not res.ok:
                continue
            result = res.json().get('chart', {}).get('result', [])
            if not result:
                continue
            meta = result[0].get('meta', {})
            price = meta.get('regularMarketPrice')
            high52 = meta.get('fiftyTwoWeekHigh')
            if price:
                return float(price), (float(high52) if high52 else None)
        except Exception as e:
            print(f'[price_high] yf error {sym}: {e}', file=sys.stderr)
    return None, None


import sys

def _fetch_all_time_high(ticker: str) -> float | None:
    """Yahoo Finance 전체 이력 월봉으로 역사상 최고가(ATH) 조회."""
    ticker = (ticker or '').strip()
    if not ticker:
        return None
    if not _is_krx_ticker(ticker):
        ticker = ticker.upper()
    time.sleep(0.5)  # rate limit 방지
    syms = ([ticker + s for s in ['.KS', '.KQ']] if _is_krx_ticker(ticker) else [ticker])
    for sym in syms:
        try:
            res = http_req.get(
                f'https://query2.finance.yahoo.com/v8/finance/chart/{sym}',
                params={'interval': '1mo', 'range': 'max'},
                headers={'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36', 'Accept': 'application/json'},
                timeout=15
            )
            if not res.ok:
                continue
            result = res.json().get('chart', {}).get('result', [])
            if not result:
                continue
            highs = result[0].get('indicators', {}).get('quote', [{}])[0].get('high', [])
            highs = [h for h in highs if h is not None and not (isinstance(h, float) and math.isnan(h))]
            if highs:
                return float(max(highs))
        except Exception as e:
            print(f'[ath] yf error {sym}: {e}', file=sys.stderr)
    return None


def _fetch_alphavantage_price(ticker: str) -> float | None:
    """Alpha Vantage API로 현재가 조회 (env: ALPHAVANTAGE_API_KEY 필요)"""
    ticker = (ticker or '').strip()
    if not ticker:
        return None
    if not _is_krx_ticker(ticker):
        ticker = ticker.upper()
    api_key = os.environ.get('ALPHAVANTAGE_API_KEY', '').strip()
    if not api_key:
        return None
    # 국내 6자리 → .KS 접미사
    av_sym = (ticker + '.KSC') if _is_krx_ticker(ticker) else ticker
    try:
        res = http_req.get(
            'https://www.alphavantage.co/query',
            params={'function': 'GLOBAL_QUOTE', 'symbol': av_sym, 'apikey': api_key},
            timeout=10
        )
        if not res.ok:
            print(f'[price] alphavantage HTTP {res.status_code} for {av_sym}', file=sys.stderr)
            return None
        data = res.json().get('Global Quote', {})
        price_str = data.get('05. price', '')
        if price_str:
            return float(price_str)
        print(f'[price] alphavantage no price for {av_sym}: {res.text[:200]}', file=sys.stderr)
    except Exception as e:
        print(f'[price] alphavantage error {ticker}: {e}', file=sys.stderr)
    return None


def _fetch_stooq_price(ticker: str) -> float | None:
    """Stooq CSV 피드로 주식 현재가(종가) 조회"""
    ticker = (ticker or '').strip()
    if not ticker:
        return None
    if not _is_krx_ticker(ticker):
        ticker = ticker.upper()
    try:
        stooq_sym = (ticker + '.kr') if _is_krx_ticker(ticker) else ticker
        res = http_req.get(
            f'https://stooq.com/q/l/?s={stooq_sym}&f=sd2t2ohlcv&h&e=csv',
            timeout=8,
            headers={'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'}
        )
        print(f'[price] stooq {stooq_sym}: status={res.status_code} body={res.text[:100]}', file=sys.stderr)
        if not res.ok:
            return None
        lines = [line.strip() for line in res.text.strip().split('\n') if line.strip()]
        if len(lines) < 2:
            return None
        headers = [h.strip().lower() for h in lines[0].split(',')]
        parts = [p.strip() for p in lines[1].split(',')]
        if 'close' in headers:
            idx = headers.index('close')
            if idx < len(parts):
                close = parts[idx]
                if close and close not in ('N/D', '0', ''):
                    return float(close)
    except Exception as e:
        print(f'[price] stooq error {ticker}: {e}', file=sys.stderr)
    return None


def _fetch_yf_direct_price(ticker: str) -> float | None:
    """Yahoo Finance v8 API 직접 호출 (regularMarketPrice 또는 Close 종가 반영)"""
    ticker = (ticker or '').strip()
    if not ticker:
        return None
    if not _is_krx_ticker(ticker):
        ticker = ticker.upper()
    time.sleep(0.5)  # rate limit 방지
    try:
        yf_sym = (ticker + '.KS') if _is_krx_ticker(ticker) else ticker
        res = http_req.get(
            f'https://query2.finance.yahoo.com/v8/finance/chart/{yf_sym}',
            params={'interval': '1d', 'range': '5d'},
            headers={'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36', 'Accept': 'application/json'},
            timeout=8
        )
        print(f'[price] yf_direct {yf_sym}: status={res.status_code} body={res.text[:100]}', file=sys.stderr)
        if not res.ok:
            return None
        result = res.json().get('chart', {}).get('result', [])
        if result:
            market_price = result[0].get('meta', {}).get('regularMarketPrice')
            if market_price is not None:
                return float(market_price)
            closes = result[0].get('indicators', {}).get('quote', [{}])[0].get('close', [])
            closes = [c for c in closes if c is not None and not (isinstance(c, float) and math.isnan(c))]
            if closes:
                return float(closes[-1])
    except Exception as e:
        print(f'[price] yf_direct error {ticker}: {e}', file=sys.stderr)
    return None


def _fetch_stock_price(ticker: str) -> float | None:
    """pykrx (국내) / yfinance (해외) 우선 시도 후 외부 HTTP API 순차 시도"""
    ticker = (ticker or '').strip()
    if not ticker:
        return None
    if not _is_krx_ticker(ticker):
        ticker = ticker.upper()
    if _is_krx_ticker(ticker):
        # 1순위: pykrx
        if HAS_PYKRX:
            try:
                import datetime
                today_str = datetime.date.today().strftime("%Y%m%d")
                prev_str = (datetime.date.today() - datetime.timedelta(days=7)).strftime("%Y%m%d")
                df = krx_stock.get_market_ohlcv_by_date(prev_str, today_str, ticker)
                if not df.empty and '종가' in df.columns:
                    return float(df['종가'].iloc[-1])
            except Exception as e:
                print(f'[price] pykrx error {ticker}: {e}', file=sys.stderr)

        # 2순위: 네이버 모바일 주식 API (HTML 스크래핑보다 안정적)
        try:
            res = http_req.get(
                f"https://m.stock.naver.com/api/stock/{ticker}/basic",
                headers={'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36', 'Referer': 'https://m.stock.naver.com/'},
                timeout=5
            )
            if res.ok:
                data = res.json()
                price = data.get('closePrice') or data.get('currentPrice')
                if price:
                    return float(str(price).replace(',', ''))
        except Exception as e:
            print(f'[price] naver mobile API error {ticker}: {e}', file=sys.stderr)

        # 3순위: 네이버 금융 시세 API
        try:
            res = http_req.get(
                f"https://polling.finance.naver.com/api/realtime/domestic/stock/{ticker}",
                headers={'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36', 'Referer': 'https://finance.naver.com/'},
                timeout=5
            )
            if res.ok:
                data = res.json()
                price = (data.get('datas') or [{}])[0].get('closePrice') or \
                        (data.get('datas') or [{}])[0].get('currentPrice')
                if price:
                    return float(str(price).replace(',', ''))
        except Exception as e:
            print(f'[price] naver polling API error {ticker}: {e}', file=sys.stderr)

        # 4순위: Yahoo Finance KS/KQ 심볼
        try:
            for suffix in ['.KS', '.KQ']:
                sym = ticker + suffix
                res = http_req.get(
                    f'https://query2.finance.yahoo.com/v8/finance/chart/{sym}',
                    params={'interval': '1d', 'range': '5d'},
                    headers={'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'},
                    timeout=5
                )
                if res.ok:
                    result = res.json().get('chart', {}).get('result', [])
                    if result:
                        closes = result[0].get('indicators', {}).get('quote', [{}])[0].get('close', [])
                        closes = [c for c in closes if c is not None]
                        if closes:
                            return float(closes[-1])
        except Exception as e:
            print(f'[price] yahoo KS/KQ error {ticker}: {e}', file=sys.stderr)

    else:
        # 해외 주식/ETF는 대문자로 변환하여 조회 시도
        ticker = ticker.upper()
        price = _fetch_yf_direct_price(ticker)
        if price is not None and not (isinstance(price, float) and math.isnan(price)):
            return price
        if HAS_YFINANCE:
            try:
                t = yf.Ticker(ticker)
                hist = t.history(period="5d")
                if not hist.empty and 'Close' in hist.columns:
                    val = float(hist['Close'].iloc[-1])
                    if not math.isnan(val):
                        return val
            except Exception as e:
                print(f'[price] yfinance error {ticker}: {e}', file=sys.stderr)

    if price is None or (isinstance(price, float) and math.isnan(price)):
        price = _fetch_stooq_price(ticker)
    if price is None or (isinstance(price, float) and math.isnan(price)):
        price = _fetch_alphavantage_price(ticker)
    if price is not None and isinstance(price, float) and (math.isnan(price) or math.isinf(price)):
        return None
    return price



def _fetch_crypto_prices(symbols: list[str]) -> dict[str, float]:
    """업비트(Upbit) API 우선 조회 후, 실패 시 CoinGecko로 KRW 현재가 조회. {symbol_upper: price}"""
    if not symbols:
        return {}
    result = {}

    # 1. 업비트 API 시도 (예: BTC -> KRW-BTC)
    for sym in symbols:
        sym_upper = sym.upper()
        try:
            market = f"KRW-{sym_upper}"
            res = http_req.get(f"https://api.upbit.com/v1/ticker?markets={market}", timeout=5, headers={'User-Agent': 'Mozilla/5.0'})
            if res.ok:
                data = res.json()
                if data and isinstance(data, list) and len(data) > 0:
                    trade_price = data[0].get('trade_price')
                    if trade_price:
                        result[sym_upper] = float(trade_price)
                        continue
        except Exception as e:
            print(f'[price] upbit error {sym}: {e}', file=sys.stderr)

    # 2. 업비트에서 조회 실패한 심볼에 대해 CoinGecko 시도
    remaining = [sym for sym in symbols if sym.upper() not in result]
    for sym in remaining:
        sym_upper = sym.upper()
        try:
            # /search 로 심볼 → coin id 획득 (전체 목록 다운로드 대신)
            search_res = http_req.get(
                'https://api.coingecko.com/api/v3/search',
                params={'query': sym},
                timeout=8
            )
            if not search_res.ok:
                continue
            coins = search_res.json().get('coins', [])
            coin_id = None
            for coin in coins:
                if coin.get('symbol', '').upper() == sym_upper:
                    coin_id = coin['id']
                    break
            if not coin_id:
                continue

            price_res = http_req.get(
                'https://api.coingecko.com/api/v3/simple/price',
                params={'ids': coin_id, 'vs_currencies': 'krw'},
                timeout=8
            )
            if not price_res.ok:
                continue
            p = price_res.json().get(coin_id, {}).get('krw')
            if p:
                result[sym_upper] = float(p)
        except Exception:
            continue
    return result


import threading

_price_update_status = {"is_updating": False, "last_result": None}
_price_update_lock = threading.Lock()

def _bg_price_update():
    global _price_update_status
    with _price_update_lock:
        _price_update_status["is_updating"] = True
    
    try:
        results = _run_price_update_logic()
        with _price_update_lock:
            _price_update_status["last_result"] = results
            # 가격 업데이트가 완료되었으므로 캐시를 무효화합니다.
            _clear_summary_cache()
    except Exception as e:
        import traceback
        with _price_update_lock:
            _price_update_status["last_result"] = {
                "stocks": [], "etf": [], "crypto": [], "split_plans": [],
                "errors": [f"백그라운드 스레드 예외: {str(e)}", traceback.format_exc()]
            }
    finally:
        with _price_update_lock:
            _price_update_status["is_updating"] = False

def _run_price_update_logic():
    import traceback
    db = get_db()
    results = {'stocks': [], 'etf': [], 'crypto': [], 'errors': []}

    try:
        # ── 주식 ──
        cur = db.cursor()
        cur.execute("SELECT id, name, ticker, COALESCE(ath,0) as ath FROM stocks WHERE ticker IS NOT NULL AND ticker != ''")
        stock_rows = cur.fetchall()
        cur.close()

        for row in stock_rows:
            sid, name, ticker, existing_ath = row['id'], row['name'], row['ticker'], row['ath']
            try:
                price = _fetch_stock_price(ticker)
            except Exception as e:
                price = None
                results['errors'].append(f"주식 [{name}({ticker})]: {e}")
            if price:
                new_ath = max(existing_ath or 0, price)
                if not existing_ath:
                    fetched = _fetch_all_time_high(ticker)
                    if fetched:
                        new_ath = fetched
                cur = db.cursor()
                cur.execute("UPDATE stocks SET current_price=%s, ath=%s WHERE id=%s", (price, new_ath, sid))
                cur.close()
                results['stocks'].append({'id': sid, 'name': name, 'ticker': ticker, 'price': price, 'ok': True})
            else:
                if not any(f"주식 [{name}({ticker})]" in e for e in results['errors']):
                    results['errors'].append(f"주식 [{name}({ticker})]: 가격 조회 실패")
                results['stocks'].append({'id': sid, 'name': name, 'ticker': ticker, 'price': None, 'ok': False})

        # ── ETF ──
        cur = db.cursor()
        cur.execute("SELECT id, name, ticker, COALESCE(ath,0) as ath FROM etf WHERE ticker IS NOT NULL AND ticker != ''")
        etf_rows = cur.fetchall()
        cur.close()

        for row in etf_rows:
            eid, name, ticker, existing_ath = row['id'], row['name'], row['ticker'], row['ath']
            try:
                price = _fetch_stock_price(ticker)
            except Exception as e:
                price = None
                results['errors'].append(f"ETF [{name}({ticker})]: {e}")
            if price:
                new_ath = max(existing_ath or 0, price)
                if not existing_ath:
                    fetched = _fetch_all_time_high(ticker)
                    if fetched:
                        new_ath = fetched
                cur = db.cursor()
                cur.execute("UPDATE etf SET current_price=%s, ath=%s WHERE id=%s", (price, new_ath, eid))
                cur.close()
                results['etf'].append({'id': eid, 'name': name, 'ticker': ticker, 'price': price, 'ok': True})
            else:
                if not any(f"ETF [{name}({ticker})]" in e for e in results['errors']):
                    results['errors'].append(f"ETF [{name}({ticker})]: 가격 조회 실패")
                results['etf'].append({'id': eid, 'name': name, 'ticker': ticker, 'price': None, 'ok': False})

        # ── 코인 ──
        cur = db.cursor()
        cur.execute("SELECT id, name, symbol FROM crypto WHERE symbol IS NOT NULL AND symbol != ''")
        crypto_rows = cur.fetchall()
        cur.close()

        if crypto_rows:
            symbols = [row['symbol'] for row in crypto_rows]
            try:
                cg_prices = _fetch_crypto_prices(symbols)
            except Exception as e:
                cg_prices = {}
                results['errors'].append(f"코인 가격 조회 오류: {e}")

            for row in crypto_rows:
                cid, name, symbol = row['id'], row['name'], row['symbol']
                price = cg_prices.get(symbol.upper())
                if price:
                    cur = db.cursor()
                    cur.execute("UPDATE crypto SET current_price = %s WHERE id = %s", (price, cid))
                    cur.close()
                    results['crypto'].append({'id': cid, 'name': name, 'symbol': symbol, 'price': price, 'ok': True})
                else:
                    results['errors'].append(f"코인 [{name}({symbol})]: 가격 조회 실패")
                    results['crypto'].append({'id': cid, 'name': name, 'symbol': symbol, 'price': None, 'ok': False})

        # ── 분할매수 계획 종목 ──
        cur = db.cursor()
        cur.execute("SELECT id, name, ticker FROM split_buy_plans WHERE ticker IS NOT NULL AND ticker != ''")
        plan_rows = cur.fetchall()
        cur.close()

        results['split_plans'] = []
        for row in plan_rows:
            pid, name, ticker = row['id'], row['name'], row['ticker']
            try:
                price = _fetch_stock_price(ticker)
            except Exception as e:
                price = None
                results['errors'].append(f"분할매수 계획 [{name}({ticker})]: {e}")
            if price:
                cur = db.cursor()
                cur.execute("UPDATE split_buy_plans SET current_price = %s WHERE id = %s", (price, pid))
                cur.close()
                results['split_plans'].append({'id': pid, 'name': name, 'ticker': ticker, 'price': price, 'ok': True})
            else:
                if not any(f"분할매수 계획 [{name}({ticker})]" in e for e in results['errors']):
                    results['errors'].append(f"분할매수 계획 [{name}({ticker})]: 가격 조회 실패")
                results['split_plans'].append({'id': pid, 'name': name, 'ticker': ticker, 'price': None, 'ok': False})

        db.commit()
    except Exception as e:
        results['errors'].append(f"서버 오류: {traceback.format_exc()}")
    finally:
        db.close()

    return results

@app.route('/api/price-update', methods=['POST'])
def api_price_update():
    """등록된 모든 종목(주식/ETF/코인)의 현재가를 외부 API로 백그라운드 스레드에서 조회 후 DB 업데이트"""
    global _price_update_status
    
    with _price_update_lock:
        if _price_update_status["is_updating"]:
            return jsonify({"ok": True, "status": "already_updating"})
            
    t = threading.Thread(target=_bg_price_update)
    t.start()
    return jsonify({"ok": True, "status": "started"})

@app.route('/api/price-update/status', methods=['GET'])
def api_price_update_status():
    """백그라운드 현재가 업데이트의 진행 상태 조회"""
    global _price_update_status
    with _price_update_lock:
        return jsonify(_price_update_status)


@app.route('/api/price-test')
def api_price_test():
    """주식 가격 소스별 진단 (예: /api/price-test?ticker=005930)"""
    ticker = request.args.get('ticker', '005930')
    av_key = os.environ.get('ALPHAVANTAGE_API_KEY', '')
    result = {'ticker': ticker, 'av_key_set': bool(av_key), 'sources': {}}

    # 1. Alpha Vantage
    try:
        av_sym = (ticker + '.KSC') if _is_krx_ticker(ticker) else ticker
        r = http_req.get('https://www.alphavantage.co/query',
            params={'function': 'GLOBAL_QUOTE', 'symbol': av_sym, 'apikey': av_key or 'demo'},
            timeout=10)
        result['sources']['alphavantage'] = {'status': r.status_code, 'body': r.text[:300]}
    except Exception as e:
        result['sources']['alphavantage'] = {'error': str(e)}

    # 2. Stooq
    try:
        stooq_sym = (ticker + '.kr') if _is_krx_ticker(ticker) else ticker
        r = http_req.get(f'https://stooq.com/q/l/?s={stooq_sym}&f=sd2t2ohlcv&h&e=csv',
            timeout=8, headers={'User-Agent': 'Mozilla/5.0'})
        result['sources']['stooq'] = {'status': r.status_code, 'body': r.text[:300]}
    except Exception as e:
        result['sources']['stooq'] = {'error': str(e)}

    # 3. Yahoo Finance 직접
    try:
        yf_sym = (ticker + '.KS') if _is_krx_ticker(ticker) else ticker
        r = http_req.get(f'https://query2.finance.yahoo.com/v8/finance/chart/{yf_sym}',
            params={'interval': '1d', 'range': '5d'},
            headers={'User-Agent': 'Mozilla/5.0', 'Accept': 'application/json'},
            timeout=8)
        result['sources']['yahoo'] = {'status': r.status_code, 'body': r.text[:300]}
    except Exception as e:
        result['sources']['yahoo'] = {'error': str(e)}

    return jsonify(result)


# ── API: USD/KRW 환율 ────────────────────────────────────────
_exchange_rate_cache = {'rate': 1380.0, 'last_updated': 0}

def get_current_exchange_rate():
    import time
    now = time.time()
    # 1시간 동안 캐시 유지
    if now - _exchange_rate_cache['last_updated'] < 3600:
        return _exchange_rate_cache['rate']

    try:
        r = http_req.get(
            'https://query2.finance.yahoo.com/v8/finance/chart/USDKRW=X',
            params={'interval': '1d', 'range': '5d'},
            headers={'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36', 'Accept': 'application/json'},
            timeout=8
        )
        if r.ok:
            result = r.json().get('chart', {}).get('result', [])
            if result:
                closes = result[0].get('indicators', {}).get('quote', [{}])[0].get('close', [])
                closes = [c for c in closes if c is not None]
                rate = None
                if closes:
                    rate = round(closes[-1], 2)
                else:
                    meta_price = result[0].get('meta', {}).get('regularMarketPrice')
                    if meta_price:
                        rate = round(meta_price, 2)
                if rate:
                    _exchange_rate_cache['rate'] = rate
                    _exchange_rate_cache['last_updated'] = now
                    return rate
    except Exception:
        pass

    # 보조 API (exchangerate.host) — Yahoo 실패 시 시도
    try:
        r2 = http_req.get('https://open.er-api.com/v6/latest/USD', timeout=8)
        if r2.ok:
            krw = r2.json().get('rates', {}).get('KRW')
            if krw:
                rate = round(krw, 2)
                _exchange_rate_cache['rate'] = rate
                _exchange_rate_cache['last_updated'] = now
                return rate
    except Exception:
        pass

    # 모두 실패 시 캐시 수명을 살짝 늘려(5분) 잦은 재시도 방지
    _exchange_rate_cache['last_updated'] = now - 3300
    return _exchange_rate_cache['rate']

def is_foreign_ticker(ticker):
    return bool(ticker) and not bool(re.match(r'^\d{6}$', str(ticker)))

def _sf(v):
    """NaN/Inf/None 을 0.0으로 안전 변환."""
    try:
        f = float(v) if v is not None else 0.0
        return 0.0 if (math.isnan(f) or math.isinf(f)) else f
    except (TypeError, ValueError):
        return 0.0

def _calc_asset_totals(rows, ex_rate):
    """(val, cost) 계산 공통 로직 — qty 음수 방지(GREATEST 0) 포함."""
    val = 0.0; cost = 0.0
    for r in rows:
        qty = max(0.0, _sf(r['buy_qty']) - _sf(r['sell_qty']))
        buy_qty_f = _sf(r['buy_qty'])
        avg = _sf(r['total_buy_amt']) / buy_qty_f if buy_qty_f > 0 else 0.0
        eval_amt = round(qty * _sf(r['current_price']))
        cost_amt = round(qty * avg)
        mul = ex_rate if is_foreign_ticker(r['ticker']) else 1.0
        val  += eval_amt * mul
        cost += cost_amt * mul
    return float(val), float(cost)

def _fetch_stock_rows(db):
    cur = db.cursor()
    cur.execute("""
        SELECT s.ticker, s.current_price,
            COALESCE(SUM(CASE WHEN t.tx_type IN ('buy','매수') THEN t.quantity ELSE 0 END), 0) AS buy_qty,
            COALESCE(SUM(CASE WHEN t.tx_type IN ('sell','매도') THEN t.quantity ELSE 0 END), 0) AS sell_qty,
            COALESCE(SUM(CASE WHEN t.tx_type IN ('buy','매수') THEN t.price * t.quantity ELSE 0 END), 0) AS total_buy_amt
        FROM stocks s LEFT JOIN stock_tx t ON t.stock_id = s.id
        GROUP BY s.id
    """)
    rows = cur.fetchall(); cur.close()
    return rows

def _fetch_etf_rows(db):
    cur = db.cursor()
    cur.execute("""
        SELECT e.ticker, e.current_price,
            COALESCE(SUM(CASE WHEN t.tx_type IN ('buy','매수') THEN t.quantity ELSE 0 END), 0) AS buy_qty,
            COALESCE(SUM(CASE WHEN t.tx_type IN ('sell','매도') THEN t.quantity ELSE 0 END), 0) AS sell_qty,
            COALESCE(SUM(CASE WHEN t.tx_type IN ('buy','매수') THEN t.price * t.quantity ELSE 0 END), 0) AS total_buy_amt
        FROM etf e LEFT JOIN etf_tx t ON t.etf_id = e.id
        GROUP BY e.id
    """)
    rows = cur.fetchall(); cur.close()
    return rows

def get_stocks_etf_totals(db, ex_rate):
    """(stocks_val, stocks_cost, etf_val, etf_cost) — 세 페이지 공통 계산 함수."""
    s_val, s_cost = _calc_asset_totals(_fetch_stock_rows(db), ex_rate)
    e_val, e_cost = _calc_asset_totals(_fetch_etf_rows(db), ex_rate)
    return s_val, s_cost, e_val, e_cost

def get_real_estate_value(db):
    """
    부동산 평가가치 합계.
    잔금(최종 납입)이 실제 완료된 매물만 current_price 전액을 반영하고,
    아직 잔금을 치르지 않은 매물은 지금까지 실제 지급한 금액(계약금/중도금 등, actual_date 도달분)만 반영한다.
    결제 일정이 전혀 등록되지 않은 매물(레거시 데이터)은 기존처럼 current_price 전액을 반영한다.
    """
    cur = db.cursor()
    cur.execute("SELECT id, current_price FROM real_estate")
    properties = cur.fetchall()
    total = 0.0
    for prop in properties:
        rid = prop['id']
        price = float(prop['current_price'] or 0)
        cur.execute(
            "SELECT payment_type, amount FROM real_estate_payments "
            "WHERE real_estate_id=%s AND direction='buy' AND actual_date IS NOT NULL AND actual_date::date <= CURRENT_DATE",
            (rid,)
        )
        payments = cur.fetchall()
        if not payments:
            total += price
            continue
        final_paid = any(p['payment_type'] == '잔금' for p in payments)
        if final_paid:
            total += price
        else:
            total += sum(float(p['amount'] or 0) for p in payments)
    cur.close()
    return total

# 하위 호환 래퍼 (기존 호출부 유지)
def get_stocks_total_value(db, ex_rate):
    return get_stocks_etf_totals(db, ex_rate)[0]

def get_etf_total_value(db, ex_rate):
    return get_stocks_etf_totals(db, ex_rate)[2]

def get_stocks_total_and_cost(db, ex_rate):
    s_val, s_cost, _, _ = get_stocks_etf_totals(db, ex_rate)
    return s_val, s_cost

def get_etf_total_and_cost(db, ex_rate):
    _, _, e_val, e_cost = get_stocks_etf_totals(db, ex_rate)
    return e_val, e_cost

@app.route('/api/exchange-rate')
def api_exchange_rate():
    """Yahoo Finance에서 USD/KRW 환율 조회. 실패 시 1380 반환"""
    return jsonify({'rate': get_current_exchange_rate()})


@app.route('/api/stocks-etf-total')
def api_stocks_etf_total():
    """주식+ETF 합산 평가액 — 대시보드·테크트리와 완전히 동일한 계산 기준."""
    db = get_db()
    ex_rate = get_current_exchange_rate()
    s_val, s_cost, e_val, e_cost = get_stocks_etf_totals(db, ex_rate)
    db.close()
    return jsonify({
        'stocks_val':  int(s_val),
        'etf_val':     int(e_val),
        'combined':    int(s_val + e_val),
        'ex_rate':     ex_rate,
    })


# ── API: 거주지 ──────────────────────────────────────────────
@app.route('/api/residence', methods=['GET', 'POST'])
def api_residence():
    db = get_db()
    if request.method == 'GET':
        cur = db.cursor()
        cur.execute("SELECT * FROM residence ORDER BY id DESC")
        rows = cur.fetchall()
        cur.close()
        db.close()
        return jsonify(rows_to_list(rows))

    data = request.json
    cur = db.cursor()
    cur.execute(
    "INSERT INTO residence (address, deposit, monthly_rent, maintenance, start_date, end_date) VALUES (%s,%s,%s,%s,%s,%s)",
    (data.get('address'), data.get('deposit', 0), data.get('monthly_rent', 0),
    data.get('maintenance', 0), data.get('start_date'), data.get('end_date'))
    )
    cur.close()
    db.commit()
    db.close()
    return jsonify({'ok': True}), 201


@app.route('/api/residence/<int:rid>', methods=['PUT', 'DELETE'])
def api_residence_detail(rid):
    db = get_db()
    if request.method == 'PUT':
        data = request.json
        cur = db.cursor()
        cur.execute(
        "UPDATE residence SET address=%s, deposit=%s, monthly_rent=%s, maintenance=%s, start_date=%s, end_date=%s WHERE id=%s",
        (data.get('address'), data.get('deposit', 0), data.get('monthly_rent', 0),
        data.get('maintenance', 0), data.get('start_date'), data.get('end_date'), rid)
        )
        cur.close()
        db.commit()
        db.close()
        return jsonify({'ok': True})
    cur = db.cursor()
    cur.execute("DELETE FROM residence WHERE id = %s", (rid,))
    cur.close()
    db.commit()
    db.close()
    return jsonify({'ok': True})


# ── API: 부동산 ──────────────────────────────────────────────
def _re_enrich(db, rows):
    """real_estate 목록에 현재 계약 정보·비용 집계·실수익률 추가"""
    if not rows:
        return []
        
    re_ids = [r['id'] for r in rows]
    
    cur = db.cursor()
    # 부동산별 가장 최근의 계약 가져오기 (PostgreSQL의 DISTINCT ON 사용)
    cur.execute("""
        SELECT DISTINCT ON (real_estate_id)
            real_estate_id, contract_type, deposit, monthly_rent, end_date
        FROM tenant_contracts
        WHERE real_estate_id IN %s
        ORDER BY real_estate_id, end_date DESC, id DESC
    """, (tuple(re_ids),))
    contracts = {r['real_estate_id']: r for r in cur.fetchall()}
    cur.close()

    cur = db.cursor()
    # 부동산별 취득비용 및 순손익(임대수익 - 기타비용) 합계를 한번에 가져오기
    cur.execute("""
        SELECT real_estate_id,
            COALESCE(SUM(CASE WHEN cost_type='취득비용' THEN amount ELSE 0 END), 0) AS acq_cost,
            COALESCE(SUM(CASE WHEN cost_type='임대수익' THEN amount ELSE -amount END), 0) AS net_extra
        FROM property_costs
        WHERE real_estate_id IN %s
        GROUP BY real_estate_id
    """, (tuple(re_ids),))
    costs = {r['real_estate_id']: r for r in cur.fetchall()}
    cur.close()

    result = []
    for r in rows:
        rid = r['id']
        contract = contracts.get(rid)
        cost = costs.get(rid, {'acq_cost': 0, 'net_extra': 0})

        acq_cost = cost['acq_cost']
        net_extra = cost['net_extra']

        deposit = contract['deposit'] if contract else 0
        purchase = r['purchase_price']
        current  = r['current_price']
        real_inv = int(purchase - deposit + acq_cost)   # 실투자금
        net_gain = int((current - purchase) + net_extra)  # 순손익
        real_roi = round(float(net_gain) / real_inv * 100, 1) if real_inv > 0 else None

        row = dict(r)
        row['contract_type']  = contract['contract_type'] if contract else None
        row['deposit']        = deposit
        row['monthly_rent']   = contract['monthly_rent'] if contract else 0
        row['contract_end']   = contract['end_date'] if contract else None
        row['real_inv']       = real_inv
        row['net_gain']       = net_gain
        row['real_roi']       = real_roi
        result.append(row)
    return result


@app.route('/api/real-estate', methods=['GET', 'POST'])
def api_real_estate():
    db = get_db()
    if request.method == 'GET':
        cur = db.cursor()
        cur.execute("SELECT * FROM real_estate ORDER BY name")
        rows = cur.fetchall()
        cur.close()
        result = _re_enrich(db, rows)
        db.close()
        return jsonify(result)

    data = request.json
    cur = db.cursor()
    cur.execute(
    "INSERT INTO real_estate (name, re_type, purchase_date, purchase_price, current_price, memo) VALUES (%s,%s,%s,%s,%s,%s) RETURNING id",
    (data.get('name'), data.get('re_type'), data.get('purchase_date'),
    data.get('purchase_price', 0), data.get('current_price', 0), data.get('memo'))
    )
    new_id = cur.fetchone()[0]
    cur.close()
    db.commit()
    db.close()
    return jsonify({'ok': True, 'id': new_id}), 201


@app.route('/api/real-estate/<int:rid>', methods=['PUT', 'DELETE'])
def api_real_estate_detail(rid):
    db = get_db()
    if request.method == 'PUT':
        data = request.json
        cur = db.cursor()
        cur.execute(
        "UPDATE real_estate SET name=%s, re_type=%s, purchase_date=%s, purchase_price=%s, current_price=%s, memo=%s WHERE id=%s",
        (data.get('name'), data.get('re_type'), data.get('purchase_date'),
        data.get('purchase_price', 0), data.get('current_price', 0), data.get('memo'), rid)
        )
        cur.close()
        db.commit()
        db.close()
        return jsonify({'ok': True})
    cur = db.cursor()
    cur.execute("DELETE FROM real_estate WHERE id = %s", (rid,))
    cur.close()
    db.commit()
    db.close()
    return jsonify({'ok': True})


@app.route('/api/re-summary')
def api_re_summary():
    db = get_db()
    cur = db.cursor()
    cur.execute("SELECT * FROM real_estate")
    rows = cur.fetchall()
    cur.close()
    enriched = _re_enrich(db, rows)
    db.close()
    total_purchase = int(sum(r['purchase_price'] for r in enriched))
    total_deposit  = int(sum(r['deposit'] for r in enriched))
    total_real_inv = int(sum(r['real_inv'] for r in enriched))
    total_net_gain = int(sum(r['net_gain'] for r in enriched))
    avg_roi = round(float(total_net_gain) / total_real_inv * 100, 1) if total_real_inv > 0 else None
    return jsonify({
        'count': len(enriched),
        'total_purchase': total_purchase,
        'total_deposit': total_deposit,
        'total_real_inv': total_real_inv,
        'total_net_gain': total_net_gain,
        'avg_roi': avg_roi,
    })


@app.route('/api/re-summary-sold')
def api_re_summary_sold():
    db = get_db()
    cur = db.cursor()
    cur.execute("SELECT real_inv, profit FROM sold_real_estate")
    rows = cur.fetchall()
    cur.execute("SELECT COUNT(*) AS cnt FROM sold_real_estate")
    cnt = cur.fetchone()['cnt']
    cur.close(); db.close()
    data = rows_to_list(rows)
    total_real_inv = int(sum(r.get('real_inv') or 0 for r in data))
    total_profit   = int(sum(r.get('profit')   or 0 for r in data))
    avg_roi = round(float(total_profit) / total_real_inv * 100, 1) if total_real_inv > 0 else None
    return jsonify({
        'count':         cnt,
        'total_real_inv': total_real_inv,
        'total_profit':   total_profit,
        'avg_roi':        avg_roi,
    })


@app.route('/api/re-expiring')
def api_re_expiring():
    """3개월 이내 만료 계약 목록"""
    db = get_db()
    cur = db.cursor()
    cur.execute("""
    SELECT c.*, r.name as re_name
    FROM tenant_contracts c
    JOIN real_estate r ON c.real_estate_id = r.id
    WHERE c.end_date IS NOT NULL
    AND c.end_date >= CURRENT_DATE
    AND c.end_date <= CURRENT_DATE + INTERVAL '3 months'
    ORDER BY c.end_date
    """)
    rows = cur.fetchall()
    cur.close()
    db.close()
    return jsonify(rows_to_list(rows))


@app.route('/api/tenant-contracts', methods=['GET', 'POST'])
def api_tenant_contracts():
    db = get_db()
    if request.method == 'GET':
        rid = request.args.get('real_estate_id')
        cur = db.cursor()
        cur.execute(
        "SELECT * FROM tenant_contracts WHERE real_estate_id=%s ORDER BY end_date DESC", (rid,)
        )
        rows = cur.fetchall()
        cur.close()
        db.close()
        return jsonify(rows_to_list(rows))
    d = request.json or {}
    cur = db.cursor()
    cur.execute(
    "INSERT INTO tenant_contracts (real_estate_id, contract_type, deposit, monthly_rent, start_date, end_date, memo)"
    " VALUES (%s,%s,%s,%s,%s,%s,%s)",
    (d.get('real_estate_id'), d.get('contract_type'), d.get('deposit', 0),
    d.get('monthly_rent', 0), d.get('start_date'), d.get('end_date'), d.get('memo'))
    )
    cur.close()
    db.commit(); db.close()
    return jsonify({'ok': True}), 201


@app.route('/api/tenant-contracts/<int:rid>', methods=['PUT', 'DELETE'])
def api_tenant_contracts_detail(rid):
    db = get_db()
    if request.method == 'DELETE':
        cur = db.cursor()
        cur.execute("DELETE FROM tenant_contracts WHERE id=%s", (rid,))
        cur.close()
        db.commit(); db.close()
        return jsonify({'ok': True})
    d = request.json or {}
    cur = db.cursor()
    cur.execute(
    "UPDATE tenant_contracts SET contract_type=%s, deposit=%s, monthly_rent=%s, start_date=%s, end_date=%s, memo=%s WHERE id=%s",
    (d.get('contract_type'), d.get('deposit', 0), d.get('monthly_rent', 0),
    d.get('start_date'), d.get('end_date'), d.get('memo'), rid)
    )
    cur.close()
    db.commit(); db.close()
    return jsonify({'ok': True})


@app.route('/api/property-costs', methods=['GET', 'POST'])
def api_property_costs():
    db = get_db()
    if request.method == 'GET':
        rid = request.args.get('real_estate_id')
        cur = db.cursor()
        cur.execute(
        "SELECT * FROM property_costs WHERE real_estate_id=%s ORDER BY date DESC", (rid,)
        )
        rows = cur.fetchall()
        cur.close()
        db.close()
        return jsonify(rows_to_list(rows))
    d = request.json or {}
    cur = db.cursor()
    cur.execute(
    "INSERT INTO property_costs (real_estate_id, cost_type, name, amount, date, memo) VALUES (%s,%s,%s,%s,%s,%s)",
    (d.get('real_estate_id'), d.get('cost_type'), d.get('name'),
    d.get('amount', 0), d.get('date'), d.get('memo'))
    )
    cur.close()
    db.commit(); db.close()
    return jsonify({'ok': True}), 201


@app.route('/api/property-costs/<int:rid>', methods=['PUT', 'DELETE'])
def api_property_costs_detail(rid):
    db = get_db()
    if request.method == 'DELETE':
        cur = db.cursor()
        cur.execute("DELETE FROM property_costs WHERE id=%s", (rid,))
        cur.close()
        db.commit(); db.close()
        return jsonify({'ok': True})
    d = request.json or {}
    cur = db.cursor()
    cur.execute(
    "UPDATE property_costs SET cost_type=%s, name=%s, amount=%s, date=%s, memo=%s WHERE id=%s",
    (d.get('cost_type'), d.get('name'), d.get('amount', 0),
    d.get('date'), d.get('memo'), rid)
    )
    cur.close()
    db.commit(); db.close()
    return jsonify({'ok': True})


# ── API: 앱 설정 ──────────────────────────────────────────────
@app.route('/api/settings/<key>', methods=['GET', 'PUT'])
def api_settings(key):
    db = get_db()
    cur = db.cursor()
    if request.method == 'GET':
        cur.execute("SELECT value FROM app_settings WHERE key=%s", (key,))
        row = cur.fetchone()
        cur.close(); db.close()
        return jsonify({'key': key, 'value': row['value'] if row else None})
    value = (request.json or {}).get('value')
    cur.execute(
        "INSERT INTO app_settings (key, value) VALUES (%s,%s) "
        "ON CONFLICT (key) DO UPDATE SET value=EXCLUDED.value",
        (key, value)
    )
    cur.close()
    db.commit(); db.close()
    return jsonify({'ok': True})


@app.route('/api/real-estate/<int:rid>/save-sell-schedule', methods=['POST'])
def api_real_estate_save_sell_schedule(rid):
    d = request.json or {}
    db = get_db()
    cur = db.cursor()
    cur.execute(
        "UPDATE real_estate SET sell_date=%s, sell_tax=%s, sell_other_costs=%s, sell_memo=%s WHERE id=%s",
        (d.get('sell_date') or None, int(d.get('sell_tax', 0)), int(d.get('sell_other_costs', 0)), d.get('sell_memo', ''), rid)
    )
    cur.close()
    db.commit(); db.close()
    return jsonify({'ok': True})


# ── API: 매도 부동산 ──────────────────────────────────────────
@app.route('/api/real-estate/<int:rid>/sell', methods=['POST'])
def api_real_estate_sell(rid):
    d = request.json or {}
    db = get_db()
    cur = db.cursor()

    cur.execute("SELECT * FROM real_estate WHERE id=%s", (rid,))
    re = cur.fetchone()
    cur.close()
    if not re:
        db.close()
        return jsonify({'error': '부동산을 찾을 수 없습니다'}), 404

    # 취득비용 합계
    cur = db.cursor()
    cur.execute(
        "SELECT COALESCE(SUM(amount),0) as v FROM property_costs "
        "WHERE real_estate_id=%s AND cost_type='취득비용'", (rid,)
    )
    acq_cost = cur.fetchone()['v']
    cur.close()

    # 현재 보증금
    cur = db.cursor()
    cur.execute(
        "SELECT COALESCE(deposit,0) as v FROM tenant_contracts "
        "WHERE real_estate_id=%s ORDER BY end_date DESC LIMIT 1", (rid,)
    )
    row = cur.fetchone()
    cur.close()
    deposit = row['v'] if row else 0

    purchase_price = re['purchase_price']
    sell_price   = int(d.get('sell_price', 0))
    tax          = int(d.get('tax', 0))
    other_costs  = int(d.get('other_costs', 0))
    real_inv     = purchase_price - deposit + acq_cost
    profit       = sell_price - purchase_price - tax - other_costs
    roi          = round(profit / real_inv * 100, 1) if real_inv > 0 else 0.0

    from datetime import date
    cur = db.cursor()
    cur.execute(
        "INSERT INTO sold_real_estate "
        "(name, re_type, purchase_date, purchase_price, real_inv, sell_date, sell_price, tax, other_costs, profit, roi, memo, created_at) "
        "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id",
        (re['name'], re['re_type'], re['purchase_date'], purchase_price, real_inv,
         d.get('sell_date'), sell_price, tax, other_costs, profit, roi,
         d.get('memo'), str(date.today()))
    )
    sold_id = cur.fetchone()['id']
    cur.close()

    cur = db.cursor()
    # 1. 기존 매수/매도 결제 단계 일정을 새로운 sold_real_estate_id에 매핑하고 real_estate_id는 NULL로 분리
    cur.execute(
        "UPDATE real_estate_payments SET sold_real_estate_id = %s, real_estate_id = NULL WHERE real_estate_id = %s",
        (sold_id, rid)
    )

    # 2. 이번 매도 확정 폼에서 넘어온 매도 거래 단계(payments) 저장
    cur.execute("DELETE FROM real_estate_payments WHERE sold_real_estate_id = %s AND direction = 'sell'", (sold_id,))
    payments = d.get('payments', [])
    for p in payments:
        cur.execute(
            "INSERT INTO real_estate_payments (sold_real_estate_id, direction, payment_type, scheduled_date, actual_date, amount, memo) "
            "VALUES (%s, 'sell', %s, %s, %s, %s, '')",
            (sold_id, p.get('payment_type'), p.get('scheduled_date') or None, p.get('actual_date') or None, p.get('amount', 0))
        )
    cur.close()

    # 원본 데이터 삭제 (관련 레코드 먼저; real_estate_payments는 분리 완료되어 영향 없음)
    cur = db.cursor()
    cur.execute("DELETE FROM tenant_contracts WHERE real_estate_id=%s", (rid,))
    cur.execute("DELETE FROM property_costs WHERE real_estate_id=%s", (rid,))
    cur.execute("DELETE FROM real_estate WHERE id=%s", (rid,))
    cur.close()

    db.commit(); db.close()
    return jsonify({'ok': True, 'profit': profit, 'roi': roi}), 201


@app.route('/api/sold-real-estate', methods=['GET'])
def api_sold_real_estate():
    db = get_db()
    cur = db.cursor()
    cur.execute("SELECT * FROM sold_real_estate ORDER BY sell_date DESC")
    rows = cur.fetchall()
    cur.close(); db.close()
    return jsonify(rows_to_list(rows))


@app.route('/api/sold-real-estate/<int:sid>', methods=['PUT', 'DELETE'])
def api_sold_real_estate_detail(sid):
    db = get_db()
    if request.method == 'PUT':
        d = request.json or {}
        purchase_price = int(d.get('purchase_price', 0))
        real_inv = int(d.get('real_inv', 0))
        sell_price = int(d.get('sell_price', 0))
        tax = int(d.get('tax', 0))
        other_costs = int(d.get('other_costs', 0))
        
        # 차익 및 수익률 자동 재계산 (데이터 일관성 보장)
        profit = sell_price - purchase_price - tax - other_costs
        roi = round(profit / real_inv * 100, 1) if real_inv > 0 else 0.0

        cur = db.cursor()
        cur.execute(
            "UPDATE sold_real_estate SET name=%s, re_type=%s, purchase_date=%s, purchase_price=%s, real_inv=%s, "
            "sell_date=%s, sell_price=%s, tax=%s, other_costs=%s, profit=%s, roi=%s, memo=%s, lease_memo=%s WHERE id=%s",
            (d.get('name'), d.get('re_type'), d.get('purchase_date'), purchase_price, real_inv,
             d.get('sell_date'), sell_price, tax, other_costs, profit, roi, d.get('memo'), d.get('lease_memo'), sid)
        )
        cur.close()
        db.commit(); db.close()
        return jsonify({'ok': True})

    cur = db.cursor()
    cur.execute("DELETE FROM sold_real_estate WHERE id=%s", (sid,))
    cur.close()
    db.commit(); db.close()
    return jsonify({'ok': True})


# ── API: 부동산 거래 단계 (계약금/중도금/잔금) ────────────────────
@app.route('/api/real-estate-payments', methods=['GET', 'POST'])
def api_re_payments():
    db = get_db()
    if request.method == 'GET':
        rid = request.args.get('real_estate_id')
        sid = request.args.get('sold_real_estate_id')
        cur = db.cursor()
        if rid:
            cur.execute(
                "SELECT * FROM real_estate_payments WHERE real_estate_id=%s ORDER BY scheduled_date, id",
                (rid,)
            )
        elif sid:
            cur.execute(
                "SELECT * FROM real_estate_payments WHERE sold_real_estate_id=%s ORDER BY scheduled_date, id",
                (sid,)
            )
        else:
            cur.execute(
                "SELECT p.*, r.name AS re_name FROM real_estate_payments p "
                "LEFT JOIN real_estate r ON p.real_estate_id=r.id "
                "ORDER BY p.scheduled_date, p.id"
            )
        rows = cur.fetchall()
        cur.close(); db.close()
        return jsonify(rows_to_list(rows))

    d = request.json
    cur = db.cursor()
    cur.execute(
        "INSERT INTO real_estate_payments (real_estate_id, direction, payment_type, scheduled_date, actual_date, amount, memo) "
        "VALUES (%s,%s,%s,%s,%s,%s,%s) RETURNING id",
        (d.get('real_estate_id'), d.get('direction'), d.get('payment_type'),
         d.get('scheduled_date') or None, d.get('actual_date') or None,
         d.get('amount', 0), d.get('memo'))
    )
    new_id = cur.fetchone()[0]

    # 매수일정 잔금일에 취득일 자동 등록
    if d.get('direction') == 'buy' and d.get('payment_type') == '잔금' and d.get('actual_date'):
        cur.execute(
            "UPDATE real_estate SET purchase_date = %s WHERE id = %s",
            (d.get('actual_date'), d.get('real_estate_id'))
        )

    # 매수 실지급 → 현금 자동 조정 생성
    if d.get('direction') == 'buy' and d.get('actual_date'):
        re_id = d.get('real_estate_id')
        if re_id:
            cur.execute("SELECT name FROM real_estate WHERE id=%s", (re_id,))
            r = cur.fetchone()
            rname = r['name'] if r else '부동산'
        else:
            rname = '부동산'
        ptype = d.get('payment_type', '')
        amt = -int(d.get('amount', 0))
        _upsert_cash_adj(cur, 're_payment', new_id, amt, f"{rname} {ptype}", d.get('actual_date'))

    cur.close()
    db.commit(); db.close()
    return jsonify({'id': new_id}), 201


@app.route('/api/real-estate-payments/<int:pid>', methods=['PUT', 'DELETE'])
def api_re_payment_detail(pid):
    db = get_db()
    if request.method == 'DELETE':
        cur = db.cursor()
        _remove_cash_adj(cur, 're_payment', pid)
        cur.execute("DELETE FROM real_estate_payments WHERE id=%s", (pid,))
        cur.close()
        db.commit(); db.close()
        return jsonify({'ok': True})

    d = request.json
    cur = db.cursor()
    cur.execute(
        "UPDATE real_estate_payments SET direction=%s, payment_type=%s, scheduled_date=%s, "
        "actual_date=%s, amount=%s, memo=%s WHERE id=%s",
        (d.get('direction'), d.get('payment_type'),
         d.get('scheduled_date') or None, d.get('actual_date') or None,
         d.get('amount', 0), d.get('memo'), pid)
    )

    # 매수일정 잔금일에 취득일 자동 등록
    if d.get('direction') == 'buy' and d.get('payment_type') == '잔금' and d.get('actual_date'):
        cur.execute("SELECT real_estate_id FROM real_estate_payments WHERE id=%s", (pid,))
        row = cur.fetchone()
        if row and row[0]:
            cur.execute(
                "UPDATE real_estate SET purchase_date = %s WHERE id = %s",
                (d.get('actual_date'), row[0])
            )

    # 매수 실지급 → 현금 자동 조정 갱신
    if d.get('direction') == 'buy' and d.get('actual_date'):
        cur.execute(
            "SELECT r.name FROM real_estate_payments p "
            "LEFT JOIN real_estate r ON r.id=p.real_estate_id WHERE p.id=%s", (pid,)
        )
        row = cur.fetchone()
        rname = row['name'] if row and row.get('name') else '부동산'
        ptype = d.get('payment_type', '')
        amt = -int(d.get('amount', 0))
        _upsert_cash_adj(cur, 're_payment', pid, amt, f"{rname} {ptype}", d.get('actual_date'))
    else:
        _remove_cash_adj(cur, 're_payment', pid)

    cur.close()
    db.commit(); db.close()
    return jsonify({'ok': True})


@app.route('/api/real-estate-payments/active')
def api_re_payments_active():
    """진행 중인 거래 단계 요약 (대시보드용)"""
    db = get_db()
    cur = db.cursor()
    # 아직 완료되지 않은 거래 단계가 있는 부동산 목록
    cur.execute("""
        SELECT p.*, r.name AS re_name, r.current_price
        FROM real_estate_payments p
        LEFT JOIN real_estate r ON p.real_estate_id = r.id
        WHERE p.real_estate_id IS NOT NULL
        ORDER BY p.real_estate_id, p.scheduled_date, p.id
    """)
    rows = cur.fetchall()
    cur.close(); db.close()
    return jsonify(rows_to_list(rows))


# ── API: 대출 ────────────────────────────────────────────────
@app.route('/api/loans', methods=['GET', 'POST'])
def api_loans():
    db = get_db()
    if request.method == 'GET':
        cur = db.cursor()
        cur.execute("SELECT * FROM loans ORDER BY name")
        rows = cur.fetchall()
        cur.close()
        db.close()
        return jsonify(rows_to_list(rows))

    data = request.json
    cur = db.cursor()
    cur.execute(
    "INSERT INTO loans (name, institution, principal, remaining, monthly_payment, interest_rate, loan_date, end_date, memo) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)",
    (data.get('name'), data.get('institution'), data.get('principal', 0),
    data.get('remaining', 0), data.get('monthly_payment', 0),
    data.get('interest_rate', 0), data.get('loan_date'), data.get('end_date'), data.get('memo'))
    )
    cur.close()
    db.commit()
    db.close()
    return jsonify({'ok': True}), 201


@app.route('/api/loans/<int:rid>', methods=['PUT', 'DELETE'])
def api_loans_detail(rid):
    db = get_db()
    if request.method == 'PUT':
        data = request.json
        cur = db.cursor()
        cur.execute(
        "UPDATE loans SET name=%s, institution=%s, principal=%s, remaining=%s, monthly_payment=%s, interest_rate=%s, loan_date=%s, end_date=%s, memo=%s WHERE id=%s",
        (data.get('name'), data.get('institution'), data.get('principal', 0),
        data.get('remaining', 0), data.get('monthly_payment', 0),
        data.get('interest_rate', 0), data.get('loan_date'), data.get('end_date'), data.get('memo'), rid)
        )
        cur.close()
        db.commit()
        db.close()
        return jsonify({'ok': True})
    cur = db.cursor()
    cur.execute("DELETE FROM loans WHERE id = %s", (rid,))
    cur.close()
    db.commit()
    db.close()
    return jsonify({'ok': True})


# ── API: 연금 ────────────────────────────────────────────────
@app.route('/api/pension', methods=['GET', 'POST'])
def api_pension():
    db = get_db()
    if request.method == 'GET':
        cur = db.cursor()
        cur.execute("SELECT * FROM pension ORDER BY pension_type, name")
        rows = cur.fetchall()
        cur.close()
        db.close()
        return jsonify(rows_to_list(rows))

    data = request.json
    cur = db.cursor()
    cur.execute(
    "INSERT INTO pension (pension_type, name, institution, monthly_payment, accumulated, return_rate, memo) VALUES (%s,%s,%s,%s,%s,%s,%s)",
    (data.get('pension_type'), data.get('name'), data.get('institution'),
    data.get('monthly_payment', 0), data.get('accumulated', 0),
    data.get('return_rate', 0), data.get('memo'))
    )
    cur.close()
    db.commit()
    db.close()
    return jsonify({'ok': True}), 201


@app.route('/api/pension/<int:rid>', methods=['PUT', 'DELETE'])
def api_pension_detail(rid):
    db = get_db()
    if request.method == 'PUT':
        data = request.json
        cur = db.cursor()
        cur.execute(
        "UPDATE pension SET pension_type=%s, name=%s, institution=%s, monthly_payment=%s, accumulated=%s, return_rate=%s, memo=%s WHERE id=%s",
        (data.get('pension_type'), data.get('name'), data.get('institution'),
        data.get('monthly_payment', 0), data.get('accumulated', 0),
        data.get('return_rate', 0), data.get('memo'), rid)
        )
        cur.close()
        db.commit()
        db.close()
        return jsonify({'ok': True})
    cur = db.cursor()
    cur.execute("DELETE FROM pension WHERE id = %s", (rid,))
    cur.close()
    db.commit()
    db.close()
    return jsonify({'ok': True})


# ── API: 목표저축 ────────────────────────────────────────────
@app.route('/api/goals', methods=['GET', 'POST'])
def api_goals():
    db = get_db()
    if request.method == 'GET':
        cur = db.cursor()
        cur.execute("SELECT * FROM goals WHERE name != '자본주의테크트리' ORDER BY target_date")
        rows = cur.fetchall()
        cur.close()
        db.close()
        return jsonify(rows_to_list(rows))

    data = request.json
    cur = db.cursor()
    cur.execute(
    "INSERT INTO goals (name, target_amount, current_amount, monthly_saving, target_date, memo) VALUES (%s,%s,%s,%s,%s,%s)",
    (data.get('name'), data.get('target_amount', 0), data.get('current_amount', 0),
    data.get('monthly_saving', 0), data.get('target_date'), data.get('memo'))
    )
    cur.close()
    db.commit()
    db.close()
    return jsonify({'ok': True}), 201


@app.route('/api/goals/<int:rid>', methods=['PUT', 'DELETE'])
def api_goals_detail(rid):
    db = get_db()
    if request.method == 'PUT':
        data = request.json
        cur = db.cursor()
        cur.execute(
        "UPDATE goals SET name=%s, target_amount=%s, current_amount=%s, monthly_saving=%s, target_date=%s, memo=%s WHERE id=%s",
        (data.get('name'), data.get('target_amount', 0), data.get('current_amount', 0),
        data.get('monthly_saving', 0), data.get('target_date'), data.get('memo'), rid)
        )
        cur.close()
        db.commit()
        db.close()
        return jsonify({'ok': True})
    cur = db.cursor()
    cur.execute("DELETE FROM goals WHERE id = %s", (rid,))
    cur.close()
    db.commit()
    db.close()
    return jsonify({'ok': True})


# ── 현금 자동 조정 헬퍼 ─────────────────────────────────────
def _upsert_cash_adj(cur, source_type, source_id, amount, description, adj_date=None):
    """자동조정 기록 후 즉시 현금 잔액에 반영 (잔액 최대 계좌 차감)"""
    if adj_date is None:
        adj_date = date.today().isoformat()

    # 이전 동일 소스의 기존 조정액 조회 (역적용 위해)
    cur.execute(
        "SELECT amount FROM cash_auto_adjustments WHERE source_type=%s AND source_id=%s",
        (source_type, source_id)
    )
    prev = cur.fetchone()
    prev_amount = int(prev['amount']) if prev else 0

    # 자동조정 테이블 upsert
    cur.execute("""
        INSERT INTO cash_auto_adjustments (adj_date, source_type, source_id, amount, description, applied)
        VALUES (%s, %s, %s, %s, %s, TRUE)
        ON CONFLICT (source_type, source_id) DO UPDATE
            SET adj_date=EXCLUDED.adj_date, amount=EXCLUDED.amount,
                description=EXCLUDED.description, applied=TRUE
    """, (adj_date, source_type, source_id, amount, description))

    # 현금 최대 잔액 계좌에 차액 즉시 반영
    delta = amount - prev_amount
    if delta != 0:
        cur.execute("SELECT id, amount FROM cash_deposits ORDER BY amount DESC LIMIT 1")
        acct = cur.fetchone()
        if acct:
            cur.execute(
                "UPDATE cash_deposits SET amount=%s, updated_date=%s WHERE id=%s",
                (int(acct['amount']) + delta, adj_date, acct['id'])
            )

def _remove_cash_adj(cur, source_type, source_id):
    """자동조정 삭제 시 현금 잔액도 역적용"""
    cur.execute(
        "SELECT amount FROM cash_auto_adjustments WHERE source_type=%s AND source_id=%s",
        (source_type, source_id)
    )
    prev = cur.fetchone()
    if prev:
        prev_amount = int(prev['amount'])
        cur.execute("SELECT id, amount FROM cash_deposits ORDER BY amount DESC LIMIT 1")
        acct = cur.fetchone()
        if acct:
            cur.execute(
                "UPDATE cash_deposits SET amount=%s, updated_date=%s WHERE id=%s",
                (int(acct['amount']) - prev_amount, date.today().isoformat(), acct['id'])
            )
    cur.execute(
        "DELETE FROM cash_auto_adjustments WHERE source_type=%s AND source_id=%s",
        (source_type, source_id)
    )


# ── API: 현금/예금 ───────────────────────────────────────────
@app.route('/api/cash-deposits', methods=['GET', 'POST'])
def api_cash_deposits():
    db = get_db()
    if request.method == 'GET':
        cur = db.cursor()
        cur.execute("SELECT * FROM cash_deposits ORDER BY name")
        rows = cur.fetchall()
        cur.close()
        db.close()
        return jsonify(rows_to_list(rows))

    data = request.json
    today = date.today().isoformat()
    cur = db.cursor()
    cur.execute(
    "INSERT INTO cash_deposits (name, amount, memo, updated_date) VALUES (%s,%s,%s,%s)",
    (data.get('name'), data.get('amount', 0), data.get('memo'), today)
    )
    cur.close()
    db.commit()
    db.close()
    return jsonify({'ok': True}), 201


@app.route('/api/cash-deposits/<int:rid>', methods=['PUT', 'DELETE'])
def api_cash_deposits_detail(rid):
    db = get_db()
    if request.method == 'PUT':
        data = request.json
        today = date.today().isoformat()
        cur = db.cursor()
        cur.execute(
        "UPDATE cash_deposits SET name=%s, amount=%s, memo=%s, updated_date=%s WHERE id=%s",
        (data.get('name'), data.get('amount', 0), data.get('memo'), today, rid)
        )
        # 수기 수정 시 자동 조정 전체 초기화 (이미 반영된 것 포함)
        cur.execute("DELETE FROM cash_auto_adjustments")
        cur.close()
        db.commit()
        db.close()
        return jsonify({'ok': True})
    cur = db.cursor()
    cur.execute("DELETE FROM cash_deposits WHERE id = %s", (rid,))
    cur.close()
    db.commit()
    db.close()
    return jsonify({'ok': True})

# ── API: 현금 자동 조정 ──────────────────────────────────────
@app.route('/api/cash-auto-adjustments', methods=['GET', 'DELETE'])
def api_cash_auto_adj():
    db = get_db()
    if request.method == 'DELETE':
        cur = db.cursor()
        cur.execute("DELETE FROM cash_auto_adjustments WHERE applied=FALSE")
        cur.close()
        db.commit(); db.close()
        return jsonify({'ok': True})

    cur = db.cursor()
    cur.execute("""
        SELECT * FROM cash_auto_adjustments
        ORDER BY adj_date DESC, id DESC
    """)
    rows = cur.fetchall()
    cur.close(); db.close()
    return jsonify(rows_to_list(rows))


@app.route('/api/cash-auto-adjustments/<int:aid>', methods=['DELETE'])
def api_cash_auto_adj_item(aid):
    db = get_db()
    cur = db.cursor()
    cur.execute("DELETE FROM cash_auto_adjustments WHERE id=%s", (aid,))
    cur.close()
    db.commit(); db.close()
    return jsonify({'ok': True})


@app.route('/api/cash-auto-adjustments/apply', methods=['POST'])
def api_cash_auto_adj_apply():
    """자동 조정을 현금 잔액에 반영 (잔액이 가장 큰 계좌에 적용)"""
    db = get_db()
    cur = db.cursor()

    cur.execute("SELECT COALESCE(SUM(amount),0) FROM cash_auto_adjustments WHERE applied=FALSE")
    total_adj = int(cur.fetchone()[0])

    if total_adj == 0:
        cur.close(); db.close()
        return jsonify({'ok': True, 'message': '반영할 조정 없음'})

    cur.execute("SELECT id, name, amount FROM cash_deposits ORDER BY amount DESC LIMIT 1")
    account = cur.fetchone()
    if not account:
        cur.close(); db.close()
        return jsonify({'error': '현금 계좌가 없습니다.'}), 400

    new_amount = int(account['amount']) + total_adj
    today = date.today().isoformat()
    cur.execute(
        "UPDATE cash_deposits SET amount=%s, updated_date=%s WHERE id=%s",
        (new_amount, today, account['id'])
    )
    cur.execute("UPDATE cash_auto_adjustments SET applied=TRUE WHERE applied=FALSE")

    cur.close()
    db.commit(); db.close()
    return jsonify({'ok': True, 'account': account['name'], 'adjusted': total_adj, 'new_amount': new_amount})


@app.route('/api/cash-auto-adjustments/recalc-foreign', methods=['POST'])
def api_recalc_foreign_cash_adj():
    """달러 ETF/주식 거래의 현금자동조정을 원화로 재계산"""
    db = get_db()
    cur = db.cursor()
    fixed = 0

    # ETF 외화 거래 재계산
    cur.execute("""
        SELECT t.id, t.price, t.quantity, t.fee, t.tx_date, t.exchange_rate,
               e.name, e.ticker
        FROM etf_tx t
        JOIN etf e ON e.id = t.etf_id
        WHERE t.tx_type IN ('buy','매수')
    """)
    for row in cur.fetchall():
        if not is_foreign_ticker(row['ticker']):
            continue
        rate = float(row['exchange_rate'] or 1.0)
        amt = -round((float(row['price']) * float(row['quantity']) + float(row['fee'] or 0)) * rate)
        ename = f"{row['name']}({row['ticker']})"
        c2 = db.cursor()
        _upsert_cash_adj(c2, 'etf_tx', row['id'], amt, f"{ename} ETF 매수", row['tx_date'])
        c2.close()
        fixed += 1

    # 주식 외화 거래 재계산
    cur.execute("""
        SELECT t.id, t.price, t.quantity, t.fee, t.tx_date, t.exchange_rate,
               s.name, s.ticker
        FROM stock_tx t
        JOIN stocks s ON s.id = t.stock_id
        WHERE t.tx_type IN ('buy','매수')
    """)
    for row in cur.fetchall():
        if not is_foreign_ticker(row['ticker']):
            continue
        rate = float(row['exchange_rate'] or 1.0)
        amt = -round((float(row['price']) * float(row['quantity']) + float(row['fee'] or 0)) * rate)
        sname = f"{row['name']}({row['ticker']})"
        c2 = db.cursor()
        _upsert_cash_adj(c2, 'stock_tx', row['id'], amt, f"{sname} 매수", row['tx_date'])
        c2.close()
        fixed += 1

    cur.close()
    db.commit(); db.close()
    return jsonify({'ok': True, 'fixed': fixed})

# ── API: 구분별 실현손익 ─────────────────────────────────────
@app.route('/api/stock-category-pnl')
def api_stock_category_pnl():
    """구분(category)별 실현손익 – period: monthly(최근12개월) | yearly | all"""
    category = request.args.get('category', '전체')
    period   = request.args.get('period', 'monthly')

    db  = get_db()
    ex_rate = get_current_exchange_rate()
    cur = db.cursor()

    cur.execute("""
        SELECT 'stock' as source, t.stock_id as asset_id, t.tx_date::text as tx_date, t.tx_type, t.price, t.quantity, t.fee, s.category, s.ticker, s.name
        FROM stock_tx t JOIN stocks s ON s.id = t.stock_id
        ORDER BY t.stock_id, t.tx_date, t.id
    """)
    stock_txs = [dict(r) for r in cur.fetchall()]

    cur.execute("""
        SELECT 'etf' as source, t.etf_id as asset_id, t.tx_date::text as tx_date, t.tx_type, t.price, t.quantity, t.fee, e.category, e.ticker, e.name
        FROM etf_tx t JOIN etf e ON e.id = t.etf_id
        ORDER BY t.etf_id, t.tx_date, t.id
    """)
    etf_txs = [dict(r) for r in cur.fetchall()]

    cur.execute("""
        SELECT id, name, listing_date::text as listing_date, ipo_price, quantity, realized_pnl, fee
        FROM ipo
        WHERE listing_date IS NOT NULL AND listing_date != ''
    """)
    ipo_txs = [dict(r) for r in cur.fetchall()]
    cur.close()
    db.close()

    from collections import defaultdict
    tx_by_asset = defaultdict(list)
    for tx in stock_txs + etf_txs:
        tx_by_asset[(tx['source'], tx['asset_id'])].append(tx)

    realized_records = []
    for (source, asset_id), txs in tx_by_asset.items():
        qty = 0.0
        avg_cost = 0.0
        for tx in txs:
            tq = float(tx['quantity'] or 0.0)
            tp = float(tx['price'] or 0.0)
            if tx['tx_type'] in ('buy', '매수'):
                new_qty = qty + tq
                if new_qty > 0:
                    avg_cost = (qty * avg_cost + tq * tp) / new_qty
                else:
                    avg_cost = 0.0
                qty = new_qty
            else: # sell
                fee_val = float(tx['fee'] or 0.0)
                pnl = (tp - avg_cost) * tq - fee_val
                cost = avg_cost * tq
                mul = ex_rate if is_foreign_ticker(tx['ticker']) else 1.0
                realized_records.append({
                    'date': tx['tx_date'],
                    'category': tx['category'],
                    'pnl': pnl * mul,
                    'cost': cost * mul
                })
                qty = max(0.0, qty - tq)
                if qty == 0.0:
                    avg_cost = 0.0

    for r in ipo_txs:
        pnl = float(r['realized_pnl'] or 0.0) - float(r['fee'] or 0.0)
        cost = float(r['ipo_price'] or 0.0) * float(r['quantity'] or 0.0)
        realized_records.append({
            'date': r['listing_date'],
            'category': '공모주',
            'pnl': pnl,
            'cost': cost
        })

    # Determine monthly period keys if needed for win/loss filtering
    if period == 'monthly':
        today = date.today()
        keys = []
        for i in range(11, -1, -1):
            mo = today.month - i
            yr = today.year
            while mo <= 0:
                mo += 12; yr -= 1
            keys.append(f"{yr}-{mo:02d}")
    else:
        keys = []

    win_count = 0
    loss_count = 0
    for r in realized_records:
        if category and category != '전체' and r['category'] != category:
            continue
        if period == 'monthly':
            ym = r['date'][:7]
            if ym not in keys:
                continue
        if r['pnl'] > 0:
            win_count += 1
        elif r['pnl'] < 0:
            loss_count += 1

    period_map = defaultdict(lambda: {'pnl': 0.0, 'cost': 0.0})
    for r in realized_records:
        if category and category != '전체' and r['category'] != category:
            continue
        date_str = r['date']
        pkey = date_str[:4] if period == 'yearly' else date_str[:7]
        period_map[pkey]['pnl'] += r['pnl']
        period_map[pkey]['cost'] += r['cost']

    data_map = {k: v for k, v in period_map.items()}

    if period == 'monthly':
        # cumulative includes all history before our 12-month window
        cumulative = sum(v['pnl'] for k, v in data_map.items() if k < keys[0])
    else:
        keys = sorted(data_map.keys())
        cumulative = 0

    result = []
    for k in keys:
        d = data_map.get(k, {'pnl': 0, 'cost': 0})
        cumulative += d['pnl']
        result.append({
            'label':          k,
            'realized_pnl':   round(d['pnl']),
            'cost_basis':     round(d['cost']),
            'return_rate':    round(d['pnl'] / d['cost'] * 100, 2) if d['cost'] else 0,
            'cumulative_pnl': round(cumulative),
        })

    return jsonify({
        'data': result,
        'win_count': win_count,
        'loss_count': loss_count
    })


# ── API: 공모주 실현손익 차트 ──────────────────────────────────
@app.route('/api/ipo-pnl')
def api_ipo_pnl():
    """공모주 월별/연도별/전체 실현손익 및 누계"""
    period = request.args.get('period', 'monthly')
    date_fmt = 'YYYY' if period == 'yearly' else 'YYYY-MM'

    db = get_db()
    cur = db.cursor()
    cur.execute(f"""
        SELECT
            to_char(listing_date::date, '{date_fmt}') AS period_key,
            COALESCE(SUM(realized_pnl - fee), 0) AS net_pnl,
            COALESCE(SUM(realized_pnl), 0) AS realized_pnl,
            COALESCE(SUM(fee), 0) AS total_fee,
            COALESCE(SUM(ipo_price * quantity), 0) AS cost_basis
        FROM ipo
        WHERE listing_date IS NOT NULL AND listing_date != ''
        GROUP BY period_key
        ORDER BY period_key
    """)
    rows = cur.fetchall()

    # 개별 IPO 건수 기반 승/패 집계 (기간 필터 적용)
    if period == 'monthly':
        today = date.today()
        win_loss_keys = []
        for i in range(11, -1, -1):
            mo = today.month - i; yr = today.year
            while mo <= 0: mo += 12; yr -= 1
            win_loss_keys.append(f"{yr}-{mo:02d}")
        cur.execute(f"""
            SELECT COUNT(*) FILTER (WHERE (realized_pnl - fee) > 0) AS win_count,
                   COUNT(*) FILTER (WHERE (realized_pnl - fee) < 0) AS loss_count
            FROM ipo
            WHERE listing_date IS NOT NULL AND listing_date != ''
              AND to_char(listing_date::date, 'YYYY-MM') = ANY(%s)
        """, (win_loss_keys,))
    else:
        cur.execute("""
            SELECT COUNT(*) FILTER (WHERE (realized_pnl - fee) > 0) AS win_count,
                   COUNT(*) FILTER (WHERE (realized_pnl - fee) < 0) AS loss_count
            FROM ipo
            WHERE listing_date IS NOT NULL AND listing_date != ''
        """)
    wl = cur.fetchone()
    win_count  = int(wl['win_count']  or 0)
    loss_count = int(wl['loss_count'] or 0)

    cur.close()
    db.close()

    data_map = {r['period_key']: {
        'net_pnl': float(r['net_pnl']),
        'realized_pnl': float(r['realized_pnl']),
        'fee': float(r['total_fee']),
        'cost': float(r['cost_basis'])
    } for r in rows}

    if period == 'monthly':
        today = date.today()
        keys = []
        for i in range(11, -1, -1):
            mo = today.month - i
            yr = today.year
            while mo <= 0:
                mo += 12; yr -= 1
            keys.append(f"{yr}-{mo:02d}")
        cumulative = sum(v['net_pnl'] for k, v in data_map.items() if k < keys[0])
    else:
        keys = sorted(data_map.keys())
        cumulative = 0

    result = []
    for k in keys:
        d = data_map.get(k, {'net_pnl': 0, 'realized_pnl': 0, 'fee': 0, 'cost': 0})
        cumulative += d['net_pnl']
        result.append({
            'label': k,
            'net_pnl': round(d['net_pnl']),
            'realized_pnl': round(d['realized_pnl']),
            'fee': round(d['fee']),
            'cost_basis': round(d['cost']),
            'cumulative_pnl': round(cumulative),
        })

    return jsonify({
        'data': result,
        'win_count':  win_count,
        'loss_count': loss_count,
    })


# ── API: 월별 실현손익 ───────────────────────────────────────
@app.route('/api/investment-monthly')
def api_investment_monthly():
    """최근 12개월 월별 실현손익 및 누계 (주식+ETF+공모주 합산)"""
    db = get_db()
    ex_rate = get_current_exchange_rate()

    cur = db.cursor()
    cur.execute("""
        WITH avg_costs AS (
            SELECT stock_id,
                SUM(CASE WHEN tx_type IN ('buy','매수') THEN price*quantity ELSE 0 END) /
                NULLIF(SUM(CASE WHEN tx_type IN ('buy','매수') THEN quantity ELSE 0 END), 0) AS avg_cost
            FROM stock_tx GROUP BY stock_id
        )
        SELECT to_char(t.tx_date::date, 'YYYY-MM') AS ym,
            COALESCE(SUM(((t.price - ac.avg_cost) * t.quantity - t.fee) * (CASE WHEN s.ticker IS NOT NULL AND s.ticker != '' AND s.ticker !~ '^[0-9]{6}$' THEN %s ELSE 1 END)), 0) AS realized_pnl
        FROM stock_tx t
        JOIN avg_costs ac ON ac.stock_id = t.stock_id
        JOIN stocks s ON s.id = t.stock_id
        WHERE t.tx_type IN ('sell','매도')
        GROUP BY ym ORDER BY ym
    """, (ex_rate,))
    stocks_by_month = {r['ym']: float(r['realized_pnl']) for r in cur.fetchall()}
    cur.close()

    cur = db.cursor()
    cur.execute("""
        WITH avg_costs AS (
            SELECT etf_id,
                SUM(CASE WHEN tx_type IN ('buy','매수') THEN price*quantity ELSE 0 END) /
                NULLIF(SUM(CASE WHEN tx_type IN ('buy','매수') THEN quantity ELSE 0 END), 0) AS avg_cost
            FROM etf_tx GROUP BY etf_id
        )
        SELECT to_char(t.tx_date::date, 'YYYY-MM') AS ym,
            COALESCE(SUM(((t.price - ac.avg_cost) * t.quantity - t.fee) * (CASE WHEN e.ticker IS NOT NULL AND e.ticker != '' AND e.ticker !~ '^[0-9]{6}$' THEN %s ELSE 1 END)), 0) AS realized_pnl
        FROM etf_tx t
        JOIN avg_costs ac ON ac.etf_id = t.etf_id
        JOIN etf e ON e.id = t.etf_id
        WHERE t.tx_type IN ('sell','매도')
        GROUP BY ym ORDER BY ym
    """, (ex_rate,))
    etf_by_month = {r['ym']: float(r['realized_pnl']) for r in cur.fetchall()}
    cur.close()

    cur = db.cursor()
    cur.execute("""
        SELECT to_char(listing_date::date, 'YYYY-MM') AS ym,
            COALESCE(SUM(realized_pnl), 0) AS realized_pnl
        FROM ipo
        GROUP BY ym ORDER BY ym
    """)
    ipo_by_month = {r['ym']: float(r['realized_pnl']) for r in cur.fetchall()}
    cur.close()
    db.close()

    today = date.today()
    months = []
    cumulative = 0
    for i in range(11, -1, -1):
        mo = today.month - i
        yr = today.year
        while mo <= 0:
            mo += 12
            yr -= 1
        ym = f"{yr}-{mo:02d}"
        pnl = round(stocks_by_month.get(ym, 0) + etf_by_month.get(ym, 0) + ipo_by_month.get(ym, 0))
        cumulative += pnl
        months.append({'ym': ym, 'realized_pnl': pnl, 'cumulative_pnl': cumulative})

    return jsonify(months)


def _save_daily_snapshot(db):
    """
    오늘 날짜의 순자산을 계산하여 daily_snapshots에 upsert.
    대시보드 또는 tech-tree 조회 시 자동 호출되어 매일 1회 기록됨.
    """
    today_str = date.today().isoformat()
    ex_rate = get_current_exchange_rate()
    stocks_val, _, etf_val, _ = get_stocks_etf_totals(db, ex_rate)
    
    cur = db.cursor()
    cur.execute("SELECT COALESCE(SUM(current_price * quantity),0) FROM crypto")
    crypto_val = float(cur.fetchone()[0] or 0)
    cur.close()
    
    re_total_price = get_real_estate_value(db)
    
    cur = db.cursor()
    cur.execute("""
    SELECT COALESCE(SUM(deposit), 0) FROM tenant_contracts
    WHERE id IN (SELECT MAX(id) FROM tenant_contracts WHERE real_estate_id IS NOT NULL GROUP BY real_estate_id)
    """)
    re_total_deposit = float(cur.fetchone()[0] or 0)
    cur.close()

    cur = db.cursor()
    cur.execute("SELECT COALESCE(SUM(deposit), 0) FROM residence")
    residence_deposit = float(cur.fetchone()[0] or 0)
    cur.close()

    re_val = re_total_price - re_total_deposit + residence_deposit

    cur = db.cursor()
    cur.execute("SELECT COALESCE(SUM(amount),0) FROM cash_deposits")
    cash_val = float(cur.fetchone()[0] or 0)
    cur.close()

    cur = db.cursor()
    cur.execute("SELECT COALESCE(SUM(accumulated),0) FROM pension")
    pension_val = float(cur.fetchone()[0] or 0)
    cur.close()

    cur = db.cursor()
    cur.execute("SELECT COALESCE(SUM(remaining),0) FROM loans")
    loan_total = float(cur.fetchone()[0] or 0)
    cur.close()

    cash_int = int(cash_val or 0)
    stocks_int = int(stocks_val or 0) + int(etf_val or 0)
    re_int = int(re_val or 0)
    crypto_int = int(crypto_val or 0)
    pension_int = int(pension_val or 0)
    
    total = cash_int + stocks_int + re_int + crypto_int + pension_int
    net_worth = total - int(loan_total or 0)

    cur = db.cursor()
    cur.execute("""
        INSERT INTO daily_snapshots (day, cash, stocks, real_estate, crypto, pension, total, net_worth, updated_at)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, CURRENT_TIMESTAMP)
        ON CONFLICT (day) DO UPDATE SET
            cash=excluded.cash, stocks=excluded.stocks,
            real_estate=excluded.real_estate, crypto=excluded.crypto,
            pension=excluded.pension, total=excluded.total,
            net_worth=excluded.net_worth, updated_at=excluded.updated_at
    """, (today_str, cash_int, stocks_int, re_int, crypto_int, pension_int, total, net_worth))
    cur.close()


@app.route('/api/networth-history')
def api_networth_history():
    """
    순자산 변화 이력 반환. 모든 기간 daily_snapshots 기반.

    daily   → 최근 90일 (1행=1일)
    weekly  → 최근 52주, 주별 마지막 행
    monthly → 최근 24개월, 월별 마지막 행
    yearly  → 최근 10년, 연별 마지막 행
    """
    period = request.args.get('period', 'daily')
    try:
        days = int(request.args.get('days', 7))
        if days not in (7, 15, 30):
            days = 7
    except (ValueError, TypeError):
        days = 7
    db     = get_db()
    cur    = db.cursor()

    asset_cols = "cash, stocks, real_estate, crypto, pension"
    asset_lag  = "LAG(cash) OVER (ORDER BY {o}), LAG(stocks) OVER (ORDER BY {o}), LAG(real_estate) OVER (ORDER BY {o}), LAG(crypto) OVER (ORDER BY {o}), LAG(pension) OVER (ORDER BY {o})"

    if period == 'daily':
        cur.execute("""
            WITH base AS (
                SELECT day, net_worth, cash, stocks, real_estate, crypto, pension
                FROM daily_snapshots
                WHERE day >= CURRENT_DATE - INTERVAL '{days} days'
            )
            SELECT
                day::text AS label,
                net_worth,
                cash, stocks, real_estate, crypto, pension,
                net_worth - COALESCE(LAG(net_worth) OVER (ORDER BY day), net_worth) AS change_amt,
                ROUND(COALESCE(
                    (net_worth - LAG(net_worth) OVER (ORDER BY day))::numeric
                    / NULLIF(LAG(net_worth) OVER (ORDER BY day), 0) * 100, 0
                ), 2) AS change_pct,
                COALESCE(LAG(cash)        OVER (ORDER BY day), cash)        AS prev_cash,
                COALESCE(LAG(stocks)      OVER (ORDER BY day), stocks)      AS prev_stocks,
                COALESCE(LAG(real_estate) OVER (ORDER BY day), real_estate) AS prev_real_estate,
                COALESCE(LAG(crypto)      OVER (ORDER BY day), crypto)      AS prev_crypto,
                COALESCE(LAG(pension)     OVER (ORDER BY day), pension)     AS prev_pension
            FROM base
            ORDER BY day
        """.format(days=days))

    elif period == 'weekly':
        cur.execute("""
            WITH weekly AS (
                SELECT
                    DATE_TRUNC('week', day)::date AS week_start,
                    net_worth, cash, stocks, real_estate, crypto, pension,
                    ROW_NUMBER() OVER (PARTITION BY DATE_TRUNC('week', day) ORDER BY day DESC) AS rn
                FROM daily_snapshots
                WHERE day >= CURRENT_DATE - INTERVAL '52 weeks'
            )
            SELECT
                week_start::text AS label,
                net_worth, cash, stocks, real_estate, crypto, pension,
                net_worth - COALESCE(LAG(net_worth) OVER (ORDER BY week_start), net_worth) AS change_amt,
                ROUND(COALESCE(
                    (net_worth - LAG(net_worth) OVER (ORDER BY week_start))::numeric
                    / NULLIF(LAG(net_worth) OVER (ORDER BY week_start), 0) * 100, 0
                ), 2) AS change_pct,
                COALESCE(LAG(cash)        OVER (ORDER BY week_start), cash)        AS prev_cash,
                COALESCE(LAG(stocks)      OVER (ORDER BY week_start), stocks)      AS prev_stocks,
                COALESCE(LAG(real_estate) OVER (ORDER BY week_start), real_estate) AS prev_real_estate,
                COALESCE(LAG(crypto)      OVER (ORDER BY week_start), crypto)      AS prev_crypto,
                COALESCE(LAG(pension)     OVER (ORDER BY week_start), pension)     AS prev_pension
            FROM weekly
            WHERE rn = 1
            ORDER BY week_start
        """)

    elif period == 'monthly':
        cur.execute("""
            WITH monthly AS (
                SELECT
                    TO_CHAR(day, 'YYYY-MM') AS ym,
                    net_worth, cash, stocks, real_estate, crypto, pension,
                    ROW_NUMBER() OVER (PARTITION BY TO_CHAR(day, 'YYYY-MM') ORDER BY day DESC) AS rn
                FROM daily_snapshots
                WHERE day >= CURRENT_DATE - INTERVAL '24 months'
            )
            SELECT
                ym AS label,
                net_worth, cash, stocks, real_estate, crypto, pension,
                net_worth - COALESCE(LAG(net_worth) OVER (ORDER BY ym), net_worth) AS change_amt,
                ROUND(COALESCE(
                    (net_worth - LAG(net_worth) OVER (ORDER BY ym))::numeric
                    / NULLIF(LAG(net_worth) OVER (ORDER BY ym), 0) * 100, 0
                ), 2) AS change_pct,
                COALESCE(LAG(cash)        OVER (ORDER BY ym), cash)        AS prev_cash,
                COALESCE(LAG(stocks)      OVER (ORDER BY ym), stocks)      AS prev_stocks,
                COALESCE(LAG(real_estate) OVER (ORDER BY ym), real_estate) AS prev_real_estate,
                COALESCE(LAG(crypto)      OVER (ORDER BY ym), crypto)      AS prev_crypto,
                COALESCE(LAG(pension)     OVER (ORDER BY ym), pension)     AS prev_pension
            FROM monthly
            WHERE rn = 1
            ORDER BY ym
        """)

    elif period == 'yearly':
        cur.execute("""
            WITH yearly AS (
                SELECT
                    TO_CHAR(day, 'YYYY') AS yr,
                    net_worth, cash, stocks, real_estate, crypto, pension,
                    ROW_NUMBER() OVER (PARTITION BY TO_CHAR(day, 'YYYY') ORDER BY day DESC) AS rn
                FROM daily_snapshots
                WHERE day >= CURRENT_DATE - INTERVAL '10 years'
            )
            SELECT
                yr AS label,
                net_worth, cash, stocks, real_estate, crypto, pension,
                net_worth - COALESCE(LAG(net_worth) OVER (ORDER BY yr), net_worth) AS change_amt,
                ROUND(COALESCE(
                    (net_worth - LAG(net_worth) OVER (ORDER BY yr))::numeric
                    / NULLIF(LAG(net_worth) OVER (ORDER BY yr), 0) * 100, 0
                ), 2) AS change_pct,
                COALESCE(LAG(cash)        OVER (ORDER BY yr), cash)        AS prev_cash,
                COALESCE(LAG(stocks)      OVER (ORDER BY yr), stocks)      AS prev_stocks,
                COALESCE(LAG(real_estate) OVER (ORDER BY yr), real_estate) AS prev_real_estate,
                COALESCE(LAG(crypto)      OVER (ORDER BY yr), crypto)      AS prev_crypto,
                COALESCE(LAG(pension)     OVER (ORDER BY yr), pension)     AS prev_pension
            FROM yearly
            WHERE rn = 1
            ORDER BY yr
        """)

    else:
        cur.close()
        db.close()
        return jsonify({'rows': [], 'summary': {'current': 0, 'change_amt': 0, 'change_pct': 0}})

    rows = cur.fetchall()
    cur.close()
    db.close()

    valid = [r for r in rows if r['net_worth'] is not None]
    first_nw   = float(valid[0]['net_worth'])  if valid else 0
    current_nw = float(valid[-1]['net_worth']) if valid else 0
    total_change     = current_nw - first_nw
    total_change_pct = round(total_change / first_nw * 100, 2) if first_nw else 0

    return jsonify({
        'rows': rows_to_list(rows),
        'summary': {
            'current':    current_nw,
            'change_amt': total_change,
            'change_pct': total_change_pct,
        }
    })


@app.route('/api/etf-invest-plan', methods=['POST'])
def api_etf_invest_plan():
    """
    지수/레버리지 ETF 분할매수 계획 계산.
    """
    d = request.json or {}
    etf_id       = d.get('etf_id')
    strategy     = d.get('strategy', 'dca')
    total_budget = float(d.get('total_budget', 0))
    periods      = int(d.get('periods', 12))
    step         = float(d.get('drawdown_step', 5))
    split_count  = int(d.get('split_count', 5))
    
    db = get_db()
    
    # 현재가 조회
    if etf_id:
        cur = db.cursor()
        cur.execute("SELECT name, ticker, current_price FROM etf WHERE id=%s", (etf_id,))
        row = cur.fetchone(); cur.close()
        current_price = float(d.get('current_price') or (row['current_price'] if row else 1) or 1)
        etf_name = row['name'] if row else ''
    else:
        current_price = float(d.get('current_price', 1))
        etf_name = d.get('etf_name', '')
    
    db.close()
    
    plan = []
    
    if strategy == 'dca':
        # ── 정액 분할매수 ──
        monthly_amount = total_budget / periods
        for i in range(periods):
            plan.append({
                'step':        i + 1,
                'label':       f"{i + 1}개월차",
                'trigger':     '매월 정기',
                'amount':      round(monthly_amount),
                'price':       round(current_price),  # DCA는 가격 무관
                'shares':      round(monthly_amount / current_price, 4),
                'cumulative':  round(monthly_amount * (i + 1)),
                'drawdown_pct': 0,
            })
    
    elif strategy == 'drawdown':
        # ── 하락률 비례 역피라미딩 ──
        # 가중치: 1구간=10%, 이후 매 구간마다 균등 배분 증가
        # 총합이 100%가 되도록 정규화
        raw_weights = [1]
        for i in range(1, split_count):
            raw_weights.append(raw_weights[-1] * 1.5)  # 1.5배씩 증가
        total_w = sum(raw_weights)
        weights = [w / total_w for w in raw_weights]
        
        cumulative = 0
        for i in range(split_count):
            dd_pct = i * step  # 하락률 (0, 5, 10, 15, 20, ...)
            trigger_price = round(current_price * (1 - dd_pct / 100))
            amount = round(total_budget * weights[i])
            cumulative += amount
            plan.append({
                'step':         i + 1,
                'label':        f"{'현재가' if i == 0 else f'-{dd_pct:.0f}% 하락 시'}",
                'trigger':      f"{'즉시 매수' if i == 0 else f'{trigger_price:,}원 이하'}",
                'trigger_price': trigger_price,
                'drawdown_pct': dd_pct,
                'weight_pct':   round(weights[i] * 100, 1),
                'amount':       amount,
                'price':        trigger_price,
                'shares':       round(amount / trigger_price, 4) if trigger_price > 0 else 0,
                'cumulative':   cumulative,
            })
    
    # 요약 정보
    total_shares = sum(p['shares'] for p in plan)
    avg_price    = round(total_budget / total_shares) if total_shares > 0 else 0
    
    return jsonify({
        'etf_name':     etf_name,
        'strategy':     strategy,
        'total_budget': round(total_budget),
        'plan':         plan,
        'summary': {
            'total_shares': round(total_shares, 4),
            'avg_price':    avg_price,
            'breakeven':    avg_price,  # 평균단가 = 손익분기점
        }
    })


# ── [NEW] 생애주기별 자산 시뮬레이터 페이지 및 API ───────────────────
@app.route('/lifecycle')
def lifecycle():
    return render_template('lifecycle.html')


@app.route('/api/lifecycle-profile', methods=['GET', 'POST'])
def api_lifecycle_profile():
    db = get_db()
    if request.method == 'GET':
        cur = db.cursor()
        cur.execute("SELECT * FROM lifecycle_profile ORDER BY id")
        rows = cur.fetchall()
        cur.close()
        db.close()
        return jsonify(rows_to_list(rows))
    d = request.json or {}
    cur = db.cursor()
    cur.execute(
        "INSERT INTO lifecycle_profile (role, name, birth_year) VALUES (%s,%s,%s)",
        (d.get('role'), d.get('name'), d.get('birth_year'))
    )
    cur.close()
    db.commit()
    db.close()
    return jsonify({'ok': True}), 201


@app.route('/api/lifecycle-profile/<int:rid>', methods=['PUT', 'DELETE'])
def api_lifecycle_profile_detail(rid):
    db = get_db()
    if request.method == 'DELETE':
        cur = db.cursor()
        cur.execute("DELETE FROM lifecycle_profile WHERE id=%s", (rid,))
        cur.close()
        db.commit()
        db.close()
        return jsonify({'ok': True})
    d = request.json or {}
    cur = db.cursor()
    cur.execute(
        "UPDATE lifecycle_profile SET role=%s, name=%s, birth_year=%s WHERE id=%s",
        (d.get('role'), d.get('name'), d.get('birth_year'), rid)
    )
    cur.close()
    db.commit()
    db.close()
    return jsonify({'ok': True})


@app.route('/api/lifecycle-events', methods=['GET', 'POST'])
def api_lifecycle_events():
    db = get_db()
    if request.method == 'GET':
        cur = db.cursor()
        cur.execute("SELECT * FROM lifecycle_events ORDER BY event_year, id")
        rows = cur.fetchall()
        cur.close()
        db.close()
        return jsonify(rows_to_list(rows))
    d = request.json or {}
    cur = db.cursor()
    cur.execute(
        "INSERT INTO lifecycle_events (event_year, event_type, asset_name, amount, memo) "
        "VALUES (%s,%s,%s,%s,%s)",
        (d.get('event_year'), d.get('event_type'),
         d.get('asset_name'), d.get('amount', 0), d.get('memo'))
    )
    cur.close()
    db.commit()
    db.close()
    return jsonify({'ok': True}), 201


@app.route('/api/lifecycle-events/<int:rid>', methods=['PUT', 'DELETE'])
def api_lifecycle_events_detail(rid):
    db = get_db()
    if request.method == 'DELETE':
        cur = db.cursor()
        cur.execute("DELETE FROM lifecycle_events WHERE id=%s", (rid,))
        cur.close()
        db.commit()
        db.close()
        return jsonify({'ok': True})
    d = request.json or {}
    cur = db.cursor()
    cur.execute("""
        UPDATE lifecycle_events 
        SET event_year=%s, event_type=%s, asset_name=%s, amount=%s, memo=%s 
        WHERE id=%s
    """, (d.get('event_year'), d.get('event_type'), d.get('asset_name'), d.get('amount', 0), d.get('memo'), rid))
    cur.close()
    db.commit()
    db.close()
    return jsonify({'ok': True})


@app.route('/api/lifecycle-settings', methods=['GET', 'POST'])
def api_lifecycle_settings():
    db = get_db()
    if request.method == 'GET':
        cur = db.cursor()
        cur.execute("SELECT * FROM lifecycle_settings LIMIT 1")
        row = cur.fetchone()
        cur.close()
        db.close()
        return jsonify(dict(row) if row else {})
    d = request.json or {}
    cur = db.cursor()
    cur.execute("""
        UPDATE lifecycle_settings SET
            sim_years=%s, annual_return_stocks=%s, annual_return_re=%s,
            annual_return_cash=%s, annual_expense_growth=%s,
            override_annual_inflow=%s, updated_at=CURRENT_TIMESTAMP
        WHERE id=1
    """, (
        d.get('sim_years', 30), d.get('annual_return_stocks', 7.00),
        d.get('annual_return_re', 3.00), d.get('annual_return_cash', 2.00),
        d.get('annual_expense_growth', 2.00), d.get('override_annual_inflow')
    ))
    cur.close()
    db.commit()
    db.close()
    return jsonify({'ok': True})


@app.route('/api/lifecycle-simulate')
def api_lifecycle_simulate():
    db = get_db()
    today = date.today()
    current_year = today.year

    # ── 1. 시뮬레이션 설정 로드 ──
    cur = db.cursor()
    cur.execute("SELECT * FROM lifecycle_settings LIMIT 1")
    settings = dict(cur.fetchone() or {})
    cur.close()
    
    sim_years    = settings.get('sim_years', 30)
    r_stocks     = float(settings.get('annual_return_stocks', 7.00)) / 100
    r_re         = float(settings.get('annual_return_re', 3.00)) / 100
    r_cash       = float(settings.get('annual_return_cash', 2.00)) / 100
    exp_growth   = float(settings.get('annual_expense_growth', 2.00)) / 100
    override_inflow = settings.get('override_annual_inflow')

    # ── 2. 현재 자산 기준점 ──
    ex_rate = get_current_exchange_rate()
    stocks_val, _, etf_val, _ = get_stocks_etf_totals(db, ex_rate)
    
    cur = db.cursor()
    cur.execute("SELECT COALESCE(SUM(amount),0) FROM cash_deposits")
    base_cash = float(cur.fetchone()[0] or 0)
    cur.close()

    base_stocks = float(stocks_val or 0) + float(etf_val or 0)

    cur = db.cursor()
    cur.execute("SELECT COALESCE(SUM(current_price * quantity),0) FROM crypto")
    base_crypto = float(cur.fetchone()[0] or 0)
    cur.close()

    re_total_price = get_real_estate_value(db)

    cur = db.cursor()
    cur.execute("""
    SELECT COALESCE(SUM(deposit), 0) FROM tenant_contracts
    WHERE id IN (SELECT MAX(id) FROM tenant_contracts WHERE real_estate_id IS NOT NULL GROUP BY real_estate_id)
    """)
    re_total_deposit = float(cur.fetchone()[0] or 0)
    cur.close()

    cur = db.cursor()
    cur.execute("SELECT COALESCE(SUM(deposit), 0) FROM residence")
    residence_deposit = float(cur.fetchone()[0] or 0)
    cur.close()

    base_re = re_total_price - re_total_deposit + residence_deposit

    cur = db.cursor()
    cur.execute("SELECT COALESCE(SUM(accumulated),0) FROM pension")
    base_pension = float(cur.fetchone()[0] or 0)
    cur.close()

    cur = db.cursor()
    cur.execute("SELECT COALESCE(SUM(monthly_payment),0) FROM pension")
    base_pension_monthly = float(cur.fetchone()[0] or 0)
    cur.close()

    cur = db.cursor()
    cur.execute("SELECT COALESCE(SUM(remaining),0) FROM loans")
    base_loans = float(cur.fetchone()[0] or 0)
    cur.close()

    # ── 3. 연평균 유입속도 (순유입: 수입 - 지출) ──
    if override_inflow is not None:
        annual_net_inflow = float(override_inflow)
        monthly_exp_avg = 0.0
        monthly_card_avg = 0.0
        loan_repayment = 0.0
    else:
        avg = _calc_annual_avg_income(db)
        
        cur = db.cursor()
        cur.execute("""
            SELECT COALESCE(AVG(monthly_exp),0) FROM (
                SELECT to_char(date::date,'YYYY-MM') as ym, SUM(amount) as monthly_exp
                FROM budget
                WHERE date >= CURRENT_DATE - INTERVAL '12 months'
                GROUP BY ym
            ) sub
        """)
        monthly_exp_avg = float(cur.fetchone()[0] or 0)
        cur.close()
        
        cur = db.cursor()
        cur.execute("""
            SELECT COALESCE(AVG(monthly_card),0) FROM (
                SELECT to_char(date::date,'YYYY-MM') as ym, SUM(amount) as monthly_card
                FROM card_tx
                WHERE date >= CURRENT_DATE - INTERVAL '12 months' AND budget_id IS NULL
                GROUP BY ym
            ) sub
        """)
        monthly_card_avg = float(cur.fetchone()[0] or 0)
        cur.close()
        
        cur = db.cursor()
        cur.execute("SELECT COALESCE(SUM(monthly_payment), 0) FROM loans")
        loan_repayment = float(cur.fetchone()[0] or 0)
        cur.close()

        annual_expense_est = (monthly_exp_avg + monthly_card_avg + loan_repayment) * 12
        annual_net_inflow = (avg['labor_annual'] + avg['passive_annual']) - annual_expense_est

    # ── 4. 이벤트 로드 ──
    cur = db.cursor()
    cur.execute("SELECT * FROM lifecycle_events ORDER BY event_year, id")
    events = cur.fetchall()
    cur.close()
    event_map = {}
    for e in events:
        yr = e['event_year']
        if yr not in event_map:
            event_map[yr] = []
        event_map[yr].append(dict(e))

    # ── 5. 가족 구성 ──
    cur = db.cursor()
    cur.execute("SELECT * FROM lifecycle_profile ORDER BY id")
    family = cur.fetchall()
    cur.close()

    # ── 6. 연도별 시뮬레이션 루프 ──
    result = []
    
    cash    = base_cash
    stocks  = base_stocks
    re      = base_re
    crypto  = base_crypto
    pension = base_pension
    loans   = base_loans

    retired = False

    for i in range(sim_years + 1):
        year = current_year + i

        family_ages = [
            {
                'name': m['name'],
                'role': m['role'],
                'age':  year - m['birth_year']
            }
            for m in family
        ]

        year_events = event_map.get(year, [])
        event_cash_delta = 0
        for evt in year_events:
            etype  = evt['event_type']
            amount = float(evt['amount'] or 0)
            if etype == 'sell_realestate':
                re   -= amount
                cash += amount
                event_cash_delta += amount
            elif etype == 'sell_stock':
                stocks -= amount
                cash   += amount
                event_cash_delta += amount
            elif etype == 'buy_asset':
                cash   -= amount
                stocks += amount
                event_cash_delta -= amount
            elif etype == 'extra_income':
                cash += amount
                event_cash_delta += amount
            elif etype == 'extra_expense':
                cash -= amount
                event_cash_delta -= amount
            elif etype == 'loan_payoff':
                cash -= amount
                loans = max(0.0, loans - amount)
                event_cash_delta -= amount
            elif etype == 'retire':
                retired = True

        current_year_inflow = annual_net_inflow
        if retired:
            if override_inflow is None:
                avg = _calc_annual_avg_income(db)
                current_year_inflow = avg['passive_annual'] - (monthly_exp_avg + monthly_card_avg + loan_repayment) * 12

        total = cash + stocks + re + crypto + pension
        net   = total - loans

        result.append({
            'year':         year,
            'family_ages':  family_ages,
            'cash':         round(cash),
            'stocks':       round(stocks),
            'real_estate':  round(re),
            'crypto':       round(crypto),
            'pension':      round(pension),
            'loans':        round(loans),
            'total_assets': round(total),
            'net_worth':    round(net),
            'events':       year_events,
            'event_cash_delta': round(event_cash_delta),
        })

        if i < sim_years:
            cash    = cash    * (1 + r_cash)    + current_year_inflow
            stocks  = stocks  * (1 + r_stocks)
            re      = re      * (1 + r_re)
            crypto  = crypto  * (1 + r_stocks)
            pension = pension + base_pension_monthly * 12
            annual_net_inflow *= (1 - exp_growth)

    db.close()
    return jsonify({
        'simulation':      result,
        'annual_net_inflow': round(annual_net_inflow),
        'settings':        settings,
    })


# ── API: 대시보드 집계 ───────────────────────────────────────
@app.route('/api/dashboard')
@cache.cached(timeout=120, query_string=True)
def api_dashboard():
    try:
        return _retry_on_db_error(_api_dashboard_inner)()
    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'trace': traceback.format_exc()}), 200

def _api_dashboard_inner():
    db = get_db()
    today = date.today()
    year  = request.args.get('year',  today.strftime('%Y'))
    month = request.args.get('month', today.strftime('%m'))
    ym    = f"{year}-{month.zfill(2)}"

    # 소득 현황 (오늘 이후 날짜의 반복 수입 등은 제외)
    # 근로소득: 급여, 사업소득
    cur = db.cursor()
    cur.execute(
    "SELECT COALESCE(SUM(amount),0) FROM income "
    "WHERE to_char(date::date, 'YYYY-MM') = %s AND category IN ('급여', '사업소득') AND date <= CURRENT_DATE",
    (ym,)
    )
    labor_inc = cur.fetchone()[0]
    cur.close()

    # 자생소득: 그 외 모든 수입
    cur = db.cursor()
    cur.execute(
    "SELECT COALESCE(SUM(amount),0) FROM income "
    "WHERE to_char(date::date, 'YYYY-MM') = %s AND category NOT IN ('급여', '사업소득') AND date <= CURRENT_DATE",
    (ym,)
    )
    passive_inc = cur.fetchone()[0]
    cur.close()

    # 이번달 수입 합계
    cur = db.cursor()
    cur.execute(
    "SELECT COALESCE(SUM(amount),0) as total FROM income WHERE to_char(date::date, 'YYYY-MM') = %s AND date <= CURRENT_DATE", (ym,)
    )
    income_total = cur.fetchone()['total']
    cur.close()

    # 이번달 지출 합계
    cur = db.cursor()
    cur.execute(
    "SELECT COALESCE(SUM(amount),0) as total FROM budget WHERE to_char(date::date, 'YYYY-MM') = %s AND date <= CURRENT_DATE", (ym,)
    )
    expense_total = cur.fetchone()['total']
    cur.close()

    # 이번달 카드 지출
    cur = db.cursor()
    cur.execute(
    "SELECT COALESCE(SUM(amount),0) as total FROM card_tx WHERE to_char(date::date, 'YYYY-MM') = %s AND date <= CURRENT_DATE AND budget_id IS NULL", (ym,)
    )
    card_total = cur.fetchone()['total']
    cur.close()

    ex_rate = get_current_exchange_rate()
    stocks_val, stocks_cost, etf_val, etf_cost = get_stocks_etf_totals(db, ex_rate)

    # 코인 평가액
    cur = db.cursor()
    cur.execute(
    "SELECT COALESCE(SUM(current_price * quantity),0) as val FROM crypto"
    )
    crypto_val = float(cur.fetchone()['val'] or 0)
    cur.close()

    # 부동산 현재가 (시세 - 임대보증금 + 거주보증금)
    re_total_price = get_real_estate_value(db)
    cur = db.cursor()
    cur.execute("""
    SELECT COALESCE(SUM(deposit), 0) FROM tenant_contracts
    WHERE id IN (SELECT MAX(id) FROM tenant_contracts WHERE real_estate_id IS NOT NULL GROUP BY real_estate_id)
    """)
    re_total_deposit = float(cur.fetchone()[0] or 0)
    cur.close()
    cur = db.cursor()
    cur.execute("SELECT COALESCE(SUM(deposit), 0) FROM residence")
    residence_deposit = float(cur.fetchone()[0] or 0)
    cur.close()
    re_val = re_total_price - re_total_deposit + residence_deposit

    # 연금 누적액
    cur = db.cursor()
    cur.execute(
    "SELECT COALESCE(SUM(accumulated),0) as val FROM pension"
    )
    pension_val = float(cur.fetchone()['val'] or 0)
    cur.close()

    # 현금/예금
    cur = db.cursor()
    cur.execute(
    "SELECT COALESCE(SUM(amount),0) as val FROM cash_deposits"
    )
    cash_val = float(cur.fetchone()['val'] or 0)
    cur.close()

    # 부동산 거래 단계 조정
    sell_received = 0.0
    buy_paid = 0.0
    try:
        cur = db.cursor()
        cur.execute("""
            SELECT direction, COALESCE(SUM(amount),0) as total
            FROM real_estate_payments
            WHERE actual_date IS NOT NULL AND actual_date <= CURRENT_DATE
            GROUP BY direction
        """)
        for row in cur.fetchall():
            if row['direction'] == 'sell':
                sell_received = float(row['total'])
            elif row['direction'] == 'buy':
                buy_paid = float(row['total'])
        cur.close()
    except Exception:
        db.rollback()  # 테이블 없을 때 트랜잭션 중단 상태 초기화
    re_val += buy_paid - sell_received

    # 대출 잔액
    cur = db.cursor()
    cur.execute(
    "SELECT COALESCE(SUM(remaining),0) as total FROM loans"
    )
    loan_total = float(cur.fetchone()['total'] or 0)
    cur.close()

    total_assets = stocks_val + etf_val + crypto_val + re_val + pension_val + cash_val
    net_worth = total_assets - loan_total
    gross_assets = total_assets + re_total_deposit

    # 이번달 수입 카테고리별
    cur = db.cursor()
    cur.execute(
    "SELECT category, SUM(amount) as total FROM income WHERE to_char(date::date, 'YYYY-MM') = %s GROUP BY category",
    (ym,)
    )
    income_by_cat = cur.fetchall()
    cur.close()

    # 이번달 지출 카테고리별
    cur = db.cursor()
    cur.execute(
    "SELECT category, SUM(amount) as total FROM budget WHERE to_char(date::date, 'YYYY-MM') = %s GROUP BY category",
    (ym,)
    )
    expense_by_cat = cur.fetchall()
    cur.close()

    # 대출 목록
    cur = db.cursor()
    cur.execute(
    "SELECT name, remaining FROM loans ORDER BY remaining DESC"
    )
    loans_list = cur.fetchall()
    cur.close()

    # 목표저축 목록 (자본주의테크트리 항목 제외)
    cur = db.cursor()
    cur.execute(
    "SELECT name, target_amount, current_amount FROM goals WHERE name != '자본주의테크트리' ORDER BY target_date"
    )
    raw_goals = cur.fetchall()
    cur.close()

    goals_list = [{'name': g['name'], 'target_amount': g['target_amount'], 'current_amount': g['current_amount']} for g in raw_goals]

    # 자본주의테크트리 목표 자산 (app_settings에서 읽기)
    cur = db.cursor()
    cur.execute("SELECT value FROM app_settings WHERE key = 'techTreeTarget'")
    tt_goal = cur.fetchone()
    cur.close()
    if tt_goal and tt_goal['value']:
        goals_list.append({
            'name': '자본주의테크트리 (목표 자산)',
            'target_amount': int(tt_goal['value']),
            'current_amount': int(net_worth),
        })

    # 투자 수익률 (코인 비용만 별도 조회)
    cur = db.cursor()
    cur.execute("SELECT COALESCE(SUM(buy_price * quantity),0) as c FROM crypto")
    crypto_cost = cur.fetchone()['c']
    cur.close()

    # 부동산 매입원가 vs 현재 시세
    cur = db.cursor()
    cur.execute("SELECT COALESCE(SUM(purchase_price),0) as c, COALESCE(SUM(current_price),0) as v FROM real_estate")
    re_inv_row = cur.fetchone()
    re_cost_inv   = float(re_inv_row['c'] or 0)
    re_price_now  = float(re_inv_row['v'] or 0)
    cur.close()

    # 연금 원금(역산) vs 누적액
    cur = db.cursor()
    cur.execute("SELECT COALESCE(SUM(accumulated),0) as val, COALESCE(SUM(accumulated * return_rate / 100.0),0) as profit FROM pension")
    p_row = cur.fetchone()
    pension_total_v    = float(p_row['val']    or 0)
    pension_profit_sum = float(p_row['profit'] or 0)
    pension_cost_inv   = pension_total_v - pension_profit_sum
    cur.close()

    try:
        _save_daily_snapshot(db)
        cur = db.cursor()
        cur.execute("""
        INSERT INTO asset_snapshots (month, cash, stocks, real_estate, crypto, pension, total, updated_at)
        VALUES (%s, %s, %s, %s, %s, %s, %s, CURRENT_TIMESTAMP)
        ON CONFLICT(month) DO UPDATE SET
        cash=excluded.cash, stocks=excluded.stocks, real_estate=excluded.real_estate,
        crypto=excluded.crypto, pension=excluded.pension, total=excluded.total,
        updated_at=excluded.updated_at
        """, (ym, cash_val, stocks_val + etf_val, re_val, crypto_val, pension_val,
        cash_val + stocks_val + etf_val + re_val + crypto_val + pension_val))
        cur.close()
        db.commit()
    except Exception as snapshot_err:
        db.rollback()
        print(f"Error saving daily snapshot in dashboard: {snapshot_err}")

    db.close()

    return jsonify({
        'income_total':    income_total,
        'expense_total':   expense_total + card_total,
        'net_worth':       net_worth,
        'total_assets':    gross_assets,
        'loan_total':      loan_total,
        'asset_breakdown': {
            'stocks_and_etf': stocks_val + etf_val,  # 주식+ETF 합산 (테크트리·투자관리와 동일 기준)
            'stocks':  stocks_val,
            'etf':     etf_val,
            'crypto':  crypto_val,
            'realestate': re_val,
            'pension': pension_val,
            'cash':    cash_val,
        },
        'income_by_cat':  rows_to_list(income_by_cat),
        'expense_by_cat': rows_to_list(expense_by_cat),
        'loans':          rows_to_list(loans_list),
        'goals':          goals_list,
        'investment_returns': {
            'stocks':      {'cost': stocks_cost,   'value': stocks_val},
            'etf':         {'cost': etf_cost,       'value': etf_val},
            'crypto':      {'cost': crypto_cost,    'value': crypto_val},
            'real_estate': {'cost': re_cost_inv,    'value': re_price_now},
            'pension':     {'cost': pension_cost_inv, 'value': pension_total_v},
        },
        'payment_adjustments': {
            'sell_received': sell_received,
            'buy_paid': buy_paid,
            'has_active': sell_received > 0 or buy_paid > 0,
        },
    })


def _calc_annual_avg_income(db):
    """
    실제 기록이 있는 월의 평균을 기준으로 연환산 수입을 계산한다.
    
    반환값:
      labor_monthly_avg   : 월평균 근로소득 (급여+사업소득)
      passive_monthly_avg : 월평균 자생소득 (나머지 수입)
      labor_annual        : 연환산 근로소득 (× 12)
      passive_annual      : 연환산 자생소득 (× 12)
      labor_months        : 집계에 사용된 월 수 (신뢰도 지표)
      passive_months      : 집계에 사용된 월 수 (신뢰도 지표)
    """
    # ── 근로소득: 월별 합계를 구하고, 합계 > 0인 월만 평균에 포함 ──
    cur = db.cursor()
    cur.execute("""
        SELECT 
            to_char(date::date, 'YYYY-MM') AS ym,
            SUM(amount) AS monthly_total
        FROM income
        WHERE category IN ('급여', '사업소득')
          AND date <= CURRENT_DATE           -- 미래 사전등록 수입 제외
        GROUP BY ym
        HAVING SUM(amount) > 0              -- 0인 달은 제외 (데이터 없는 달과 동일 취급)
        ORDER BY ym DESC
        LIMIT 12                            -- 최근 12개월 이내만 사용 (이직/폐업 등 과거 왜곡 방지)
    """)
    labor_rows = cur.fetchall()
    cur.close()

    if labor_rows:
        labor_monthly_avg = sum(r['monthly_total'] for r in labor_rows) / len(labor_rows)
    else:
        labor_monthly_avg = 0

    # ── 자생소득: 동일 방식 ──
    cur = db.cursor()
    cur.execute("""
        SELECT 
            to_char(date::date, 'YYYY-MM') AS ym,
            SUM(amount) AS monthly_total
        FROM income
        WHERE category NOT IN ('급여', '사업소득')
          AND date <= CURRENT_DATE
        GROUP BY ym
        HAVING SUM(amount) > 0
        ORDER BY ym DESC
        LIMIT 12
    """)
    passive_rows = cur.fetchall()
    cur.close()

    # 부동산 임대 수입은 매달 고정으로 발생하므로 그대로 합산
    cur = db.cursor()
    cur.execute("""
        SELECT COALESCE(SUM(monthly_rent), 0)
        FROM tenant_contracts
        WHERE contract_type = '월세'
          AND id IN (SELECT MAX(id) FROM tenant_contracts WHERE real_estate_id IS NOT NULL GROUP BY real_estate_id)
    """)
    rental_monthly = cur.fetchone()[0]
    cur.close()

    cur = db.cursor()
    cur.execute("""
        SELECT COALESCE(SUM(deposit * 0.04 / 12), 0)
        FROM tenant_contracts
        WHERE contract_type = '전세'
          AND id IN (SELECT MAX(id) FROM tenant_contracts WHERE real_estate_id IS NOT NULL GROUP BY real_estate_id)
    """)
    leverage_monthly = cur.fetchone()[0]
    cur.close()

    if passive_rows:
        passive_monthly_avg = sum(r['monthly_total'] for r in passive_rows) / len(passive_rows)
    else:
        passive_monthly_avg = 0

    passive_monthly_avg += (rental_monthly + leverage_monthly)

    return {
        'labor_monthly_avg':   round(labor_monthly_avg),
        'passive_monthly_avg': round(passive_monthly_avg),
        'labor_annual':        round(labor_monthly_avg * 12),
        'passive_annual':      round(passive_monthly_avg * 12),
        'labor_months':        len(labor_rows),
        'passive_months':      len(passive_rows),
    }


@app.route('/api/tech-tree-data')
@cache.cached(timeout=180)
def api_tech_tree_data():
    try:
        return _retry_on_db_error(_api_tech_tree_data_inner)()
    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'trace': traceback.format_exc()}), 500

def _api_tech_tree_data_inner():
    db = get_db()
    ex_rate = get_current_exchange_rate()
    stocks_val, _, etf_val, _ = get_stocks_etf_totals(db, ex_rate)

    today = date.today()
    ym = today.strftime('%Y-%m')
    
    # 인덱스 친화적인 날짜 범위 계산
    ym_start = today.strftime('%Y-%m-01')
    if today.month == 12:
        ym_end = date(today.year + 1, 1, 1).strftime('%Y-%m-01')
    else:
        ym_end = date(today.year, today.month + 1, 1).strftime('%Y-%m-01')

    # 단일 CTE 쿼리로 여러 집계 데이터 일괄 조회
    cur = db.cursor()
    cur.execute("""
    WITH
        crypto_val AS (SELECT COALESCE(SUM(current_price * quantity), 0) AS val FROM crypto),
        re_total_price AS (SELECT COALESCE(SUM(current_price), 0) AS val FROM real_estate),
        re_total_deposit AS (
            SELECT COALESCE(SUM(deposit), 0) AS val FROM tenant_contracts
            WHERE id IN (SELECT MAX(id) FROM tenant_contracts WHERE real_estate_id IS NOT NULL GROUP BY real_estate_id)
        ),
        residence_deposit AS (SELECT COALESCE(SUM(deposit), 0) AS val FROM residence),
        cash_val AS (SELECT COALESCE(SUM(amount), 0) AS val FROM cash_deposits),
        pension_val AS (SELECT COALESCE(SUM(accumulated), 0) AS val FROM pension),
        
        labor_inc AS (
            SELECT COALESCE(SUM(amount), 0) AS val FROM income
            WHERE date >= %s AND date < %s AND category IN ('급여', '사업소득') AND date <= CURRENT_DATE
        ),
        passive_inc AS (
            SELECT COALESCE(SUM(amount), 0) AS val FROM income
            WHERE date >= %s AND date < %s AND category NOT IN ('급여', '사업소득') AND date <= CURRENT_DATE
        ),
        
        rental_inc AS (
            SELECT COALESCE(SUM(monthly_rent), 0) AS val FROM tenant_contracts
            WHERE contract_type = '월세'
              AND id IN (SELECT MAX(id) FROM tenant_contracts WHERE real_estate_id IS NOT NULL GROUP BY real_estate_id)
        ),
        leverage_inc AS (
            SELECT COALESCE(SUM(deposit * 0.04 / 12), 0) AS val FROM tenant_contracts
            WHERE contract_type = '전세'
              AND id IN (SELECT MAX(id) FROM tenant_contracts WHERE real_estate_id IS NOT NULL GROUP BY real_estate_id)
        ),
        expense_total AS (
            SELECT COALESCE(SUM(amount), 0) AS val FROM budget WHERE date >= %s AND date < %s
        ),
        card_total AS (
            SELECT COALESCE(SUM(amount), 0) AS val FROM card_tx WHERE date >= %s AND date < %s AND budget_id IS NULL
        ),
        loan_repayment AS (SELECT COALESCE(SUM(monthly_payment), 0) AS val FROM loans),
        loan_total AS (SELECT COALESCE(SUM(remaining), 0) AS val FROM loans),
        
        s_buy AS (
            SELECT COALESCE(SUM(price*quantity), 0) AS val FROM stock_tx 
            WHERE tx_date >= %s AND tx_date < %s AND tx_type IN ('buy','매수')
        ),
        s_sell AS (
            SELECT COALESCE(SUM(price*quantity), 0) AS val FROM stock_tx
            WHERE tx_date >= %s AND tx_date < %s AND tx_type IN ('sell','매도')
        ),
        c_buy AS (
            SELECT COALESCE(SUM(buy_price*quantity), 0) AS val FROM crypto 
            WHERE buy_date >= %s AND buy_date < %s
        ),
        target_amount AS (SELECT value AS val FROM app_settings WHERE key = 'techTreeTarget')

    SELECT
        (SELECT val FROM crypto_val) AS crypto_val,
        (SELECT val FROM re_total_price) AS re_total_price,
        (SELECT val FROM re_total_deposit) AS re_total_deposit,
        (SELECT val FROM residence_deposit) AS residence_deposit,
        (SELECT val FROM cash_val) AS cash_val,
        (SELECT val FROM pension_val) AS pension_val,
        (SELECT val FROM labor_inc) AS labor_inc,
        (SELECT val FROM passive_inc) AS passive_inc,
        (SELECT val FROM rental_inc) AS rental_inc,
        (SELECT val FROM leverage_inc) AS leverage_inc,
        (SELECT val FROM expense_total) AS expense_total,
        (SELECT val FROM card_total) AS card_total,
        (SELECT val FROM loan_repayment) AS loan_repayment,
        (SELECT val FROM loan_total) AS loan_total,
        (SELECT val FROM s_buy) AS s_buy,
        (SELECT val FROM s_sell) AS s_sell,
        (SELECT val FROM c_buy) AS c_buy,
        (SELECT val FROM target_amount) AS target_amount
    """, (
        ym_start, ym_end,  # labor_inc
        ym_start, ym_end,  # passive_inc
        ym_start, ym_end,  # expense_total
        ym_start, ym_end,  # card_total
        ym_start, ym_end,  # s_buy
        ym_start, ym_end,  # s_sell
        ym_start, ym_end   # c_buy
    ))
    row = cur.fetchone()
    cur.close()

    crypto_val = float(row['crypto_val'] or 0)
    re_total_price = get_real_estate_value(db)
    re_total_deposit = float(row['re_total_deposit'] or 0)
    residence_deposit = float(row['residence_deposit'] or 0)
    re_val = re_total_price - re_total_deposit + residence_deposit

    cash_val = float(row['cash_val'] or 0)
    pension_val = float(row['pension_val'] or 0)
    labor_inc = float(row['labor_inc'] or 0)
    passive_inc = float(row['passive_inc'] or 0)
    rental_inc = float(row['rental_inc'] or 0)
    leverage_inc = float(row['leverage_inc'] or 0)
    
    passive_inc += (rental_inc + leverage_inc)

    # 당월 주식/ETF 실현손익 (FIFO 기준, 미실현 평가손익 제외)
    from collections import defaultdict as _defaultdict
    cur = db.cursor()
    cur.execute("""
        SELECT 'stock' as source, t.stock_id as asset_id, t.tx_date::text as tx_date,
               t.tx_type, t.price, t.quantity, COALESCE(t.fee,0) as fee, s.ticker
        FROM stock_tx t JOIN stocks s ON s.id = t.stock_id
        ORDER BY t.stock_id, t.tx_date, t.id
    """)
    all_stock_tx = cur.fetchall()
    cur.execute("""
        SELECT 'etf' as source, t.etf_id as asset_id, t.tx_date::text as tx_date,
               t.tx_type, t.price, t.quantity, COALESCE(t.fee,0) as fee, e.ticker
        FROM etf_tx t JOIN etf e ON e.id = t.etf_id
        ORDER BY t.etf_id, t.tx_date, t.id
    """)
    all_etf_tx = cur.fetchall()
    cur.close()

    tx_by_asset_tt = _defaultdict(list)
    for tx in list(all_stock_tx) + list(all_etf_tx):
        tx_by_asset_tt[(tx['source'], tx['asset_id'])].append(tx)

    stock_realized_pnl = 0.0
    for (source, asset_id), txs in tx_by_asset_tt.items():
        qty = 0.0; avg_cost = 0.0
        for tx in txs:
            tq = float(tx['quantity']); tp = float(tx['price'])
            if tx['tx_type'] in ('buy', '매수'):
                new_qty = qty + tq
                avg_cost = (qty * avg_cost + tq * tp) / new_qty if new_qty > 0 else 0.0
                qty = new_qty
            else:
                tx_ym = (tx['tx_date'] or '')[:7]
                if tx_ym == ym:
                    pnl = (tp - avg_cost) * tq - float(tx['fee'] or 0)
                    mul = ex_rate if is_foreign_ticker(tx['ticker']) else 1.0
                    stock_realized_pnl += pnl * mul
                qty = max(0.0, qty - tq)
                if qty == 0.0:
                    avg_cost = 0.0

    passive_inc += stock_realized_pnl

    # 당월 공모주 실현손익
    cur = db.cursor()
    cur.execute("""
        SELECT COALESCE(SUM(realized_pnl - COALESCE(fee, 0)), 0) AS val
        FROM ipo
        WHERE SUBSTRING(listing_date, 1, 7) = %s
    """, (ym,))
    ipo_pnl = float(cur.fetchone()[0] or 0)
    cur.close()
    passive_inc += ipo_pnl

    # 당월 코인 매도 실현손익
    try:
        cur = db.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS crypto_sell (
                id SERIAL PRIMARY KEY, sell_date DATE NOT NULL,
                name TEXT NOT NULL, pnl INTEGER DEFAULT 0,
                memo TEXT, created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        db.commit()
        cur.execute("""
            SELECT COALESCE(SUM(pnl), 0) AS val FROM crypto_sell
            WHERE TO_CHAR(sell_date, 'YYYY-MM') = %s
        """, (ym,))
        crypto_sell_pnl = float(cur.fetchone()[0] or 0)
        cur.close()
    except Exception:
        crypto_sell_pnl = 0.0
    passive_inc += crypto_sell_pnl

    # 고정비(빨대) 합계 계산 - /api/straws 와 동일한 기준 (3개월 연속 + 편차 ±10%)
    import math as _math
    cur = db.cursor()
    cur.execute("""
        SELECT name,
               to_char(date::date, 'YYYY-MM') AS month,
               AVG(amount) AS avg_amount
        FROM budget
        WHERE date >= (CURRENT_DATE - INTERVAL '4 months')
          AND amount > 0
          AND name IS NOT NULL AND name != ''
        GROUP BY name, to_char(date::date, 'YYYY-MM')
    """)
    _straw_rows = cur.fetchall()
    cur.close()

    _name_months = {}
    for r in _straw_rows:
        _name_months.setdefault(r['name'], []).append({'month': r['month'], 'avg': float(r['avg_amount'] or 0)})

    _recent_3 = set()
    _y, _m = today.year, today.month
    for _ in range(3):
        _recent_3.add(f"{_y}-{_m:02d}")
        _m -= 1
        if _m == 0: _m = 12; _y -= 1

    straw_total = 0
    _straw_list = []
    for name, entries in _name_months.items():
        months_present = {e['month'] for e in entries}
        if not _recent_3.issubset(months_present):
            continue
        amounts = [e['avg'] for e in entries if e['month'] in _recent_3]
        if not amounts:
            continue
        mean_amt = sum(amounts) / len(amounts)
        if mean_amt <= 0:
            continue
        variance = sum((a - mean_amt) ** 2 for a in amounts) / len(amounts)
        std_dev  = _math.sqrt(variance)
        if mean_amt > 0 and (std_dev / mean_amt) > 0.10:
            continue
        straw_total += round(mean_amt)
        _straw_list.append({'name': name, 'amount': round(mean_amt)})

    expense_total = float(row['expense_total'] or 0)
    card_total = float(row['card_total'] or 0)
    loan_repayment = float(row['loan_repayment'] or 0)
    total_exp = expense_total + card_total + loan_repayment

    # [신규] 월간 변동성 계산 (이번달 순유입액 기준)
    monthly_stats = {
        'cash': {'change': (labor_inc + passive_inc) - total_exp, 'percent': 0},
        'stocks': {'change': 0, 'percent': 0},
        'real_estate': {'change': 0, 'percent': 0},
        'crypto': {'change': 0, 'percent': 0}
    }
    
    s_buy = float(row['s_buy'] or 0)
    s_sell = float(row['s_sell'] or 0)
    monthly_stats['stocks']['change'] = s_buy - s_sell
    monthly_stats['crypto']['change'] = float(row['c_buy'] or 0)

    # 변동률 계산 (현재값 대비)
    def calc_pct(val, change):
        val, change = float(val), float(change)
        prev = val - change
        return round((change / prev * 100), 1) if prev > 0 else 0
    
    monthly_stats['cash']['percent'] = calc_pct(cash_val, monthly_stats['cash']['change'])
    monthly_stats['stocks']['percent'] = calc_pct(stocks_val + etf_val, monthly_stats['stocks']['change'])
    monthly_stats['crypto']['percent'] = calc_pct(crypto_val, monthly_stats['crypto']['change'])

    # 목표 자산
    target_amount = int(row['target_amount']) if row['target_amount'] else 1000000000  # 기본 10억

    # [신규] 월별 자산 스냅샷 자동 저장/업데이트
    cur = db.cursor()
    cur.execute("""
    INSERT INTO asset_snapshots (month, cash, stocks, real_estate, crypto, pension, total, updated_at)
    VALUES (%s, %s, %s, %s, %s, %s, %s, CURRENT_TIMESTAMP)
    ON CONFLICT(month) DO UPDATE SET
    cash=excluded.cash, stocks=excluded.stocks, real_estate=excluded.real_estate,
    crypto=excluded.crypto, pension=excluded.pension, total=excluded.total,
    updated_at=excluded.updated_at
    """, (ym, cash_val, stocks_val + etf_val, re_val, crypto_val, pension_val, 
    cash_val + stocks_val + etf_val + re_val + crypto_val + pension_val))
    cur.close()
    
    try:
        _save_daily_snapshot(db)
    except Exception as snapshot_err:
        print(f"Error saving daily snapshot in tech tree: {snapshot_err}")
    db.commit()

    avg_income = _calc_annual_avg_income(db)

    db.close()
    return jsonify({
        'assets': {
            'cash': int(cash_val or 0),
            'stocks': int(stocks_val or 0) + int(etf_val or 0),
            'real_estate': int(re_val or 0),
            'crypto': int(crypto_val or 0),
            'pension': int(pension_val or 0),
            'loan_total': int(float(row['loan_total'] or 0)),
        },
        'income': {
            'labor': int(labor_inc or 0),
            'passive': int(passive_inc or 0),
            'stock_pnl': int(stock_realized_pnl or 0),
            'labor_annual_avg': avg_income['labor_annual'],
            'passive_annual_avg': avg_income['passive_annual'],
            'labor_months': avg_income['labor_months'],
            'passive_months': avg_income['passive_months']
        },
        'expense': int(total_exp or 0),
        'straw_total': int(straw_total or 0),
        'target_amount': int(target_amount or 0),
        'monthly_stats': {
            k: {
                'change': int(v['change'] or 0),
                'percent': v['percent']
            } for k, v in monthly_stats.items()
        }
    })

@app.route('/api/straws')
def api_straws():
    """최근 3개월 이상 연속 발생 + 금액 편차 ±10% 이내인 고정비(빨대) 목록 반환"""
    db = get_db()
    today_d = date.today()
    ym_now  = today_d.strftime('%Y-%m')
    cur = db.cursor()

    # 최근 4개월치 데이터 (3개월 연속 여부 확인을 위해 1개 여유)
    cur.execute("""
        SELECT name, category,
               to_char(date::date, 'YYYY-MM') AS month,
               AVG(amount) AS avg_amount
        FROM budget
        WHERE date >= (CURRENT_DATE - INTERVAL '4 months')
          AND amount > 0
          AND name IS NOT NULL AND name != ''
        GROUP BY name, category, to_char(date::date, 'YYYY-MM')
    """)
    rows = cur.fetchall()

    # 이번달 수입 (수입 대비 비율 계산용)
    cur.execute(
        "SELECT COALESCE(SUM(amount),0) FROM income "
        "WHERE to_char(date::date,'YYYY-MM')=%s AND date<=CURRENT_DATE", (ym_now,)
    )
    monthly_income = float(cur.fetchone()[0] or 0)
    cur.close()
    db.close()

    from collections import defaultdict
    import math
    name_months = defaultdict(list)
    name_cat    = {}
    for r in rows:
        name_months[r['name']].append({'month': r['month'], 'avg': float(r['avg_amount'] or 0)})
        if r['category']:
            name_cat[r['name']] = r['category']

    # 최근 3개월 목록 구성
    recent_3 = set()
    y, m = today_d.year, today_d.month
    for _ in range(3):
        recent_3.add(f"{y}-{m:02d}")
        m -= 1
        if m == 0: m = 12; y -= 1

    straws = []
    for name, entries in name_months.items():
        months_present = {e['month'] for e in entries}
        # 3개월 모두 발생했는지 확인
        if not recent_3.issubset(months_present):
            continue
        amounts = [e['avg'] for e in entries if e['month'] in recent_3]
        if not amounts:
            continue
        mean_amt = sum(amounts) / len(amounts)
        if mean_amt <= 0:
            continue
        # 표준편차 / 평균 ≤ 10% (변동계수)
        variance = sum((a - mean_amt) ** 2 for a in amounts) / len(amounts)
        std_dev  = math.sqrt(variance)
        if mean_amt > 0 and (std_dev / mean_amt) > 0.10:
            continue
        straws.append({
            'name':     name,
            'category': name_cat.get(name, ''),
            'amount':   round(mean_amt),
            'cnt':      len(amounts),
            'months':   sorted(months_present)
        })

    straws.sort(key=lambda x: -x['amount'])
    total_straw = sum(s['amount'] for s in straws)
    income_ratio = round(total_straw / monthly_income * 100, 1) if monthly_income > 0 else 0

    return jsonify({
        'straws': straws,
        'total_amount': total_straw,
        'monthly_income': round(monthly_income),
        'income_ratio': income_ratio
    })

@app.route('/api/tech-tree-goal', methods=['POST'])
def api_tech_tree_goal():
    db = get_db()
    data = request.json
    target = str(data.get('target_amount', 1000000000))
    cur = db.cursor()
    cur.execute(
        "INSERT INTO app_settings (key, value) VALUES ('techTreeTarget', %s) "
        "ON CONFLICT (key) DO UPDATE SET value=EXCLUDED.value",
        (target,)
    )
    cur.close()
    db.commit()
    db.close()
    return jsonify({'ok': True})


# ── API: 자산별 연도별 성장 통계 (Flow Rate & Milestone) ─────────
@app.route('/api/tech-tree-yearly-stats')
def api_tech_tree_yearly_stats():
    try:
        return _api_tech_tree_yearly_stats_inner()
    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'trace': traceback.format_exc()}), 500

def _api_tech_tree_yearly_stats_inner():
    db = get_db()
    today = date.today()
    
    # 1. 목표 자산 가져오기
    cur = db.cursor()
    cur.execute("SELECT value FROM app_settings WHERE key = 'techTreeTarget'")
    goal = cur.fetchone()
    cur.close()
    target_amount = int(goal['value']) if goal and goal['value'] else 1000000000
    
    # 2. 실시간 현재 자산 가져오기
    ex_rate = get_current_exchange_rate()
    stocks_val, _, etf_val, _ = get_stocks_etf_totals(db, ex_rate)

    cur = db.cursor()
    cur.execute("SELECT COALESCE(SUM(current_price * quantity),0) FROM crypto")
    crypto_val = float(cur.fetchone()[0] or 0)
    cur.close()

    re_total_price = get_real_estate_value(db)

    cur = db.cursor()
    cur.execute("""
    SELECT COALESCE(SUM(deposit), 0) FROM tenant_contracts
    WHERE id IN (SELECT MAX(id) FROM tenant_contracts WHERE real_estate_id IS NOT NULL GROUP BY real_estate_id)
    """)
    re_total_deposit = float(cur.fetchone()[0] or 0)
    cur.close()

    cur = db.cursor()
    cur.execute("SELECT COALESCE(SUM(deposit), 0) FROM residence")
    residence_deposit = float(cur.fetchone()[0] or 0)
    cur.close()
    re_val = re_total_price - re_total_deposit + residence_deposit

    cur = db.cursor()
    cur.execute("SELECT COALESCE(SUM(amount),0) FROM cash_deposits")
    cash_val = float(cur.fetchone()[0] or 0)
    cur.close()

    cur = db.cursor()
    cur.execute("SELECT COALESCE(SUM(accumulated),0) FROM pension")
    pension_val = float(cur.fetchone()[0] or 0)
    cur.close()

    cur = db.cursor()
    cur.execute("SELECT COALESCE(SUM(remaining),0) FROM loans")
    loan_total = float(cur.fetchone()[0] or 0)
    cur.close()

    current_total = int(cash_val + stocks_val + etf_val + re_val + crypto_val + pension_val - loan_total)
    current_percent = round((current_total / target_amount) * 100, 1) if target_amount > 0 else 0
    
    # 3. 스냅샷 데이터 조회하여 연도별 마지막 스냅샷 추출
    cur = db.cursor()
    cur.execute("SELECT month, cash, stocks, real_estate, crypto, pension, total FROM asset_snapshots ORDER BY month ASC")
    all_snapshots = cur.fetchall()
    cur.close()
    
    yearly_map = {}
    for s in all_snapshots:
        m = s['month']
        year = m[:4]
        yearly_map[year] = {
            'year': year,
            'month': m,
            'total': s['total'],
            'breakdown': {
                'cash': s['cash'],
                'stocks': s['stocks'],
                'real_estate': s['real_estate'],
                'crypto': s['crypto'],
                'pension': s['pension']
            }
        }
    
    # 올해 실시간 데이터 반영
    curr_year_str = str(today.year)
    curr_month_str = today.strftime('%Y-%m')
    
    yearly_map[curr_year_str] = {
        'year': curr_year_str,
        'month': curr_month_str,
        'total': current_total,
        'breakdown': {
            'cash': int(cash_val),
            'stocks': int(stocks_val + etf_val),
            'real_estate': int(re_val),
            'crypto': int(crypto_val),
            'pension': int(pension_val)
        }
    }
    
    # 맵 정렬 및 증감액 연산
    sorted_years = sorted(yearly_map.keys())
    yearly_history = []
    
    for idx, yr in enumerate(sorted_years):
        item = yearly_map[yr]
        total_assets = item['total']
        item['percent'] = round((total_assets / target_amount) * 100, 1) if target_amount > 0 else 0
        
        if idx == 0:
            item['change_amount'] = 0
            item['change_percent'] = 0.0
        else:
            prev_total = yearly_map[sorted_years[idx - 1]]['total']
            item['change_amount'] = total_assets - prev_total
            item['change_percent'] = round((total_assets - prev_total) / prev_total * 100, 1) if prev_total > 0 else 0.0
            
        yearly_history.append(item)
        
    # 4. 연간 유입 속도 (Flow Rate) 연산
    flow_rate_amount = 0
    changes = [h['change_amount'] for h in yearly_history if h['change_amount'] > 0]
    if len(changes) > 0:
        flow_rate_amount = sum(changes) / len(changes)
    else:
        # 과거 데이터가 없는 경우 이번달 소득/지출 기준 추정
        ym = today.strftime('%Y-%m')
        cur = db.cursor()
        cur.execute("SELECT COALESCE(SUM(amount),0) FROM income WHERE to_char(date::date, 'YYYY-MM') = %s AND category IN ('급여', '사업소득') AND date <= CURRENT_DATE", (ym,))
        labor_inc = cur.fetchone()[0] or 0
        cur.close()
        
        cur = db.cursor()
        cur.execute("SELECT COALESCE(SUM(amount),0) FROM income WHERE to_char(date::date, 'YYYY-MM') = %s AND category NOT IN ('급여', '사업소득') AND date <= CURRENT_DATE", (ym,))
        passive_inc = cur.fetchone()[0] or 0
        cur.close()
        
        cur = db.cursor()
        cur.execute("""
        SELECT COALESCE(SUM(monthly_rent), 0) FROM tenant_contracts 
        WHERE contract_type = '월세' AND id IN (SELECT MAX(id) FROM tenant_contracts WHERE real_estate_id IS NOT NULL GROUP BY real_estate_id)
        """)
        rental_inc = cur.fetchone()[0] or 0
        cur.close()
        
        cur = db.cursor()
        cur.execute("""
        SELECT COALESCE(SUM(deposit * 0.04 / 12), 0) FROM tenant_contracts 
        WHERE contract_type = '전세' AND id IN (SELECT MAX(id) FROM tenant_contracts WHERE real_estate_id IS NOT NULL GROUP BY real_estate_id)
        """)
        leverage_inc = cur.fetchone()[0] or 0
        cur.close()
        
        total_income = labor_inc + passive_inc + rental_inc + leverage_inc
        
        cur = db.cursor()
        cur.execute("SELECT COALESCE(SUM(amount),0) FROM budget WHERE to_char(date::date, 'YYYY-MM') = %s", (ym,))
        expense_total = cur.fetchone()[0] or 0
        cur.close()
        
        cur = db.cursor()
        cur.execute("SELECT COALESCE(SUM(amount),0) FROM card_tx WHERE to_char(date::date, 'YYYY-MM') = %s AND budget_id IS NULL", (ym,))
        card_total = cur.fetchone()[0] or 0
        cur.close()
        
        cur = db.cursor()
        cur.execute("SELECT COALESCE(SUM(monthly_payment), 0) FROM loans")
        loan_repayment = cur.fetchone()[0] or 0
        cur.close()
        
        total_expense = expense_total + card_total + loan_repayment
        monthly_net_savings = total_income - total_expense
        flow_rate_amount = max(monthly_net_savings * 12, 0)
        
    if flow_rate_amount == 0:
        flow_rate_amount = 12000000 # 기본 연 1200만 원 (월 100만 원)
        
    flow_rate_percent = round((flow_rate_amount / target_amount) * 100, 1) if target_amount > 0 else 0
    
    # 5. 목표 완충 예상 기간 연산
    remaining_amount = max(target_amount - current_total, 0)
    remaining_years = 0
    remaining_months = 0
    expected_completion_ym = "완충 달성 완료!"
    
    if remaining_amount > 0:
        if flow_rate_amount > 0:
            total_months_needed = (remaining_amount / flow_rate_amount) * 12
            remaining_years = int(total_months_needed // 12)
            remaining_months = int(round(total_months_needed % 12))
            if remaining_months == 12:
                remaining_years += 1
                remaining_months = 0
                
            total_add_months = remaining_years * 12 + remaining_months
            expected_year = today.year
            expected_month = today.month + total_add_months
            while expected_month > 12:
                expected_year += 1
                expected_month -= 12
            expected_completion_ym = f"{expected_year}년 {expected_month}월"
        else:
            expected_completion_ym = "예측 불가 (자산 정체)"
            
    db.close()
    return jsonify({
        'target_amount': target_amount,
        'current_total': current_total,
        'current_percent': current_percent,
        'flow_rate_amount': int(flow_rate_amount),
        'flow_rate_percent': flow_rate_percent,
        'remaining_years': remaining_years,
        'remaining_months': remaining_months,
        'expected_completion_ym': expected_completion_ym,
        'yearly_history': yearly_history
    })


# ── API: 자산별 히스토리 (최근 12개월) ──────────────────────────
@app.route('/api/asset-history')
@cache.cached(timeout=300)
def api_asset_history():
  try:
    db = get_db()
    today = date.today()
    history = []
    
    # 1. DB에 저장된 스냅샷 불러오기
    cur = db.cursor()
    cur.execute("SELECT * FROM asset_snapshots ORDER BY month DESC")
    snapshots = {r['month']: r for r in cur.fetchall()}
    cur.close()

    # 현재 실시간 자산 상태 (역산용 기준점)
    cur = db.cursor()
    cur.execute("SELECT COALESCE(SUM(amount),0) FROM cash_deposits")
    curr_cash = float(cur.fetchone()[0] or 0)
    cur.close()
    cur = db.cursor()
    cur.execute("SELECT COALESCE(SUM(accumulated),0) FROM pension")
    curr_pension = float(cur.fetchone()[0] or 0)
    cur.close()
    cur = db.cursor()
    cur.execute("SELECT COALESCE(SUM(monthly_payment),0) FROM pension")
    p_monthly = float(cur.fetchone()[0] or 0)
    cur.close()
    ex_rate = get_current_exchange_rate()
    _sv, _, _ev, _ = get_stocks_etf_totals(db, ex_rate)
    curr_stocks = _sv + _ev
    cur = db.cursor()
    cur.execute("SELECT COALESCE(SUM(current_price * quantity),0) FROM crypto")
    curr_crypto = float(cur.fetchone()[0] or 0)
    cur.close()
    re_price = get_real_estate_value(db)
    cur = db.cursor()
    cur.execute("SELECT COALESCE(SUM(deposit), 0) FROM tenant_contracts WHERE id IN (SELECT MAX(id) FROM tenant_contracts WHERE real_estate_id IS NOT NULL GROUP BY real_estate_id)")
    re_dep = float(cur.fetchone()[0] or 0)
    cur.close()
    cur = db.cursor()
    cur.execute("SELECT COALESCE(SUM(deposit), 0) FROM residence")
    res_dep = float(cur.fetchone()[0] or 0)
    cur.close()
    curr_re = re_price - re_dep + res_dep

    # 12개월 범위의 연월 리스트 미리 구성
    months_list = []
    y, m = today.year, today.month
    for _ in range(12):
        months_list.append(f"{y}-{m:02d}")
        m -= 1
        if m == 0: m = 12; y -= 1

    # 12개월 범위의 날짜 경계 계산
    earliest_ym = months_list[-1]
    latest_ym = months_list[0]
    start_date = f"{earliest_ym}-01"
    ly, lm = int(latest_ym[:4]), int(latest_ym[5:])
    if lm == 12:
        end_date = f"{ly+1:04d}-01-01"
    else:
        end_date = f"{ly:04d}-{lm+1:02d}-01"

    # 최근 12개월간의 수입, 지출, 카드, 주식 거래, 코인 거래 일괄 집계 (루프 밖으로 쿼리 통합)
    cur = db.cursor()
    cur.execute("""
        SELECT to_char(date::date, 'YYYY-MM') as ym, COALESCE(SUM(amount), 0)
        FROM income
        WHERE date >= %s AND date < %s
        GROUP BY ym
    """, (start_date, end_date))
    inc_map = {r[0]: r[1] for r in cur.fetchall()}
    cur.close()

    cur = db.cursor()
    cur.execute("""
        SELECT to_char(date::date, 'YYYY-MM') as ym, COALESCE(SUM(amount), 0)
        FROM budget
        WHERE date >= %s AND date < %s
        GROUP BY ym
    """, (start_date, end_date))
    exp_map = {r[0]: r[1] for r in cur.fetchall()}
    cur.close()

    cur = db.cursor()
    cur.execute("""
        SELECT to_char(date::date, 'YYYY-MM') as ym, COALESCE(SUM(amount), 0)
        FROM card_tx
        WHERE date >= %s AND date < %s AND budget_id IS NULL
        GROUP BY ym
    """, (start_date, end_date))
    card_map = {r[0]: r[1] for r in cur.fetchall()}
    cur.close()

    cur = db.cursor()
    cur.execute("""
        SELECT to_char(tx_date::date, 'YYYY-MM') as ym, 
               COALESCE(SUM(CASE WHEN tx_type IN ('buy','매수') THEN price*quantity ELSE 0 END), 0) as s_buy,
               COALESCE(SUM(CASE WHEN tx_type IN ('sell','매도') THEN price*quantity ELSE 0 END), 0) as s_sell
        FROM stock_tx
        WHERE tx_date >= %s AND tx_date < %s
        GROUP BY ym
    """, (start_date, end_date))
    stock_tx_map = {r[0]: (r[1], r[2]) for r in cur.fetchall()}
    cur.close()

    cur = db.cursor()
    cur.execute("""
        SELECT to_char(buy_date::date, 'YYYY-MM') as ym, COALESCE(SUM(buy_price*quantity), 0)
        FROM crypto
        WHERE buy_date >= %s AND buy_date < %s
        GROUP BY ym
    """, (start_date, end_date))
    crypto_map = {r[0]: r[1] for r in cur.fetchall()}
    cur.close()

    cur = db.cursor()
    cur.execute("""
        SELECT to_char(buy_date::date, 'YYYY-MM') as ym, COALESCE(SUM(buy_price*quantity), 0)
        FROM crypto
        WHERE buy_date <= CURRENT_DATE
        GROUP BY ym
    """)
    crypto_monthly_buy = {r[0]: float(r[1]) for r in cur.fetchall()}
    cur.close()

    total_crypto_buy = sum(crypto_monthly_buy.values())
    crypto_ratio = (curr_crypto / total_crypto_buy) if total_crypto_buy > 0 else 1.0

    # 거꾸로 12개월치 데이터 생성 (메모리 맵 참조 방식으로 대기시간 격감)
    y, m = today.year, today.month
    for i in range(12):
        ym = f"{y}-{m:02d}"
        
        # 스냅샷이 있으면 스냅샷 데이터 사용, 없으면 역산 데이터 사용
        if ym in snapshots:
            s = snapshots[ym]
            snap_cash    = float(s['cash']        or 0)
            snap_stocks  = float(s['stocks']      or 0)
            snap_re      = float(s['real_estate'] or 0)
            snap_crypto  = float(s['crypto']      or 0)
            snap_pension = float(s['pension']     or 0)
            history.append({
                'month': ym,
                'cash': snap_cash,
                'stocks': snap_stocks,
                'real_estate': snap_re,
                'crypto': snap_crypto,
                'pension': snap_pension,
                'is_snapshot': True
            })
            curr_cash, curr_stocks, curr_re, curr_crypto, curr_pension = snap_cash, snap_stocks, snap_re, snap_crypto, snap_pension
        else:
            cum_crypto_buy = sum(v for k, v in crypto_monthly_buy.items() if k <= ym)
            est_crypto = float(cum_crypto_buy * crypto_ratio)
            history.append({
                'month': ym,
                'cash': float(curr_cash or 0),
                'stocks': float(curr_stocks or 0),
                'real_estate': float(curr_re or 0),
                'crypto': est_crypto,
                'pension': float(curr_pension or 0),
                'is_snapshot': False
            })
            curr_crypto = est_crypto

        # 미리 수집된 메모리 해시맵에서 값 읽기 (속도 혁명!)
        inc   = float(inc_map.get(ym, 0)  or 0)
        exp   = float(exp_map.get(ym, 0)  or 0)
        card  = float(card_map.get(ym, 0) or 0)
        s_buy_raw, s_sell_raw = stock_tx_map.get(ym, (0, 0))
        s_buy  = float(s_buy_raw  or 0)
        s_sell = float(s_sell_raw or 0)
        c_buy  = float(crypto_map.get(ym, 0) or 0)

        curr_cash    = float(curr_cash or 0)    - (inc - (exp + card) - (s_buy + c_buy) + s_sell)
        curr_stocks  = float(curr_stocks or 0)  - (s_buy - s_sell)
        curr_pension = float(curr_pension or 0) - p_monthly
        
        m -= 1
        if m == 0: m = 12; y -= 1

    db.close()
    return jsonify(history[::-1])
  except Exception as e:
    import traceback
    print("api_asset_history error:", traceback.format_exc())
    return jsonify({'error': str(e)}), 500

# ── API: 자산별 상세 내역 조회 ──────────────────────────
@app.route('/api/tech-tree-detail')
def api_tech_tree_detail():
    db = get_db()
    ttype = request.args.get('type')
    ym = date.today().strftime('%Y-%m')
    
    res = []
    if ttype == 'labor':
        cur = db.cursor()
        cur.execute("SELECT date, name, amount, category FROM income WHERE to_char(date::date, 'YYYY-MM') = %s AND category IN ('급여', '사업소득', '기타') AND date <= CURRENT_DATE", (ym,))
        rows = cur.fetchall()
        cur.close()
        res = [{'date': r[0], 'name': r[1], 'amount': r[2], 'memo': r[3]} for r in rows]
    elif ttype == 'passive':
        cur = db.cursor()
        cur.execute("SELECT date, name, amount, category FROM income WHERE to_char(date::date, 'YYYY-MM') = %s AND category NOT IN ('급여', '사업소득', '기타') AND date <= CURRENT_DATE", (ym,))
        rows = cur.fetchall()
        cur.close()
        res = [{'date': r[0], 'name': r[1], 'amount': r[2], 'memo': r[3]} for r in rows]
        # 부동산 월세 수입 추가
        cur = db.cursor()
        cur.execute("""
        SELECT r.name, tc.monthly_rent 
        FROM tenant_contracts tc
        JOIN real_estate r ON tc.real_estate_id = r.id
        WHERE tc.contract_type = '월세' 
        AND tc.id IN (SELECT MAX(id) FROM tenant_contracts WHERE real_estate_id IS NOT NULL GROUP BY real_estate_id)
        """)
        rentals = cur.fetchall()
        cur.close()
        for rent in rentals:
            if rent[1] > 0:
                res.append({'date': '임대료', 'name': rent[0], 'amount': rent[1], 'memo': '부동산 월세'})
        
        # 전세 사적 레버리지 추가
        cur = db.cursor()
        cur.execute("""
        SELECT r.name, tc.deposit * 0.04 / 12 as amount
        FROM tenant_contracts tc
        JOIN real_estate r ON tc.real_estate_id = r.id
        WHERE tc.contract_type = '전세' 
        AND tc.id IN (SELECT MAX(id) FROM tenant_contracts WHERE real_estate_id IS NOT NULL GROUP BY real_estate_id)
        """)
        leverages = cur.fetchall()
        cur.close()
        for lev in leverages:
            if lev[1] > 0:
                res.append({'date': '레버리지', 'name': f"{lev[0]} (사적레버리지)", 'amount': int(lev[1]), 'memo': '전세금 기회비용(4%)'})

        # 주식/ETF 실현손익 내역 추가
        from collections import defaultdict as _defaultdict
        cur = db.cursor()
        cur.execute("""
            SELECT 'stock' as source, t.stock_id as asset_id, t.tx_date::text as tx_date,
                   t.tx_type, t.price, t.quantity, COALESCE(t.fee,0) as fee, s.ticker, s.name
            FROM stock_tx t JOIN stocks s ON s.id = t.stock_id
            ORDER BY t.stock_id, t.tx_date, t.id
        """)
        all_stock_tx = cur.fetchall()
        cur.execute("""
            SELECT 'etf' as source, t.etf_id as asset_id, t.tx_date::text as tx_date,
                   t.tx_type, t.price, t.quantity, COALESCE(t.fee,0) as fee, e.ticker, e.name
            FROM etf_tx t JOIN etf e ON e.id = t.etf_id
            ORDER BY t.etf_id, t.tx_date, t.id
        """)
        all_etf_tx = cur.fetchall()
        cur.close()

        ex_rate = get_current_exchange_rate()
        tx_by_asset_tt = _defaultdict(list)
        for tx in list(all_stock_tx) + list(all_etf_tx):
            tx_by_asset_tt[(tx['source'], tx['asset_id'])].append(tx)

        for (source, asset_id), txs in tx_by_asset_tt.items():
            qty = 0.0; avg_cost = 0.0
            for tx in txs:
                tq = float(tx['quantity']); tp = float(tx['price'])
                if tx['tx_type'] in ('buy', '매수'):
                    new_qty = qty + tq
                    avg_cost = (qty * avg_cost + tq * tp) / new_qty if new_qty > 0 else 0.0
                    qty = new_qty
                else:
                    tx_ym = (tx['tx_date'] or '')[:7]
                    if tx_ym == ym:
                        pnl = (tp - avg_cost) * tq - float(tx['fee'] or 0)
                        mul = ex_rate if is_foreign_ticker(tx['ticker']) else 1.0
                        realized_pnl_krw = pnl * mul
                        
                        currency_symbol = '$' if is_foreign_ticker(tx['ticker']) else '₩'
                        res.append({
                            'date': tx['tx_date'],
                            'name': f"{tx['name']} ({'주식' if source == 'stock' else 'ETF'} 실현손익)",
                            'amount': round(realized_pnl_krw),
                            'memo': f"매도 {tq}주 @ {currency_symbol}{tp} / 평단: {currency_symbol}{round(avg_cost, 2)}"
                        })
                    qty = max(0.0, qty - tq)
                    if qty == 0.0:
                        avg_cost = 0.0

        # 날짜 정렬 (날짜 형식은 최신순 정렬, 임대료/레버리지 등 텍스트는 맨 아래로)
        # 공모주 실현손익 추가
        cur = db.cursor()
        cur.execute("""
            SELECT listing_date, name, (realized_pnl - COALESCE(fee,0)) AS pnl, fee
            FROM ipo
            WHERE SUBSTRING(listing_date, 1, 7) = %s AND realized_pnl != 0
        """, (ym,))
        for r in cur.fetchall():
            fee_amt = int(r[3] or 0)
            res.append({
                'date': r[0],
                'name': f"{r[1]} (공모주 익절)",
                'amount': int(r[2] or 0),
                'memo': f"수수료 {fee_amt:,}원 차감" if fee_amt else ''
            })
        cur.close()

        # 코인 매도 실현손익 추가
        try:
            cur = db.cursor()
            cur.execute("""
                SELECT sell_date::text, name, pnl, memo
                FROM crypto_sell
                WHERE TO_CHAR(sell_date, 'YYYY-MM') = %s AND pnl != 0
            """, (ym,))
            for r in cur.fetchall():
                res.append({
                    'date': r[0],
                    'name': f"{r[1]} (코인 익절)",
                    'amount': int(r[2] or 0),
                    'memo': r[3] or ''
                })
            cur.close()
        except Exception:
            pass

        def sort_key(x):
            d = str(x['date'] or '')
            if len(d) == 10 and d[4] == '-' and d[7] == '-':
                return (1, d)
            return (0, d)
        res.sort(key=sort_key, reverse=True)
    elif ttype == 'expense':
        cur = db.cursor()
        cur.execute("SELECT date, name, amount, category FROM budget WHERE to_char(date::date, 'YYYY-MM') = %s", (ym,))
        b = cur.fetchall()
        cur.close()
        cur = db.cursor()
        cur.execute("SELECT date, name, amount, category FROM card_tx WHERE to_char(date::date, 'YYYY-MM') = %s AND budget_id IS NULL", (ym,))
        c = cur.fetchall()
        cur.close()
        cur = db.cursor()
        cur.execute("SELECT '매월' as date, name, monthly_payment as amount, institution as category FROM loans")
        l = cur.fetchall()
        cur.close()
        res = [{'date': r[0], 'name': r[1], 'amount': r[2], 'memo': r[3]} for r in b + c + l]
        # 날짜 정렬 (최신순, '매월' 등 텍스트 날짜는 맨 아래로)
        def sort_key_exp(x):
            d = str(x['date'] or '')
            if len(d) == 10 and d[4] == '-' and d[7] == '-':
                return (1, d)
            return (0, d)
        res.sort(key=sort_key_exp, reverse=True)
    elif ttype == 'cash':
        cur = db.cursor()
        cur.execute("SELECT '현금' as date, name, amount, memo FROM cash_deposits")
        c = cur.fetchall()
        cur.close()
        cur = db.cursor()
        cur.execute("SELECT '목표' as date, name, current_amount as amount, memo FROM goals WHERE name != '자본주의테크트리'")
        g = cur.fetchall()
        cur.close()
        res = [{'date': r[0], 'name': r[1], 'amount': r[2], 'memo': r[3]} for r in c + g]
    elif ttype == 'stocks':
        ex_rate = get_current_exchange_rate()
        cur = db.cursor()
        cur.execute("""
        SELECT s.name, s.ticker, s.current_price,
        COALESCE(SUM(CASE WHEN t.tx_type IN ('buy','매수') THEN t.quantity ELSE 0 END), 0) AS buy_qty,
        COALESCE(SUM(CASE WHEN t.tx_type IN ('sell','매도') THEN t.quantity ELSE 0 END), 0) AS sell_qty
        FROM stocks s
        LEFT JOIN stock_tx t ON t.stock_id = s.id
        GROUP BY s.id
        """)
        s_rows = cur.fetchall()
        cur.close()
        for r in s_rows:
            qty = float(r['buy_qty'] - r['sell_qty'])
            if qty > 0:
                eval_amt = round(qty * float(r['current_price'] or 0))
                mul = ex_rate if is_foreign_ticker(r['ticker']) else 1
                res.append({'date': '주식', 'name': r['name'], 'amount': float(eval_amt * mul), 'memo': r['ticker']})

        cur = db.cursor()
        cur.execute("""
        SELECT e.name, e.ticker, e.current_price,
        COALESCE(SUM(CASE WHEN t.tx_type IN ('buy','매수') THEN t.quantity ELSE 0 END), 0) AS buy_qty,
        COALESCE(SUM(CASE WHEN t.tx_type IN ('sell','매도') THEN t.quantity ELSE 0 END), 0) AS sell_qty
        FROM etf e
        LEFT JOIN etf_tx t ON t.etf_id = e.id
        GROUP BY e.id
        """)
        e_rows = cur.fetchall()
        cur.close()
        for r in e_rows:
            qty = float(r['buy_qty'] - r['sell_qty'])
            if qty > 0:
                eval_amt = round(qty * float(r['current_price'] or 0))
                mul = ex_rate if is_foreign_ticker(r['ticker']) else 1
                res.append({'date': 'ETF', 'name': r['name'], 'amount': float(eval_amt * mul), 'memo': r['ticker']})
    elif ttype == 'real_estate':
        # 각 매물별 시세 - 최신 보증금(부채) 계산
        cur = db.cursor()
        cur.execute("""
        SELECT r.name, r.current_price, 
        COALESCE((SELECT deposit FROM tenant_contracts WHERE real_estate_id = r.id ORDER BY id DESC LIMIT 1), 0) as tenant_dep,
        r.memo
        FROM real_estate r
        """)
        rows = cur.fetchall()
        cur.close()
        res = [{'date': '부동산', 'name': r[0], 'amount': r[1] - r[2], 'memo': f"시세:{r[1]:,} / 보증금:-{r[2]:,}"} for r in rows]
        # 거주 보증금 추가 (내가 낸 돈이므로 자산)
        cur = db.cursor()
        cur.execute("SELECT address, deposit FROM residence")
        res_dep = cur.fetchall()
        cur.close()
        for rd in res_dep:
            res.append({'date': '거주보증금', 'name': rd[0], 'amount': rd[1], 'memo': '내가 낸 보증금'})
    elif ttype == 'crypto':
        cur = db.cursor()
        cur.execute("SELECT '코인' as date, name, quantity * current_price as amount, symbol as memo FROM crypto")
        rows = cur.fetchall()
        cur.close()
        res = [{'date': r[0], 'name': r[1], 'amount': r[2], 'memo': r[3]} for r in rows]
    elif ttype == 'pension':
        cur = db.cursor()
        cur.execute("SELECT pension_type as date, name, accumulated as amount, institution as memo FROM pension")
        rows = cur.fetchall()
        cur.close()
        res = [{'date': r[0], 'name': r[1], 'amount': r[2], 'memo': r[3]} for r in rows]

    db.close()
    return jsonify(res)


# ── API: 월별 결산 ───────────────────────────────────────────
@app.route('/api/monthly-summary')
def api_monthly_summary():
    db = get_db()
    year = request.args.get('year', date.today().strftime('%Y'))

    months = []
    for m in range(1, 13):
        ym = f"{year}-{m:02d}"
        # 미래 날짜의 수입(반복 수입 사전 등록분)은 해당 날짜가 돼야만 집계
        cur = db.cursor()
        cur.execute(
        "SELECT COALESCE(SUM(amount),0) as t FROM income"
        " WHERE to_char(date::date, 'YYYY-MM') = %s AND date <= CURRENT_DATE", (ym,)
        )
        inc = cur.fetchone()['t']
        cur.close()
        cur = db.cursor()
        cur.execute(
        "SELECT COALESCE(SUM(amount),0) as t FROM budget WHERE to_char(date::date, 'YYYY-MM') = %s", (ym,)
        )
        exp = cur.fetchone()['t']
        cur.close()
        cur = db.cursor()
        cur.execute(
        "SELECT COALESCE(SUM(amount),0) as t FROM card_tx WHERE to_char(date::date, 'YYYY-MM') = %s AND budget_id IS NULL", (ym,)
        )
        card = cur.fetchone()['t']
        cur.close()
        months.append({
            'month':   ym,
            'income':  inc,
            'expense': exp + card,
            'saving':  inc - (exp + card),
        })

    db.close()
    return jsonify(months)


# ── 카드 엑셀 가져오기 ───────────────────────────────────────
_HINTS = {
    'date':        ['이용일', '이용일시', '거래일', '거래일자', '이용일자', '날짜', '결제일', '승인일', '사용일', '거래 일시'],
    'name':        ['이용가맹점', '가맹점명', '상호명', '가맹점', '이용처', '거래처', '내용', '적요', '사용처', '거래내용'],
    'amount':      ['이용금액', '결제금액', '거래금액', '금액', '승인금액', '사용금액', '국내이용금액'],
    'installment': ['할부', '할부개월', '할부개월수', '분할', '할부기간'],
    'category':    ['업종', '카테고리', '이용구분', '업종명', '분류'],
}

def _detect_header(rows):
    all_hints = [h for hs in _HINTS.values() for h in hs]
    best, idx = 0, 0
    for i, row in enumerate(rows[:15]):
        score = sum(1 for c in row if any(h in str(c) for h in all_hints))
        if score > best:
            best, idx = score, i
    return idx

def _detect_mapping(headers):
    m = {}
    for field, hints in _HINTS.items():
        for h in headers:
            if any(hint in str(h) for hint in hints):
                m[field] = h; break
    return m

def _parse_date(val):
    if val is None: return None
    if hasattr(val, 'strftime'): return val.strftime('%Y-%m-%d')
    s = str(val).strip().split(' ')[0].split('T')[0]
    m = re.match(r'(\d{4})[-./](\d{1,2})[-./](\d{1,2})', s)
    if m: return f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
    m = re.match(r'^(\d{4})(\d{2})(\d{2})$', s)
    if m: return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    return None

def _parse_amount(val):
    if val is None: return 0
    if isinstance(val, (int, float)): return int(abs(val))
    try: return int(abs(float(str(val).replace(',', '').replace(' ', ''))))
    except: return 0

def _parse_file(file_bytes, filename):
    rows = []
    name = filename.lower()
    if name.endswith('.csv'):
        for enc in ['utf-8-sig', 'cp949', 'euc-kr', 'utf-8']:
            try:
                text = file_bytes.decode(enc)
                rows = [[str(c).strip() for c in r]
                        for r in csv.reader(io.StringIO(text)) if any(c for c in r)]
                break
            except: continue
    elif name.endswith('.xls'):
        if not HAS_XLRD:
            return [], []
        wb = xlrd.open_workbook(file_contents=file_bytes)
        ws = wb.sheet_by_index(0)
        for i in range(ws.nrows):
            row_vals = []
            for j in range(ws.ncols):
                cell = ws.cell(i, j)
                if cell.ctype == xlrd.XL_CELL_DATE:
                    try:
                        t = xlrd.xldate.xldate_as_tuple(cell.value, wb.datemode)
                        row_vals.append(datetime(*t[:6]))
                    except:
                        row_vals.append(cell.value)
                else:
                    row_vals.append(cell.value)
            if any(v is not None and str(v).strip() for v in row_vals):
                rows.append(row_vals)
    else:  # .xlsx
        if not HAS_OPENPYXL:
            return [], []
        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
            ws = wb.active
            rows = [[cell.value for cell in row] for row in ws.iter_rows()
                    if any(c.value is not None for c in row)]
        except Exception:
            return [], []
    if not rows: return [], []
    hi = _detect_header(rows)
    headers = [str(h).strip() if h is not None else f'컬럼{i}' for i, h in enumerate(rows[hi])]
    data    = [list(row) for row in rows[hi+1:]
               if any(v is not None and str(v).strip() for v in row)]
    return headers, data


@app.route('/api/card-excel/preview', methods=['POST'])
def api_card_excel_preview():
    f = request.files.get('file')
    if not f: return jsonify({'error': '파일이 없습니다'}), 400
    headers, data = _parse_file(f.read(), f.filename)
    if not headers: return jsonify({'error': '파일을 읽을 수 없습니다. xlsx 또는 csv를 올려주세요.'}), 400
    sample = [
        {h: (str(row[i]).strip() if i < len(row) and row[i] is not None else '')
         for i, h in enumerate(headers)}
        for row in data[:5]
    ]
    all_rows = [
        [(str(row[i]).strip() if i < len(row) and row[i] is not None else '') if not hasattr(row[i] if i < len(row) else None, 'strftime')
         else (row[i].strftime('%Y-%m-%d') if i < len(row) and row[i] is not None else '')
         for i in range(len(headers))]
        for row in data
    ]
    return jsonify({'headers': headers, 'mapping': _detect_mapping(headers),
                    'sample': sample, 'total': len(data), 'all_rows': all_rows})


@app.route('/api/card-excel/mapping/<int:card_id>', methods=['GET', 'POST'])
def api_card_excel_mapping(card_id):
    db = get_db()
    if request.method == 'GET':
        cur = db.cursor()
        cur.execute("SELECT mapping FROM card_mappings WHERE card_id=%s", (card_id,))
        row = cur.fetchone()
        cur.close()
        db.close()
        return jsonify(json.loads(row['mapping']) if row else {})
    cur = db.cursor()
    cur.execute("INSERT INTO card_mappings (card_id, mapping) VALUES (%s, %s) ON CONFLICT (card_id) DO UPDATE SET mapping = EXCLUDED.mapping",
    (card_id, json.dumps(request.json or {})))
    cur.close()
    db.commit(); db.close()
    return jsonify({'ok': True})


@app.route('/api/card-excel/import', methods=['POST'])
def api_card_excel_import():
    data     = request.json or {}
    card_id  = data.get('card_id')
    mapping  = data.get('mapping', {})
    headers  = data.get('headers', [])
    all_rows = data.get('all_rows', [])
    if not card_id or not mapping.get('date') or not mapping.get('amount'):
        return jsonify({'error': '카드, 날짜, 금액 컬럼은 필수입니다'}), 400

    hi = {h: i for i, h in enumerate(headers)}
    def get(row, col): return row[hi[col]] if col and col in hi and hi[col] < len(row) else ''

    db = get_db()
    inserted = skipped = duplicate = 0
    for row in all_rows:
        date_str = _parse_date(get(row, mapping['date']))
        amount   = _parse_amount(get(row, mapping['amount']))
        name     = str(get(row, mapping.get('name', '')) or '').strip()
        inst     = _parse_amount(get(row, mapping.get('installment', ''))) or 1
        category = str(get(row, mapping.get('category', '')) or '').strip()
        if not date_str or amount <= 0: skipped += 1; continue
        tmp_cur = db.cursor()
        tmp_cur.execute("SELECT id FROM card_tx WHERE card_id=%s AND date=%s AND name=%s AND amount=%s",
        (card_id, date_str, name, amount))
        exists = tmp_cur.fetchone()
        tmp_cur.close()
        if exists:
            duplicate += 1; continue
        # 카테고리가 없으면 힌트 자동 적용
        if not category and name:
            category = _get_category_hint(db, name)
        cur = db.cursor()
        cur.execute("INSERT INTO card_tx (card_id,date,name,category,amount,installment) VALUES (%s,%s,%s,%s,%s,%s)",
        (card_id, date_str, name, category, amount, inst))
        cur.close()
        inserted += 1
    db.commit(); db.close()
    return jsonify({'ok': True, 'inserted': inserted, 'skipped': skipped, 'duplicate': duplicate})


# ── API: 카테고리 자동 힌트 ──────────────────────────────────
def _get_category_hint(db, name):
    """가맹점명으로 카테고리 추천 (히스토리 → 키워드 규칙 순)"""
    name = (name or '').strip()
    if not name:
        return ''
    # 1. 히스토리: 동일 가맹점의 가장 최근 카테고리
    cur = db.cursor()
    cur.execute(
    "SELECT category FROM card_tx WHERE name=%s AND category IS NOT NULL AND category!='' "
    "ORDER BY date DESC LIMIT 1", (name,)
    )
    row = cur.fetchone()
    cur.close()
    if row:
        return row['category']
    # 2. 키워드 규칙 (긴 키워드 우선)
    cur = db.cursor()
    cur.execute(
    "SELECT keyword, category FROM card_category_rules ORDER BY LENGTH(keyword) DESC"
    )
    rules = cur.fetchall()
    cur.close()
    name_lower = name.lower()
    for rule in rules:
        if rule['keyword'].lower() in name_lower:
            return rule['category']
    return ''


@app.route('/api/card-category-hint')
def api_card_category_hint():
    name = request.args.get('name', '').strip()
    if not name:
        return jsonify({'category': ''})
    db = get_db()
    category = _get_category_hint(db, name)
    db.close()
    return jsonify({'category': category})


@app.route('/api/card-category-rules', methods=['GET', 'POST'])
def api_card_category_rules():
    db = get_db()
    if request.method == 'GET':
        cur = db.cursor()
        cur.execute("SELECT * FROM card_category_rules ORDER BY keyword")
        rows = cur.fetchall()
        cur.close()
        db.close()
        return jsonify(rows_to_list(rows))
    d = request.json or {}
    cur = db.cursor()
    cur.execute("INSERT INTO card_category_rules (keyword, category) VALUES (%s, %s)",
    (d.get('keyword', ''), d.get('category', '')))
    cur.close()
    db.commit(); db.close()
    return jsonify({'ok': True})


@app.route('/api/categories', methods=['GET', 'POST'])
def api_categories():
    db = get_db()
    if request.method == 'GET':
        cur = db.cursor()
        cur.execute("SELECT * FROM categories ORDER BY sort_order, name")
        rows = cur.fetchall()
        cur.close()
        db.close()
        return jsonify(rows_to_list(rows))
    d = request.json or {}
    cur = db.cursor()
    cur.execute("SELECT COALESCE(MAX(sort_order), -1) FROM categories")
    max_order = cur.fetchone()[0]
    cur.close()
    cur = db.cursor()
    cur.execute("INSERT INTO categories (name, sort_order) VALUES (%s, %s) ON CONFLICT (name) DO NOTHING",
    (d.get('name', '').strip(), max_order + 1))
    cur.close()
    db.commit(); db.close()
    return jsonify({'ok': True}), 201


@app.route('/api/categories/<int:rid>', methods=['PUT', 'DELETE'])
def api_categories_detail(rid):
    db = get_db()
    if request.method == 'DELETE':
        cur = db.cursor()
        cur.execute("DELETE FROM categories WHERE id=%s", (rid,))
        cur.close()
        db.commit(); db.close()
        return jsonify({'ok': True})
    d = request.json or {}
    cur = db.cursor()
    cur.execute("UPDATE categories SET name=%s WHERE id=%s", (d.get('name', '').strip(), rid))
    cur.close()
    db.commit(); db.close()
    return jsonify({'ok': True})


@app.route('/api/card-category-rules/<int:rid>', methods=['PUT', 'DELETE'])
def api_card_category_rules_detail(rid):
    db = get_db()
    if request.method == 'DELETE':
        cur = db.cursor()
        cur.execute("DELETE FROM card_category_rules WHERE id=%s", (rid,))
        cur.close()
        db.commit(); db.close()
        return jsonify({'ok': True})
    d = request.json or {}
    cur = db.cursor()
    cur.execute("UPDATE card_category_rules SET keyword=%s, category=%s WHERE id=%s",
    (d.get('keyword', ''), d.get('category', ''), rid))
    cur.close()
    db.commit(); db.close()
    return jsonify({'ok': True})


# ── API: 자금 그룹 ────────────────────────────────────────────
def _get_fund_group_hint(db, name):
    """가맹점명으로 자금 그룹 추천 (히스토리 → 키워드 규칙 순)"""
    name = (name or '').strip()
    if not name:
        return None
    # 1. 히스토리: 동일 가맹점의 가장 최근 자금 그룹
    cur = db.cursor()
    cur.execute(
    "SELECT fund_group_id FROM card_tx WHERE name=%s AND fund_group_id IS NOT NULL "
    "ORDER BY date DESC LIMIT 1", (name,)
    )
    row = cur.fetchone()
    cur.close()
    if row:
        return row['fund_group_id']
    # 2. 키워드 규칙 (긴 키워드 우선)
    cur = db.cursor()
    cur.execute(
    "SELECT keyword, fund_group_id FROM fund_group_rules ORDER BY LENGTH(keyword) DESC"
    )
    rules = cur.fetchall()
    cur.close()
    name_lower = name.lower()
    for rule in rules:
        if rule['keyword'].lower() in name_lower:
            return rule['fund_group_id']
    return None


@app.route('/fund-management')
def fund_management():
    return render_template('fund_management.html')


@app.route('/api/fund-groups', methods=['GET', 'POST'])
def api_fund_groups():
    db = get_db()
    if request.method == 'GET':
        cur = db.cursor()
        cur.execute("SELECT * FROM fund_groups ORDER BY sort_order, name")
        rows = cur.fetchall()
        cur.close()
        db.close()
        return jsonify(rows_to_list(rows))
    d = request.json or {}
    cur = db.cursor()
    cur.execute("SELECT COALESCE(MAX(sort_order), -1) FROM fund_groups")
    max_order = cur.fetchone()[0]
    cur.close()
    cur = db.cursor()
    cur.execute("INSERT INTO fund_groups (name, sort_order) VALUES (%s, %s) ON CONFLICT (name) DO NOTHING",
    (d.get('name', '').strip(), max_order + 1))
    cur.close()
    db.commit(); db.close()
    return jsonify({'ok': True}), 201


@app.route('/api/fund-groups/<int:rid>', methods=['PUT', 'DELETE'])
def api_fund_groups_detail(rid):
    db = get_db()
    if request.method == 'DELETE':
        cur = db.cursor()
        cur.execute("DELETE FROM fund_groups WHERE id=%s", (rid,))
        cur.close()
        db.commit(); db.close()
        return jsonify({'ok': True})
    d = request.json or {}
    cur = db.cursor()
    cur.execute("UPDATE fund_groups SET name=%s WHERE id=%s", (d.get('name', '').strip(), rid))
    cur.close()
    db.commit(); db.close()
    return jsonify({'ok': True})


@app.route('/api/fund-group-rules', methods=['GET', 'POST'])
def api_fund_group_rules():
    db = get_db()
    if request.method == 'GET':
        cur = db.cursor()
        cur.execute(
        "SELECT r.*, g.name as fund_group_name FROM fund_group_rules r "
        "LEFT JOIN fund_groups g ON r.fund_group_id = g.id ORDER BY r.keyword"
        )
        rows = cur.fetchall()
        cur.close()
        db.close()
        return jsonify(rows_to_list(rows))
    d = request.json or {}
    cur = db.cursor()
    cur.execute("INSERT INTO fund_group_rules (keyword, fund_group_id) VALUES (%s, %s)",
    (d.get('keyword', ''), d.get('fund_group_id')))
    cur.close()
    db.commit(); db.close()
    return jsonify({'ok': True}), 201


@app.route('/api/fund-group-rules/<int:rid>', methods=['PUT', 'DELETE'])
def api_fund_group_rules_detail(rid):
    db = get_db()
    if request.method == 'DELETE':
        cur = db.cursor()
        cur.execute("DELETE FROM fund_group_rules WHERE id=%s", (rid,))
        cur.close()
        db.commit(); db.close()
        return jsonify({'ok': True})
    d = request.json or {}
    cur = db.cursor()
    cur.execute("UPDATE fund_group_rules SET keyword=%s, fund_group_id=%s WHERE id=%s",
    (d.get('keyword', ''), d.get('fund_group_id'), rid))
    cur.close()
    db.commit(); db.close()
    return jsonify({'ok': True})


@app.route('/api/monthly-fund-budgets', methods=['GET', 'POST'])
def api_monthly_fund_budgets():
    db = get_db()
    if request.method == 'GET':
        year  = request.args.get('year')
        month = request.args.get('month')
        cur = db.cursor()
        cur.execute(
        "SELECT b.*, g.name as fund_group_name FROM monthly_fund_budgets b "
        "LEFT JOIN fund_groups g ON b.fund_group_id = g.id "
        "WHERE b.year=%s AND b.month=%s ORDER BY g.sort_order, g.name",
        (year, int(month))
        )
        rows = cur.fetchall()
        cur.close()
        db.close()
        return jsonify(rows_to_list(rows))
    d = request.json or {}
    cur = db.cursor()
    cur.execute(
    "INSERT INTO monthly_fund_budgets (fund_group_id, year, month, budget_amount) VALUES (%s,%s,%s,%s) "
    "ON CONFLICT(fund_group_id, year, month) DO UPDATE SET budget_amount=excluded.budget_amount",
    (d.get('fund_group_id'), d.get('year'), d.get('month'), d.get('budget_amount', 0))
    )
    cur.close()
    db.commit(); db.close()
    return jsonify({'ok': True}), 201


@app.route('/api/fund-summary')
def api_fund_summary():
    year  = request.args.get('year')
    month = request.args.get('month')
    db = get_db()
    # 자금 그룹별 실제 지출 집계
    cur = db.cursor()
    cur.execute(
    "SELECT g.id, g.name, g.sort_order, COALESCE(SUM(t.amount), 0) as actual "
    "FROM fund_groups g "
    "LEFT JOIN card_tx t ON t.fund_group_id = g.id "
    "  AND to_char(t.date::date, 'YYYY') = %s AND to_char(t.date::date, 'MM') = %s "
    "GROUP BY g.id ORDER BY g.sort_order, g.name",
    (year, month.zfill(2))
    )
    actuals = cur.fetchall()
    cur.close()
    cur = db.cursor()
    cur.execute(
    "SELECT fund_group_id, budget_amount FROM monthly_fund_budgets WHERE year=%s AND month=%s",
    (year, int(month))
    )
    budgets = {r['fund_group_id']: r['budget_amount'] for r in cur.fetchall()}
    cur.close()
    db.close()
    result = []
    for row in actuals:
        result.append({
            'id': row['id'],
            'name': row['name'],
            'actual': row['actual'],
            'budget': budgets.get(row['id'], 0),
        })
    return jsonify(result)


@app.route('/api/card-tx/auto-fund-group', methods=['POST'])
def api_card_tx_auto_fund_group():
    data    = request.json or {}
    card_id = data.get('card_id')
    year    = data.get('year')
    month   = data.get('month')

    db = get_db()
    query  = "SELECT id, name FROM card_tx WHERE fund_group_locked = 0"
    params = []
    if card_id:
        query += " AND card_id = %s"; params.append(card_id)
    if year and month:
        query += " AND to_char(date::date, 'YYYY') = %s AND to_char(date::date, 'MM') = %s"
        params += [year, month.zfill(2)]

    cur = db.cursor()
    cur.execute(query, params)
    rows = cur.fetchall()
    cur.close()
    updated = 0
    for row in rows:
        hint = _get_fund_group_hint(db, row['name'])
        if hint:
            cur = db.cursor()
            cur.execute("UPDATE card_tx SET fund_group_id=%s WHERE id=%s", (hint, row['id']))
            cur.close()
            updated += 1
    db.commit()
    db.close()
    return jsonify({'ok': True, 'updated': updated})


@app.route('/api/fund-group-hint')
def api_fund_group_hint():
    name = request.args.get('name', '').strip()
    if not name:
        return jsonify({'fund_group_id': None})
    db = get_db()
    fund_group_id = _get_fund_group_hint(db, name)
    db.close()
    return jsonify({'fund_group_id': fund_group_id})


# ── 설정 및 동기화 페이지 ─────────────────────────────────────
@app.route('/settings')
def settings_page():
    return render_template('settings.html')


@app.route('/sync')
def sync_page():
    return redirect('/settings')


@app.route('/api/settings', methods=['GET', 'POST'])
def api_general_settings():
    db = get_db()
    if request.method == 'GET':
        cur = db.cursor()
        cur.execute("SELECT key, value FROM app_settings WHERE key IN ('settings_happiness', 'settings_goals', 'settings_vision_board', 'settings_popup_config')")
        rows = cur.fetchall()
        cur.close()
        db.close()
        
        # Defaults
        res = {
            'settings_happiness': [],
            'settings_goals': {},
            'settings_vision_board': '',
            'settings_popup_config': {'happiness_enabled': True, 'vision_enabled': True, 'goals_enabled': True}
        }
        for r in rows:
            key, val = r['key'], r['value']
            if val:
                try:
                    if key in ('settings_happiness', 'settings_goals', 'settings_popup_config'):
                        res[key] = json.loads(val)
                    else:
                        res[key] = val
                except Exception:
                    res[key] = val
        return jsonify(res)
    else:
        data = request.json
        cur = db.cursor()
        for key in ('settings_happiness', 'settings_goals', 'settings_vision_board', 'settings_popup_config'):
            if key in data:
                val = data[key]
                if key in ('settings_happiness', 'settings_goals', 'settings_popup_config'):
                    val_str = json.dumps(val)
                else:
                    val_str = val
                
                cur.execute("SELECT 1 FROM app_settings WHERE key=%s", (key,))
                if cur.fetchone():
                    cur.execute("UPDATE app_settings SET value=%s WHERE key=%s", (val_str, key))
                else:
                    cur.execute("INSERT INTO app_settings (key, value) VALUES (%s, %s)", (key, val_str))
        cur.close()
        db.commit()
        db.close()
        return jsonify({'ok': True})


SOURCE_FILES = [
    'app.py', 'database.py', 'requirements.txt',
    'templates/base.html', 'templates/dashboard.html',
    'templates/income.html', 'templates/budget.html',
    'templates/cards.html', 'templates/investments.html',
    'templates/realestate.html', 'templates/loans.html',
    'templates/pension.html', 'templates/goals.html',
    'templates/monthly.html', 'templates/settings.html',
    'templates/fund_management.html',
    'static/css/style.css', 'static/js/common.js', 'static/js/dashboard.js',
]

@app.route('/api/export-source')
def api_export_source():
    import base64
    files = {}
    base = os.path.dirname(os.path.abspath(__file__))
    for rel in SOURCE_FILES:
        path = os.path.join(base, rel.replace('/', os.sep))
        if os.path.exists(path):
            with open(path, 'rb') as f:
                files[rel] = base64.b64encode(f.read()).decode()
    return jsonify({'files': files})


@app.route('/api/compare-source', methods=['POST'])
def api_compare_source():
    import base64
    files = (request.json or {}).get('files', {})
    if not files:
        return jsonify({'error': '내용이 없습니다'}), 400
    base_dir = os.path.dirname(os.path.abspath(__file__))
    result = []
    for rel, b64 in files.items():
        dest = os.path.join(base_dir, rel.replace('/', os.sep))
        incoming = base64.b64decode(b64)
        if os.path.exists(dest):
            with open(dest, 'rb') as f:
                current = f.read()
            status = 'same' if incoming == current else 'changed'
        else:
            status = 'new'
        result.append({'file': rel, 'status': status, 'size': len(incoming)})
    return jsonify(result)


@app.route('/api/import-source', methods=['POST'])
def api_import_source():
    import base64
    files = (request.json or {}).get('files', {})
    if not files:
        return jsonify({'error': '내용이 없습니다'}), 400
    base = os.path.dirname(os.path.abspath(__file__))
    for rel, b64 in files.items():
        dest = os.path.join(base, rel.replace('/', os.sep))
        os.makedirs(os.path.dirname(dest), exist_ok=True)
        with open(dest, 'wb') as f:
            f.write(base64.b64decode(b64))
    return jsonify({'ok': True, 'count': len(files)})


@app.route('/api/export-text')
def api_export_text():
    return jsonify({'error': 'PostgreSQL 환경에서는 텍스트 백업 기능을 지원하지 않습니다.'}), 501


@app.route('/api/import-text', methods=['POST'])
def api_import_text():
    return jsonify({'error': 'PostgreSQL 환경에서는 텍스트 복구 기능을 지원하지 않습니다.'}), 501


# ── API: 전체 백업 / 복구 ────────────────────────────────────

# 외래키 의존성을 고려한 삭제 순서 (자식 → 부모)
_BACKUP_DELETE_ORDER = [
    'card_tx', 'card_mappings', 'card_category_rules',
    'stock_tx', 'etf_tx',
    'tenant_contracts', 'property_costs',
    'fund_group_rules', 'monthly_fund_budgets',
    'asset_snapshots', 'sold_real_estate',
    'income', 'budget', 'budget_recurring', 'crypto', 'loans', 'pension', 'goals',
    'cash_deposits', 'residence',
    'stocks', 'etf', 'real_estate', 'card_info',
    'fund_groups', 'categories', 'budget_categories', 'budget_category_rules',
    'stock_categories', 'app_settings',
]
# 삽입 순서 (부모 → 자식)
_BACKUP_INSERT_ORDER = [
    'categories', 'budget_categories', 'budget_category_rules',
    'stock_categories', 'card_info', 'fund_groups',
    'stocks', 'etf', 'real_estate',
    'budget_recurring',
    'income', 'budget', 'crypto', 'loans', 'pension', 'goals',
    'cash_deposits', 'residence', 'app_settings', 'asset_snapshots', 'sold_real_estate',
    'card_tx', 'card_mappings', 'card_category_rules',
    'stock_tx', 'etf_tx',
    'tenant_contracts', 'property_costs',
    'fund_group_rules', 'monthly_fund_budgets',
]


@app.route('/api/backup')
def api_backup():
    """전체 테이블 데이터를 JSON 파일로 다운로드"""
    import decimal, traceback

    def _safe(v):
        """psycopg2 반환값을 JSON 직렬화 가능한 타입으로 변환"""
        if v is None:
            return None
        if isinstance(v, (datetime, date)):
            return v.isoformat()
        if isinstance(v, decimal.Decimal):
            return float(v)
        if isinstance(v, memoryview):
            return v.tobytes().decode('utf-8', errors='replace')
        return v

    db = get_db()
    backup = {'version': '1.0', 'exported_at': datetime.now().isoformat(), 'tables': {}}
    try:
        for table in _BACKUP_INSERT_ORDER:
            cur = db.cursor()
            try:
                cur.execute(f"SELECT * FROM {table}")
                rows = cur.fetchall()
                backup['tables'][table] = [
                    {k: _safe(v) for k, v in dict(row).items()}
                    for row in rows
                ]
            except Exception:
                backup['tables'][table] = []
            finally:
                cur.close()
    except Exception as e:
        db.close()
        return jsonify({'error': traceback.format_exc()}), 500
    finally:
        try:
            db.close()
        except Exception:
            pass

    filename = f"money_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    return Response(
        json.dumps(backup, ensure_ascii=False, indent=2),
        mimetype='application/json',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'}
    )


@app.route('/api/restore', methods=['POST'])
def api_restore():
    """업로드된 JSON 백업 파일로 전체 데이터 복구"""
    import traceback
    if 'file' not in request.files:
        return jsonify({'error': '파일이 없습니다.'}), 400
    f = request.files['file']
    if not f.filename.endswith('.json'):
        return jsonify({'error': 'JSON 파일만 업로드할 수 있습니다.'}), 400

    try:
        backup = json.loads(f.read().decode('utf-8'))
    except Exception:
        return jsonify({'error': '파일 파싱 실패: 올바른 JSON 형식이 아닙니다.'}), 400

    if backup.get('version') != '1.0' or 'tables' not in backup:
        return jsonify({'error': '지원하지 않는 백업 형식입니다.'}), 400

    db = get_db()
    try:
        cur = db.cursor()
        # 자식 → 부모 순서로 삭제
        for table in _BACKUP_DELETE_ORDER:
            try:
                cur.execute(f"DELETE FROM {table}")
            except Exception:
                db.rollback()
        cur.close()

        tables = backup['tables']
        for table in _BACKUP_INSERT_ORDER:
            rows = tables.get(table, [])
            if not rows:
                continue
            cols = list(rows[0].keys())
            col_str = ', '.join(f'"{c}"' for c in cols)
            placeholders = ', '.join(['%s'] * len(cols))
            cur = db.cursor()
            for row in rows:
                vals = [row.get(c) for c in cols]
                cur.execute(
                    f"INSERT INTO {table} ({col_str}) VALUES ({placeholders}) ON CONFLICT DO NOTHING",
                    vals
                )
            cur.close()

            # 시퀀스 리셋 (id 컬럼이 있는 테이블)
            if rows and 'id' in rows[0]:
                cur = db.cursor()
                try:
                    cur.execute(
                        f"SELECT setval(pg_get_serial_sequence('{table}', 'id'), "
                        f"COALESCE((SELECT MAX(id) FROM {table}), 0) + 1, false)"
                    )
                except Exception:
                    pass
                cur.close()

        db.commit()
        total = sum(len(v) for v in tables.values())
        return jsonify({'ok': True, 'restored_rows': total})
    except Exception as e:
        db.rollback()
        return jsonify({'error': f'복구 실패: {traceback.format_exc()}'}), 500
    finally:
        db.close()


@app.route('/api/reset', methods=['POST'])
def api_reset():
    """모든 테이블의 데이터를 삭제 (초기화)"""
    import traceback
    db = get_db()
    try:
        cur = db.cursor()
        for table in _BACKUP_DELETE_ORDER:
            try:
                cur.execute(f"DELETE FROM {table}")
            except Exception:
                db.rollback()
        db.commit()
        cur.close()
        return jsonify({'ok': True})
    except Exception as e:
        db.rollback()
        return jsonify({'error': f'초기화 실패: {traceback.format_exc()}'}), 500
    finally:
        db.close()


@app.route('/api/bond-rate')
def api_bond_rate():
    """국민주택채권 할인율 자동 조회 (우리은행 경유)"""
    import re as _re
    from datetime import datetime as _dt
    now = _dt.now()
    year = str(now.year)
    month = f'{now.month:02d}'

    url = 'https://svc.wooribank.com/svc/Dream?withyou=HBNHB0087'
    hdrs = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
        'Content-Type': 'application/x-www-form-urlencoded',
        'Referer': url,
    }
    try:
        resp = http_req.post(url, headers=hdrs,
                             data={'MODE': '1', 'BSDT_YM': year + month,
                                   'STD_YEAR': year, 'STD_MONTH': month},
                             timeout=15)
        resp.raise_for_status()
        text = resp.text

        rows = _re.findall(
            r'<tr[^>]*class="tableline"[^>]*>(.*?)</tr>', text, _re.DOTALL)
        records = []
        for row in rows:
            tds = _re.findall(r'<td[^>]*>(.*?)</td>', row, _re.DOTALL)
            tds = [_re.sub(r'<[^>]+>', '', td).strip() for td in tds]
            if len(tds) >= 4:
                try:
                    records.append({
                        'date':          tds[0],
                        'sell_price':    tds[1],
                        'yield_rate':    float(tds[2]),
                        'discount_rate': round(float(tds[3]), 2),
                    })
                except ValueError:
                    pass

        if not records:
            return jsonify({'error': '데이터를 찾을 수 없습니다.'}), 404

        latest = records[-1]
        return jsonify({'ok': True, 'records': records, 'latest': latest})
    except Exception as e:
        return jsonify({'error': str(e)}), 500





# ── API: 상세 자산 목록 (계산기용) ──────────────────────────────
@app.route('/api/assets-detailed')
def api_assets_detailed():
    try:
        db = get_db()
        cur = db.cursor()
        ex_rate = get_current_exchange_rate()
        
        # 주식 (수량은 stock_tx 기반 계산)
        cur.execute("""
            SELECT s.name, s.ticker, s.current_price,
                COALESCE(SUM(CASE WHEN t.tx_type IN ('buy','매수') THEN t.quantity ELSE 0 END), 0)
              - COALESCE(SUM(CASE WHEN t.tx_type IN ('sell','매도') THEN t.quantity ELSE 0 END), 0) AS qty
            FROM stocks s
            LEFT JOIN stock_tx t ON t.stock_id = s.id
            GROUP BY s.id, s.name, s.ticker, s.current_price
            HAVING (COALESCE(SUM(CASE WHEN t.tx_type IN ('buy','매수') THEN t.quantity ELSE 0 END), 0)
                  - COALESCE(SUM(CASE WHEN t.tx_type IN ('sell','매도') THEN t.quantity ELSE 0 END), 0)) > 0
        """)
        stocks = []
        for r in cur.fetchall():
            val = float(r['current_price'] or 0) * float(r['qty'] or 0)
            ticker = str(r['ticker'] or '')
            is_foreign = ticker and not re.match(r'^[0-9]{6}$', ticker)
            if is_foreign: val *= ex_rate
            stocks.append({'name': r['name'] or '이름없음', 'val': round(val)})

        # ETF (수량은 etf_tx 기반 계산)
        cur.execute("""
            SELECT e.name, e.ticker, e.current_price,
                COALESCE(SUM(CASE WHEN t.tx_type IN ('buy','매수') THEN t.quantity ELSE 0 END), 0)
              - COALESCE(SUM(CASE WHEN t.tx_type IN ('sell','매도') THEN t.quantity ELSE 0 END), 0) AS qty
            FROM etf e
            LEFT JOIN etf_tx t ON t.etf_id = e.id
            GROUP BY e.id, e.name, e.ticker, e.current_price
            HAVING (COALESCE(SUM(CASE WHEN t.tx_type IN ('buy','매수') THEN t.quantity ELSE 0 END), 0)
                  - COALESCE(SUM(CASE WHEN t.tx_type IN ('sell','매도') THEN t.quantity ELSE 0 END), 0)) > 0
        """)
        etfs = []
        for r in cur.fetchall():
            val = float(r['current_price'] or 0) * float(r['qty'] or 0)
            ticker = str(r['ticker'] or '')
            is_foreign = ticker and not re.match(r'^[0-9]{6}$', ticker)
            if is_foreign: val *= ex_rate
            etfs.append({'name': r['name'] or '이름없음', 'val': round(val)})

        # 코인
        cur.execute("SELECT name, current_price * quantity as val FROM crypto WHERE quantity > 0")
        crypto = [{'name': r['name'] or '이름없음', 'val': round(float(r['val'] or 0))} for r in cur.fetchall()]

        # 현금/예금 + 거주보증금
        cur.execute("SELECT name, amount as val FROM cash_deposits WHERE amount > 0")
        cash = [{'name': r['name'] or '이름없음', 'val': round(float(r['val'] or 0))} for r in cur.fetchall()]
        
        cur.execute("SELECT address, deposit as val FROM residence WHERE deposit > 0")
        residence = [{'name': "[거주] " + (r['address'] or '보증금'), 'val': round(float(r['val'] or 0))} for r in cur.fetchall()]

        # 부동산 (세입자 보증금 차감하여 순가치 계산)
        cur.execute("""
            SELECT re.id, re.name, re.current_price,
                COALESCE(SUM(tc.deposit), 0) AS total_deposit
            FROM real_estate re
            LEFT JOIN tenant_contracts tc ON tc.real_estate_id = re.id
            GROUP BY re.id, re.name, re.current_price
        """)
        re_list = []
        for r in cur.fetchall():
            price   = round(float(r['current_price'] or 0))
            deposit = round(float(r['total_deposit'] or 0))
            re_list.append({
                'name':          r['name'] or '이름없음',
                'val':           price,
                'deposit':       deposit,
                'net_val':       price - deposit,
            })

        # 연금
        cur.execute("SELECT name, accumulated as val FROM pension WHERE accumulated > 0")
        pension = [{'name': r['name'] or '이름없음', 'val': round(float(r['val'] or 0))} for r in cur.fetchall()]

        # 세입자 보증금 (사적 레버리지 — 순자산 차감 항목)
        cur.execute("""
            SELECT re.name as re_name, tc.contract_type, tc.deposit
            FROM tenant_contracts tc
            JOIN real_estate re ON re.id = tc.real_estate_id
            WHERE tc.deposit > 0
        """)
        tenant_deposits = [
            {'name': f"[{r['contract_type']}] {r['re_name']}", 'val': round(float(r['deposit'] or 0))}
            for r in cur.fetchall()
        ]

        cur.close()
        db.close()

        return jsonify({
            '주식+ETF': stocks + etfs,  # 테크트리·대시보드 파이차트와 동일 기준
            '코인': crypto,
            '현금/예금': cash + residence,
            '부동산': re_list,
            '연금': pension,
            '_tenant_deposits': tenant_deposits,
        })
    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'trace': traceback.format_exc()}), 500

if __name__ == '__main__':
    init_db()
    app.run(debug=True, port=5000)
